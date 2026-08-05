"""
DFP TakeoffPro – Drawing Designer
(renamed from floor_plan_designer.py — same module, three tabs: Floor Plan,
Wiring Diagram, One-Line Diagram)

Scaled 2-D floor plan drawing: walls, rooms, dimensions, layers, and
electrical / low-voltage / furniture / structural symbol libraries.
Includes a freeform (not-to-scale) Wiring Diagram tab for schematic
device/circuit drawings, and a fire-alarm One-Line Diagram tab (panel /
booster / NAC-circuit / SLC-loop tree with live loading math). Exports to
PDF with automatic scale-to-fit-page (or manual scale override) against
standard paper sizes.

DEVELOPMENT NOTES — read this before touching rendering, drag/snap, or
PDF export. These are lessons paid for the hard way across a long session;
skipping them means re-discovering the same bugs.

1. HEADLESS TESTING — QT_QPA_PLATFORM=offscreen renders shapes fine but
   renders ZERO text. `QPainter.drawText()` / `QWidget.grab()` /
   `QGraphicsScene.render()` all silently produce blank space where text
   should be in this environment — confirmed all the way down to a bare
   `painter.drawText()` on a fresh QImage. This is NOT a bug in the app;
   it's a font-backend gap in the offscreen QPA plugin on this machine.
   Don't waste time chasing "why is my text invisible" in a screenshot
   taken this way — the shapes/layout are still trustworthy signal, text
   legibility is not. For anything text-related (dimension labels, tags,
   room info), verify via the PDF export instead: `fitz` inserts text via
   its own renderer (`page.insert_text` / `doc[i].get_pixmap()`), which
   reliably shows real glyphs. Test recipe: build a scene headlessly,
   call export_*_pdf(), then either `doc[i].get_text()` for a quick
   substring check or `get_pixmap(dpi=200+)` + crop-and-view for a real
   visual check.

2. PDF EXPORT — SHAPES MUST COMMIT BEFORE TEXT. Every export function
   here uses one shared `fitz.Shape` (`page.new_shape()`) that all
   `draw_line/draw_rect/draw_circle/draw_polyline(...).finish(...)` calls
   accumulate into; nothing actually hits the page's content stream until
   `shape.commit()` is called ONCE. Meanwhile `page.insert_text()` writes
   to the content stream IMMEDIATELY, in call order. So if you insert text
   in the middle of a shape-drawing loop (before that loop's `commit()`),
   the text physically ends up UNDER every shape in that same commit batch
   in the final PDF — even shapes that were "drawn" earlier in your Python
   call order — because the whole batch is one later content-stream chunk.
   Symptom: a symbol's tag text or a dimension label rendering invisible
   wherever it happens to overlap a wall or icon fill. Fixed pattern used
   throughout this file: do ALL shape drawing across ALL categories first,
   call `shape.commit()` once, THEN do a second pass inserting all text
   (collect what you need into a list of tuples during the shape pass if
   the text depends on per-item computed values). See
   export_floor_plan_pdf's `dim_labels`/`symbol_labels` lists for the
   pattern, and export_oneline_pdf's "Pass 1 shapes / Pass 2 text" split.

3. QGraphicsItem DRAG/SNAP GOTCHAS:
   - `itemChange()` is NEVER called for ItemPositionChange/
     ItemPositionHasChanged unless the item also has the
     `QGraphicsItem.ItemSendsGeometryChanges` flag set — this is off by
     default for performance and fails completely silently (no error, no
     warning, the hook just never fires). SymbolItem shipped for a while
     with drag-to-place snapping "working" (the initial click-to-place
     path snapped fine) while drag-to-reposition an existing symbol had
     ZERO snap logic wired up, purely because this flag was missing. If
     you add a new draggable/snappable item type, set this flag or your
     itemChange override is dead code.
   - Don't set a "this was manually moved" flag from inside itemChange —
     programmatic `setPos()` calls (auto-layout algorithms, load_dict
     restoring positions, etc.) fire itemChange exactly like a real user
     drag does, so a flag set there can't tell the difference and an
     auto-arrange pass will immediately re-mark everything as
     user-positioned, defeating its own reset. See OneLineNodeBase: the
     "manual_pos" flag is set from mousePressEvent/mouseReleaseEvent
     (comparing position at press vs. release — an actual drag happened)
     or explicitly by auto_arrange()/load_dict(), never from itemChange
     itself. itemChange is only used there for live connector-line
     following during a move, which has no such ambiguity problem.

4. ONE-LINE DIAGRAM layout model (floor_plan_designer's OneLineScene):
   - `relayout()` is cheap (just redraws connector lines from current
     positions) and safe to call after any edit. `auto_arrange()` is the
     expensive full tree-layout recompute (resets every node's manual_pos)
     — only call it for the explicit "Auto-Arrange" action or on initial
     load of geometry-less (legacy) save data. New nodes get a lightweight
     one-off placement near their parent (`_place_new_node`), not a full
     relayout, so adding one circuit doesn't reshuffle everything else.
   - CircuitNodes can be "standalone" (parent_node=None, tracked in
     `scene.standalone` rather than under `scene.panel`), and Panel/Booster
     nodes have a `.hidden` flag — both exist so a diagram can show just a
     NAC circuit with a free-text `source_label` ("Panel — NAC 3") instead
     of requiring the whole upstream panel/booster to be drawn. A node's
     `has_visible_parent()` (parent exists AND isn't hidden) gates whether
     the source_label prints — check both the live canvas paint() AND the
     PDF export text pass if you touch this, they're two separate code
     paths that must stay in sync (same for connector-line suppression in
     `update_connectors()` and `walk_lines()` in the PDF export).

5. RECT-ROOM RESIZE IS SESSION-ONLY. `RoomLabelItem._rect_walls` /
   `_rect_origin` / `_rect_size` (set by `FloorPlanScene.add_rect_room`)
   let "Resize Rectangle Room…" regenerate a room's 4 walls in place, but
   these are plain Python attributes, NOT persisted in to_dict()/
   from_dict(). After a save/load round trip the room and its walls are
   still there as ordinary items, just the "resize" convenience menu item
   silently stops appearing (checked via `hasattr(self, "_rect_walls")`).
   This was a deliberate scope cut, not an oversight — full persistence
   would need stable wall IDs to survive re-serialization.

6. SYMBOL SIZE OVERRIDES. `SymbolItem.size_override` is a per-instance
   (w_in, d_in) tuple that beats the catalog default from SYMBOL_DEFS —
   only set for furniture/stairs shapes (SymbolItem.SIZE_EDITABLE_SHAPES).
   Every place that reads a symbol's footprint for drawing/dimensioning
   must go through `effective_wd()` on-screen or build a shallow-copied
   `defn` dict with `w`/`d` overridden before calling
   `_draw_symbol_icon_pdf()` in the PDF export (see the `defn_eff` pattern
   in export_floor_plan_pdf) — reading `self.defn.get("w")` directly
   bypasses the override.

7. FQQ IMPORT SCOPE. `import_fqq_xlsm()` only reads an Autocall FQQ
   quote's conventional `NAC` sheet (non-addressable, current-draw-limited
   circuits). It does not parse the `IDNAC` sheet (addressable, device-
   count-limited SLC loops) — that sheet has a different data shape
   (no per-row mA lookup) and hasn't been built yet.
"""
import math, os, json, re, sys, traceback, datetime

try:
    from version import APP_VERSION
except ImportError:
    APP_VERSION = "1.0.0"

from PyQt5.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QPushButton, QLabel, QDoubleSpinBox, QFormLayout, QFrame,
    QMessageBox, QGraphicsView, QGraphicsScene, QGraphicsItem,
    QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView, QMenu,
    QApplication, QCheckBox, QLineEdit, QComboBox, QScrollArea,
    QTabWidget, QInputDialog, QFileDialog, QLayout, QSizePolicy,
    QSpinBox, QRadioButton, QButtonGroup, QListWidget, QListWidgetItem,
    QToolButton,
)
from PyQt5.QtGui import (
    QPainter, QColor, QPen, QBrush, QFont, QPolygonF, QPainterPath,
    QCursor, QPixmap, QIcon, QImage,
)
from PyQt5.QtCore import Qt, QRectF, QPointF, QSizeF, pyqtSignal, QRect, QPoint, QSize, QTimer
import fitz


# ═══════════════════════════════════════════════════════════════════════════════
#  Shared app-data helpers (mirrors suppression_designer.py / paint_booth_designer.py)
# ═══════════════════════════════════════════════════════════════════════════════

def _app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _projects_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.join(os.path.expanduser("~"), "Documents", "DFP TakeoffPro")
    else:
        base = _app_dir()
    p = os.path.join(base, "Projects")
    os.makedirs(p, exist_ok=True)
    return p


def _submittals_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.join(os.path.expanduser("~"), "Documents", "DFP TakeoffPro")
    else:
        base = _app_dir()
    p = os.path.join(base, "Submittals")
    os.makedirs(p, exist_ok=True)
    return p


def _log_error(context, exc=None):
    try:
        base = os.path.join(os.path.expanduser("~"), "Documents", "DFP TakeoffPro")
        os.makedirs(base, exist_ok=True)
        with open(os.path.join(base, "dfp_crash.log"), "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.datetime.now()}] {context}\n")
            f.write(traceback.format_exc())
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
#  FlowLayout / FlowBar — wrapping button toolbar (see suppression_designer.py)
# ═══════════════════════════════════════════════════════════════════════════════

class FlowLayout(QLayout):
    def __init__(self, parent=None, margin=0, hspacing=6, vspacing=6):
        super().__init__(parent)
        self._hspacing = hspacing
        self._vspacing = vspacing
        self._items = []
        if margin >= 0:
            self.setContentsMargins(margin, margin, margin, margin)

    def addItem(self, item):
        self._items.append(item)

    def addStretch(self, stretch=0):
        pass

    def horizontalSpacing(self):
        return self._hspacing

    def verticalSpacing(self):
        return self._vspacing

    def count(self):
        return len(self._items)

    def itemAt(self, index):
        return self._items[index] if 0 <= index < len(self._items) else None

    def takeAt(self, index):
        return self._items.pop(index) if 0 <= index < len(self._items) else None

    def expandingDirections(self):
        return Qt.Orientations(Qt.Orientation(0))

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self._do_layout(QRect(0, 0, width, 0), True)

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._do_layout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    def minimumSize(self):
        size = QSize()
        for item in self._items:
            size = size.expandedTo(item.minimumSize())
        m = self.contentsMargins()
        size += QSize(m.left()+m.right(), m.top()+m.bottom())
        return size

    def _do_layout(self, rect, test_only):
        left, top, right, bottom = self.getContentsMargins()
        effective = rect.adjusted(left, top, -right, -bottom)
        x, y = effective.x(), effective.y()
        line_height = 0
        for item in self._items:
            hint = item.sizeHint()
            next_x = x + hint.width() + self._hspacing
            if next_x - self._hspacing > effective.right() and line_height > 0:
                x = effective.x()
                y = y + line_height + self._vspacing
                next_x = x + hint.width() + self._hspacing
                line_height = 0
            if not test_only:
                item.setGeometry(QRect(QPoint(x, y), hint))
            x = next_x
            line_height = max(line_height, hint.height())
        return y + line_height - rect.y() + bottom


class FlowBar(QWidget):
    def __init__(self, parent=None, hspacing=6, vspacing=6, margin=0):
        super().__init__(parent)
        self._flow = FlowLayout(self, hspacing=hspacing, vspacing=vspacing)
        if margin >= 0:
            self._flow.setContentsMargins(margin, margin, margin, margin)
        sp = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)
        sp.setHeightForWidth(True)
        self.setSizePolicy(sp)

    def addWidget(self, w):
        self._flow.addWidget(w)

    def addStretch(self, stretch=0):
        pass

    def setContentsMargins(self, *a):
        self._flow.setContentsMargins(*a)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        h = self._flow.heightForWidth(self.width())
        if h > 0 and self.minimumHeight() != h:
            self.setMinimumHeight(h); self.setMaximumHeight(h)


TOOLBAR_MENU_STYLE = ("QMenu { background:#232728; color:#efe6e1; border:1px solid #555; }"
                      "QMenu::item { padding:7px 20px; }"
                      "QMenu::item:selected { background:#ff7002; color:white; }"
                      "QMenu::item:disabled { color:#777; }")


class HoverMenuButton(QToolButton):
    """Toolbar dropdown that opens its menu on hover as well as click — groups
    a cluster of related, mostly tab-specific tools (e.g. every Floor Plan
    drawing tool, or every Background Trace control) behind one button so
    they aren't all sitting in the bar at once regardless of which tab is
    active. Styled to match this toolbar's existing button look."""
    _instances = []

    def __init__(self, txt, color="#333738", parent=None):
        super().__init__(parent)
        self.setText(txt)
        self.setPopupMode(QToolButton.InstantPopup)
        self.setToolButtonStyle(Qt.ToolButtonTextOnly)
        self.setStyleSheet(
            f"QToolButton{{background:{color};color:#efe6e1;border-radius:3px;"
            f"padding:5px 10px;border:none;font-weight:bold;font-size:11px;}}"
            f"QToolButton:hover{{background:#ff7002;color:white;}}"
            f"QToolButton::menu-indicator{{width:0px;}}")
        HoverMenuButton._instances.append(self)

    def _close_sibling_menus(self):
        # Clicking or hovering onto one dropdown while a sibling's menu is
        # still open let Qt forward that same click into the new popup as
        # part of its native "click outside closes popup" handling, which
        # nested our showMenu() inside the still-unwinding first popup's
        # call stack and crashed. Closing any other open menu of ours
        # synchronously, before Qt gets a chance to do that forwarding,
        # keeps only one of our popups active at a time.
        for other in HoverMenuButton._instances:
            if other is not self:
                m = other.menu()
                if m is not None and m.isVisible():
                    m.hide()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            # Do NOT call super().mousePressEvent() here — with InstantPopup
            # that calls showMenu() synchronously, and when a sibling's menu
            # is already open, Qt can redeliver this same click into this
            # button while still unwinding the sibling's own popup call
            # stack (the "click one dropdown to switch straight to another"
            # behavior). Opening a second menu synchronously in that window
            # nests one popup's event loop inside the other's and crashes.
            # Deferring to the next event-loop tick lets the sibling's
            # popup finish closing first.
            event.accept()
            self._close_sibling_menus()
            QTimer.singleShot(0, self._open_menu_now)
            return
        super().mousePressEvent(event)

    def _open_menu_now(self):
        menu = self.menu()
        if menu is not None and not menu.isVisible():
            self.showMenu()

    def enterEvent(self, event):
        super().enterEvent(event)
        # Defer to the next event-loop tick instead of calling showMenu()
        # synchronously here — enterEvent can fire while another popup is
        # still tearing down its own mouse-grab, and calling showMenu()
        # inline in that window is what was crashing the app on dismiss.
        QTimer.singleShot(0, self._maybe_show_menu)

    def _maybe_show_menu(self):
        menu = self.menu()
        if menu is not None and not menu.isVisible() and self.underMouse():
            self._close_sibling_menus()
            self.showMenu()


# ═══════════════════════════════════════════════════════════════════════════════
#  Units — internal storage is always feet (float); thickness/widths in inches
# ═══════════════════════════════════════════════════════════════════════════════

PX_PER_FT = 12.0     # on-screen editing pixels per foot at 100% view zoom
GRID_FT   = 1.0       # minor grid spacing (ft)
GRID_MAJOR_EVERY = 10  # major grid line every N minor lines
SNAP_TOL_FT = 0.5     # generic snap tolerance (endpoints, close-loop, wall-proximity)

_LEN_RE = re.compile(
    r"""^\s*
        (?:(?P<ft>\d+(?:\.\d+)?)\s*'?)?\s*
        (?:(?P<in>\d+(?:\.\d+)?)\s*"?)?
        \s*$""", re.VERBOSE)


def parse_length(text):
    """Parse a typed length like 12, 12.5, 12', 12'6, 12'6", 6" -> feet (float).
    Returns None if unparseable. Accepts a bare number as feet."""
    text = (text or "").strip()
    if not text:
        return None
    # bare decimal number (no ' or ") => feet
    if re.fullmatch(r"\d+(\.\d+)?", text):
        return float(text)
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*'\s*(?:(\d+(?:\.\d+)?)\s*\"?)?", text)
    if m:
        ft = float(m.group(1))
        inch = float(m.group(2)) if m.group(2) else 0.0
        return ft + inch/12.0
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*\"", text)
    if m:
        return float(m.group(1))/12.0
    return None


def fmt_feet(ft, precision=1):
    """Format feet as architectural feet-inches, e.g. 12.5 -> 12'-6\"."""
    neg = ft < 0
    ft = abs(ft)
    whole_ft = int(ft)
    inch = round((ft - whole_ft) * 12.0, precision)
    if inch >= 12.0:
        whole_ft += 1; inch -= 12.0
    inch_s = f"{inch:.0f}" if abs(inch - round(inch)) < 0.05 else f"{inch:.1f}"
    s = f"{whole_ft}'-{inch_s}\""
    return ("-" if neg else "") + s


def snap_angle(dx, dy, step_deg=45.0):
    """Snap a direction vector to the nearest multiple of step_deg, keep length."""
    length = math.hypot(dx, dy)
    if length == 0:
        return dx, dy
    ang = math.atan2(dy, dx)
    step = math.radians(step_deg)
    snapped = round(ang / step) * step
    return length*math.cos(snapped), length*math.sin(snapped)


def snap_grid(v, grid=GRID_FT):
    return round(v/grid)*grid


def shoelace_area(points):
    """points: list of (x,y) in feet. Returns positive area in sq ft."""
    n = len(points)
    if n < 3:
        return 0.0
    a = 0.0
    for i in range(n):
        x1, y1 = points[i]
        x2, y2 = points[(i+1) % n]
        a += x1*y2 - x2*y1
    return abs(a) / 2.0


def polygon_centroid(points):
    n = len(points)
    if n == 0:
        return (0.0, 0.0)
    cx = sum(p[0] for p in points)/n
    cy = sum(p[1] for p in points)/n
    return (cx, cy)


def hatch_lines_in_polygon(points, angle_deg, spacing):
    """Generic scanline hatch-fill: returns [((x1,y1),(x2,y2)), ...] segments
    that fill `points` (a closed simple polygon, any coordinate space) with
    parallel lines at `angle_deg` spaced `spacing` apart. Used by the PDF
    Fire Zone export to draw the same hatch textures Qt's pattern brushes
    give the on-screen view for free — fitz has no built-in pattern brush,
    so this rotates the polygon so the hatch is horizontal, finds each
    scanline's edge crossings (even-odd rule), and rotates the resulting
    segments back."""
    if len(points) < 3 or spacing <= 0:
        return []
    ang = math.radians(angle_deg)
    ca, sa = math.cos(ang), math.sin(ang)
    # Rotate polygon into hatch-aligned space (hatch lines become horizontal).
    rpts = [(x*ca + y*sa, -x*sa + y*ca) for x, y in points]
    ys = [p[1] for p in rpts]
    y0, y1 = min(ys), max(ys)
    n = len(rpts)
    segments = []
    y = y0 + spacing/2
    while y < y1:
        xs = []
        for i in range(n):
            ax, ay = rpts[i]
            bx, by = rpts[(i+1) % n]
            if (ay - y) * (by - y) < 0:
                t = (y - ay) / (by - ay)
                xs.append(ax + t*(bx - ax))
        xs.sort()
        for i in range(0, len(xs) - 1, 2):
            x1, x2 = xs[i], xs[i+1]
            # Rotate the two segment endpoints back to the original space.
            p1 = (x1*ca - y*sa, x1*sa + y*ca)
            p2 = (x2*ca - y*sa, x2*sa + y*ca)
            segments.append((p1, p2))
        y += spacing
    return segments


# Angle(s) in degrees + line spacing (in PDF points, pre symbol/zone scale)
# for each Fire Zone hatch pattern — mirrors ZONE_QT_PATTERNS' look.
ZONE_HATCH_SPECS = {
    "solid": [],
    "diag1": [(45, 9)],
    "diag2": [(135, 9)],
    "cross_diag": [(45, 10), (135, 10)],
    "horizontal": [(0, 8)],
    "vertical": [(90, 8)],
    "grid": [(0, 9), (90, 9)],
}


def point_in_any_zone(zone_items, x_ft, y_ft):
    """Whether (x_ft,y_ft) falls inside any of `zone_items` (each a
    FireZoneItem or FireZoneItem-like object with a `.points` list) —
    standard even-odd ray-casting test. Shared by both the on-screen paint
    (SymbolItem/RoomLabelItem/FireZoneItem give a symbol/text a clear white
    halo when inside a zone, so a hatch pattern doesn't run through it) and
    the Fire Safety Plan PDF export's equivalent halo logic."""
    for z in zone_items:
        pts = z.points
        n = len(pts)
        if n < 3:
            continue
        inside = False
        j = n - 1
        for i in range(n):
            xi, yi = pts[i]; xj, yj = pts[j]
            if (yi > y_ft) != (yj > y_ft) and \
                    x_ft < (xj-xi)*(y_ft-yi)/((yj-yi) or 1e-9) + xi:
                inside = not inside
            j = i
        if inside:
            return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
#  Layers
# ═══════════════════════════════════════════════════════════════════════════════

LAYER_WALLS   = "Walls"
LAYER_DIMS    = "Dimensions"
LAYER_ROOMS   = "Room Labels"
LAYER_OPENING = "Doors / Windows"
LAYER_ELEC    = "Electrical"
LAYER_LV      = "Low-Voltage / Data"
LAYER_FURN    = "Furniture"
LAYER_STRUCT  = "Structural"
LAYER_FIRE    = "Fire Safety Plan"
LAYER_NOTES   = "Notes"

LAYER_ORDER = [LAYER_WALLS, LAYER_DIMS, LAYER_ROOMS, LAYER_OPENING,
               LAYER_ELEC, LAYER_LV, LAYER_FURN, LAYER_STRUCT, LAYER_FIRE, LAYER_NOTES]

# Preset boundary colors offered for Fire Zone / separation-area polygons,
# cycled by default so successive zones on one plan are visually distinct.
ZONE_COLOR_CHOICES = ["#c0392b", "#2980b9", "#27ae60", "#e67e22",
                       "#8e44ad", "#16a085", "#2c3e50", "#d35400"]

# Fill hatch patterns offered for Fire Zones — same idea as a real Fire Alarm
# Zone Drawing (see reference sample): a zone that reuses a color on another
# floor still reads as distinct because the fill texture differs, not just
# the color. Cycled independently from color so the (color, pattern)
# combination repeats far less often than either alone.
ZONE_PATTERN_CHOICES = ["solid", "diag1", "diag2", "cross_diag", "horizontal", "vertical", "grid"]
ZONE_PATTERN_NAMES = {
    "solid": "Solid", "diag1": "Diagonal /", "diag2": "Diagonal \\",
    "cross_diag": "Crosshatch ×", "horizontal": "Horizontal", "vertical": "Vertical", "grid": "Grid",
}
# Qt hatch brush styles for on-screen rendering — PDF export draws the
# equivalent by hand (fitz has no built-in pattern brush).
ZONE_QT_PATTERNS = {
    "solid": Qt.SolidPattern, "diag1": Qt.FDiagPattern, "diag2": Qt.BDiagPattern,
    "cross_diag": Qt.DiagCrossPattern, "horizontal": Qt.HorPattern,
    "vertical": Qt.VerPattern, "grid": Qt.CrossPattern,
}

LAYER_COLORS = {
    LAYER_WALLS:   "#3a3d3e",
    LAYER_DIMS:    "#c0392b",
    LAYER_ROOMS:   "#7f8c8d",
    LAYER_OPENING: "#2c3e50",
    LAYER_ELEC:    "#e67e22",
    LAYER_LV:      "#2980b9",
    LAYER_FURN:    "#8e7355",
    LAYER_STRUCT:  "#546e7a",
    LAYER_FIRE:    "#c0392b",
    LAYER_NOTES:   "#27ae60",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  Wall thickness presets (inches)
# ═══════════════════════════════════════════════════════════════════════════════

WALL_THICKNESS_OPTIONS = [
    ("2x4 stud + 1/2\" drywall both sides (4.5\")", 4.5),
    ("2x6 stud + 1/2\" drywall both sides (6.5\")", 6.5),
    ("Interior partition (3.5\")", 3.5),
    ("Exterior / masonry (8\")", 8.0),
    ("Glass partition (1\")", 1.0),
]

CEILING_TYPES = [
    "T-Bar / Lay-in Tile", "Drywall", "Open to Structure", "Wood", "Other",
]
FLOOR_TYPES = [
    "Carpet Tile", "Broadloom Carpet", "VCT / Vinyl Tile", "Hardwood",
    "Ceramic / Porcelain Tile", "Polished / Sealed Concrete", "Other",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  Paper sizes & print scales
# ═══════════════════════════════════════════════════════════════════════════════

PAPER_SIZES_IN = {
    "Letter (8.5x11)":  (8.5, 11.0),
    "Legal (8.5x14)":   (8.5, 14.0),
    "Tabloid (11x17)":  (11.0, 17.0),
    "ANSI C (17x22)":   (17.0, 22.0),
    "ANSI D (22x34)":   (22.0, 34.0),
    "ARCH D (24x36)":   (24.0, 36.0),
    "ARCH E (36x48)":   (36.0, 48.0),
}

# (display label, feet represented by 1 inch of paper)
SCALE_OPTIONS = [
    ("1/16\" = 1'-0\"", 16.0),
    ("3/32\" = 1'-0\"", 32.0/3.0),
    ("1/8\" = 1'-0\"",  8.0),
    ("3/16\" = 1'-0\"", 16.0/3.0),
    ("1/4\" = 1'-0\"",  4.0),
    ("3/8\" = 1'-0\"",  8.0/3.0),
    ("1/2\" = 1'-0\"",  2.0),
    ("3/4\" = 1'-0\"",  4.0/3.0),
    ("1\" = 1'-0\"",    1.0),
    ("1 1/2\" = 1'-0\"",2.0/3.0),
    ("3\" = 1'-0\"",    1.0/3.0),
    ("1\" = 10'",       10.0),
    ("1\" = 20'",       20.0),
    ("1\" = 30'",       30.0),
    ("1\" = 40'",       40.0),
    ("1\" = 50'",       50.0),
]
SCALE_OPTIONS.sort(key=lambda t: t[1])   # ascending feet-per-inch = largest drawing first


def best_fit_scale(width_ft, height_ft, paper_w_in, paper_h_in, margin_in=0.5):
    """Return the largest standard scale (smallest ft-per-inch) that fits the
    content within the printable area of the given paper (already oriented)."""
    pw = max(0.1, paper_w_in - 2*margin_in)
    ph = max(0.1, paper_h_in - 2*margin_in)
    if width_ft <= 0 or height_ft <= 0:
        return SCALE_OPTIONS[0][1]
    needed = max(width_ft/pw, height_ft/ph)
    for _, fpi in SCALE_OPTIONS:
        if fpi >= needed - 1e-9:
            return fpi
    return SCALE_OPTIONS[-1][1]


def fits_at_scale(width_ft, height_ft, paper_w_in, paper_h_in, feet_per_inch, margin_in=0.5):
    pw = max(0.1, paper_w_in - 2*margin_in)
    ph = max(0.1, paper_h_in - 2*margin_in)
    return (width_ft/feet_per_inch) <= pw + 1e-6 and (height_ft/feet_per_inch) <= ph + 1e-6


# ═══════════════════════════════════════════════════════════════════════════════
#  Symbol library — electrical / low-voltage / furniture (icons drawn generically
#  by SymbolItem.paint() based on "shape" key)
# ═══════════════════════════════════════════════════════════════════════════════

SYMBOL_DEFS = {
    # ── Electrical ───────────────────────────────────────────────────────────
    "outlet_duplex":  {"name":"Duplex Outlet",   "layer":LAYER_ELEC, "shape":"outlet",  "tag":""},
    "outlet_gfci":    {"name":"GFCI Outlet",     "layer":LAYER_ELEC, "shape":"outlet",  "tag":"GFI"},
    "outlet_quad":    {"name":"Quad Outlet",     "layer":LAYER_ELEC, "shape":"outlet4", "tag":""},
    "outlet_floor":   {"name":"Floor Outlet",    "layer":LAYER_ELEC, "shape":"outlet_sq","tag":""},
    "outlet_240v":    {"name":"240V Outlet",     "layer":LAYER_ELEC, "shape":"outlet",  "tag":"240"},
    "switch_single":  {"name":"Switch",          "layer":LAYER_ELEC, "shape":"switch",  "tag":"S"},
    "switch_3way":    {"name":"3-Way Switch",    "layer":LAYER_ELEC, "shape":"switch",  "tag":"S3"},
    "dimmer":         {"name":"Dimmer",          "layer":LAYER_ELEC, "shape":"switch",  "tag":"SD"},
    "light_ceiling":  {"name":"Ceiling Light",   "layer":LAYER_ELEC, "shape":"light",   "tag":""},
    "light_recessed": {"name":"Recessed Light",  "layer":LAYER_ELEC, "shape":"light_sm","tag":""},
    "panel":          {"name":"Electrical Panel","layer":LAYER_ELEC, "shape":"panel",   "tag":"PANEL"},

    # ── Low-voltage / data ───────────────────────────────────────────────────
    "data_drop":      {"name":"Data Drop",       "layer":LAYER_LV, "shape":"hex", "tag":"D"},
    "data_phone":     {"name":"Data/Phone Combo","layer":LAYER_LV, "shape":"hex", "tag":"D/V"},
    "wap":            {"name":"Wireless AP",     "layer":LAYER_LV, "shape":"wap", "tag":""},
    "camera":         {"name":"Security Camera", "layer":LAYER_LV, "shape":"camera","tag":""},
    "card_reader":    {"name":"Card Reader",     "layer":LAYER_LV, "shape":"rrect","tag":"CR"},
    "speaker":        {"name":"Speaker",         "layer":LAYER_LV, "shape":"speaker","tag":"SPK"},
    "tv_mount":       {"name":"TV / Display",    "layer":LAYER_LV, "shape":"rrect","tag":"TV"},

    # ── Furniture ─────────────────────────────────────────────────────────────
    "desk":           {"name":"Desk",            "layer":LAYER_FURN, "shape":"rect", "w":60,"d":30,"tag":"DSK"},
    "chair":          {"name":"Chair",           "layer":LAYER_FURN, "shape":"circle","w":20,"d":20,"tag":""},
    "table_round":    {"name":"Round Table",     "layer":LAYER_FURN, "shape":"circle","w":42,"d":42,"tag":""},
    "table_rect":     {"name":"Conference Table","layer":LAYER_FURN, "shape":"rect", "w":96,"d":42,"tag":""},
    "filing_cabinet": {"name":"Filing Cabinet",  "layer":LAYER_FURN, "shape":"cabinet","w":18,"d":24,"tag":""},

    # ── Structural ────────────────────────────────────────────────────────────
    "stairs_up":      {"name":"Stairs (up)",     "layer":LAYER_STRUCT, "shape":"stairs","w":36,"d":132,"tag":"UP","dir":"up"},
    "stairs_down":    {"name":"Stairs (down)",   "layer":LAYER_STRUCT, "shape":"stairs","w":36,"d":132,"tag":"DN","dir":"down"},

    # ── Fire Safety Plan — official Calgary Fire Department symbol set for
    # fire safety plan drawings (National Fire Code of Canada, Alberta
    # Edition, requires the AHJ's own approved symbols; only symbols
    # actually used need to appear in the legend — matches how the Symbol
    # Legend export already works). "You Are Here" isn't part of Calgary's
    # own set but is near-universal on evacuation diagrams, so it's added.
    "fsp_exit":        {"name":"Fire Exit",                  "layer":LAYER_FIRE, "shape":"arrow",       "color":"#27ae60", "tag":"EXIT"},
    "fsp_pull":        {"name":"Fire Pull Station",          "layer":LAYER_FIRE, "shape":"pull_hand",   "color":"#c0392b", "tag":"PS"},
    "fsp_extinguisher":{"name":"Fire Extinguisher",          "layer":LAYER_FIRE, "shape":"cylinder",    "color":"#c0392b", "tag":"FE"},
    "fsp_gas_shutoff": {"name":"Gas Meter Shut-off",         "layer":LAYER_FIRE, "shape":"gas_meter",   "color":"#27ae60", "tag":"G"},
    "fsp_hazard":      {"name":"Hazard",                     "layer":LAYER_FIRE, "shape":"diamond_text","color":"#c0392b", "tag":"HAZ"},
    "fsp_phone":       {"name":"Firefighter Telephone",      "layer":LAYER_FIRE, "shape":"phone",       "color":"#2980b9", "tag":"T"},
    "fsp_handicap":    {"name":"Handicap Symbol",            "layer":LAYER_FIRE, "shape":"rect_outline","color":"#2980b9", "tag":""},
    "fsp_faap":        {"name":"Fire Alarm Annunciator Panel","layer":LAYER_FIRE,"shape":"textbox",     "color":"#c0392b", "tag":"FAAP"},
    "fsp_facp":        {"name":"Fire Alarm Control Panel",   "layer":LAYER_FIRE, "shape":"textbox",     "color":"#c0392b", "tag":"FACP"},
    "fsp_cacf":        {"name":"Central Alarm Control Facility","layer":LAYER_FIRE,"shape":"textbox",   "color":"#c0392b", "tag":"CACF"},
    "fsp_fsp":         {"name":"Fire Safety Plan",           "layer":LAYER_FIRE, "shape":"textbox",     "color":"#c0392b", "tag":"FSP"},
    "fsp_standpipe":   {"name":"Standpipe Connection",       "layer":LAYER_FIRE, "shape":"standpipe",   "color":"#c0392b", "tag":"STDP"},
    "fsp_keybox":      {"name":"Fire Dept. Key Box",         "layer":LAYER_FIRE, "shape":"keybox",      "color":"#c0392b", "tag":"F.D.\nKEY"},
    "fsp_sprinkler_fv":{"name":"Sprinkler Floor Valve",      "layer":LAYER_FIRE, "shape":"gate_valve",  "color":"#c0392b", "tag":"SFV", "pedestal":True},
    "fsp_elec_vault":  {"name":"Electrical Vault",           "layer":LAYER_FIRE, "shape":"elec_vault",  "color":"#f1c40f", "tag":"E"},
    "fsp_sprinkler_cv":{"name":"Sprinkler Control Valves",   "layer":LAYER_FIRE, "shape":"gate_valve",  "color":"#c0392b", "tag":"SCV", "pedestal":False},
    "fsp_elec_hazard": {"name":"Electrical Hazard",          "layer":LAYER_FIRE, "shape":"bolt",        "color":"#f1c40f", "tag":"E"},
    "fsp_oxygen":      {"name":"Oxygen Valve",                "layer":LAYER_FIRE, "shape":"line_valve",  "color":"#27ae60", "tag":"O2"},
    "fsp_water_curtain":{"name":"Water Curtain",             "layer":LAYER_FIRE, "shape":"spray_head",  "color":"#2980b9", "tag":"WC"},
    "fsp_radiation":   {"name":"Radiation Hazard",           "layer":LAYER_FIRE, "shape":"trefoil",     "color":"#1a1a1a", "tag":"RAD"},
    "fsp_pull_special":{"name":"Pull Station for Special Suppression","layer":LAYER_FIRE,"shape":"flag","color":"#2980b9", "tag":"PSS"},
    "fsp_refuge":      {"name":"Area of Refuge",             "layer":LAYER_FIRE, "shape":"triangle_text","color":"#c0392b","tag":"R"},
    "fsp_gas_cylinder":{"name":"Pressurized Gas Cylinder",   "layer":LAYER_FIRE, "shape":"gas_cyl_diamond","color":"#27ae60", "tag":"PGC"},
    "fsp_generator":   {"name":"Generator",                  "layer":LAYER_FIRE, "shape":"gen_box",     "color":"#f1c40f", "tag":"Gen"},
    "fsp_air_horn":    {"name":"Air Horn",                   "layer":LAYER_FIRE, "shape":"horn",        "color":"#c0392b", "tag":"HORN"},
    "fsp_fire_wall":   {"name":"Fire Wall with Fire Door",   "layer":LAYER_FIRE, "shape":"wall_break",  "color":"#1a1a1a", "tag":"FWD"},
    "fsp_hose_cabinet":{"name":"Fire Hose in Cabinet",       "layer":LAYER_FIRE, "shape":"hose_reel",   "color":"#c0392b", "tag":"FHC", "cabinet":True},
    "fsp_fdc":         {"name":"Fire Dept. Connection",      "layer":LAYER_FIRE, "shape":"wye",         "color":"#c0392b", "tag":"FDC"},
    "fsp_hydrant_priv":{"name":"Fire Hydrant - Private",     "layer":LAYER_FIRE, "shape":"hydrant",     "color":"#c0392b", "tag":"PRIV","cap_color":None},
    "fsp_hydrant_pub": {"name":"Fire Hydrant - Public",      "layer":LAYER_FIRE, "shape":"hydrant",     "color":"#c0392b", "tag":"PUB", "cap_color":"#27ae60"},
    "fsp_fd_entry":    {"name":"Fire Dept. Entry",           "layer":LAYER_FIRE, "shape":"arrow",       "color":"#00b4e6", "tag":"ENTRY"},
    "fsp_water_shutoff":{"name":"Domestic Water Shut-off",   "layer":LAYER_FIRE, "shape":"line_valve",  "color":"#2980b9", "tag":"W"},
    "fsp_north":       {"name":"North Symbol",               "layer":LAYER_FIRE, "shape":"compass",     "color":"#1a1a1a", "tag":"N"},
    "fsp_elev_full":   {"name":"Firefighter Elevator - Full","layer":LAYER_FIRE, "shape":"helmet",      "color":"#c0392b", "tag":"F"},
    "fsp_rhss":        {"name":"Range Hood Suppression System","layer":LAYER_FIRE,"shape":"textbox",    "color":"#2980b9", "tag":"RHSS"},
    "fsp_elev_conv":   {"name":"Firefighter Elevator - Converted","layer":LAYER_FIRE,"shape":"helmet",  "color":"#c0392b", "tag":"C"},
    "fsp_hose_rack":   {"name":"Fire Hose on Rack",          "layer":LAYER_FIRE, "shape":"hose_reel",   "color":"#c0392b", "tag":"FHR", "cabinet":False},
    "fsp_first_aid":   {"name":"First Aid Station/Room",     "layer":LAYER_FIRE, "shape":"cross_box",   "color":"#c0392b", "tag":"FA"},
    "fsp_you_are_here":{"name":"You Are Here",               "layer":LAYER_FIRE, "shape":"you_are_here","color":"#c0392b", "tag":"YOU ARE\nHERE"},
}

DOOR_WINDOW_DEFS = {
    "door_swing_l": {"name":"Door (swing left)",  "layer":LAYER_OPENING, "shape":"door", "swing":"left",  "w_in":36},
    "door_swing_r": {"name":"Door (swing right)", "layer":LAYER_OPENING, "shape":"door", "swing":"right", "w_in":36},
    "door_double":  {"name":"Double Door",        "layer":LAYER_OPENING, "shape":"door_dbl","swing":"both","w_in":72},
    "window":       {"name":"Window",             "layer":LAYER_OPENING, "shape":"window","swing":None,   "w_in":36},
}

SYMBOL_CATEGORIES = [
    ("Electrical", [k for k,v in SYMBOL_DEFS.items() if v["layer"]==LAYER_ELEC]),
    ("Low-Voltage / Data", [k for k,v in SYMBOL_DEFS.items() if v["layer"]==LAYER_LV]),
    ("Furniture", [k for k,v in SYMBOL_DEFS.items() if v["layer"]==LAYER_FURN]),
    ("Structural", [k for k,v in SYMBOL_DEFS.items() if v["layer"]==LAYER_STRUCT]),
    ("Fire Safety Plan", [k for k,v in SYMBOL_DEFS.items() if v["layer"]==LAYER_FIRE]),
    ("Doors / Windows", list(DOOR_WINDOW_DEFS.keys())),
]


def _all_symbol_defs():
    d = dict(SYMBOL_DEFS); d.update(DOOR_WINDOW_DEFS); return d


ALL_SYMBOL_DEFS = _all_symbol_defs()


# ═══════════════════════════════════════════════════════════════════════════════
#  Fire alarm one-line diagram — device/circuit library
#  mA figures are ballpark 24V-regulated values for a rough loading estimate —
#  always verify against the actual device's manufacturer spec sheet.
# ═══════════════════════════════════════════════════════════════════════════════

FA_DEVICE_TYPES = {
    # SLC (addressable) — loop capacity is device-COUNT limited, not current
    "smoke_detector": {"name":"Smoke Detector",    "category":"slc", "ma":0.0, "abbr":"SD"},
    "heat_detector":  {"name":"Heat Detector",     "category":"slc", "ma":0.0, "abbr":"HD"},
    "duct_detector":  {"name":"Duct Detector",     "category":"slc", "ma":0.0, "abbr":"DD"},
    "beam_detector":  {"name":"Beam Detector",     "category":"slc", "ma":0.0, "abbr":"BD"},
    "pull_station":   {"name":"Pull Station",      "category":"slc", "ma":0.0, "abbr":"PS"},
    "monitor_module": {"name":"Monitor Module",    "category":"slc", "ma":0.0, "abbr":"MM"},
    "control_module": {"name":"Control Module",    "category":"slc", "ma":0.0, "abbr":"CM"},
    "iso_module":     {"name":"Isolator Module",   "category":"slc", "ma":0.0, "abbr":"ISO"},
    # NAC (notification) — current-draw limited
    "horn":            {"name":"Horn",                  "category":"nac", "ma":56,  "abbr":"H"},
    "strobe_15":       {"name":"Strobe (15cd)",          "category":"nac", "ma":42,  "abbr":"S15"},
    "strobe_30":       {"name":"Strobe (30cd)",          "category":"nac", "ma":56,  "abbr":"S30"},
    "strobe_75":       {"name":"Strobe (75cd)",          "category":"nac", "ma":94,  "abbr":"S75"},
    "strobe_110":      {"name":"Strobe (110cd)",         "category":"nac", "ma":124, "abbr":"S110"},
    "horn_strobe_15":  {"name":"Horn/Strobe (15cd)",     "category":"nac", "ma":101, "abbr":"HS15"},
    "horn_strobe_30":  {"name":"Horn/Strobe (30cd)",     "category":"nac", "ma":115, "abbr":"HS30"},
    "horn_strobe_75":  {"name":"Horn/Strobe (75cd)",     "category":"nac", "ma":153, "abbr":"HS75"},
    "horn_strobe_110": {"name":"Horn/Strobe (110cd)",    "category":"nac", "ma":183, "abbr":"HS110"},
    "speaker":         {"name":"Speaker",                "category":"nac", "ma":30,  "abbr":"SPK"},
    "speaker_strobe":  {"name":"Speaker/Strobe",         "category":"nac", "ma":88,  "abbr":"SPKS"},

    # Autocall TrueAlert conventional (non-addressable) — real nameplate mA
    # per manufacturer datasheets AC4906-0001 (VO), AC4906-0010 (weatherproof),
    # and the FQQ sizing tool's own current tables (matches datasheets exactly).
    "ac_9101_15":  {"name":"Autocall A4906-9101 VO Wall (15cd)",         "category":"nac", "ma":60,  "abbr":"9101-15",  "mfr":"Autocall", "part":"A4906-9101"},
    "ac_9101_30":  {"name":"Autocall A4906-9101 VO Wall (30cd)",         "category":"nac", "ma":94,  "abbr":"9101-30",  "mfr":"Autocall", "part":"A4906-9101"},
    "ac_9101_75":  {"name":"Autocall A4906-9101 VO Wall (75cd)",         "category":"nac", "ma":186, "abbr":"9101-75",  "mfr":"Autocall", "part":"A4906-9101"},
    "ac_9101_110": {"name":"Autocall A4906-9101 VO Wall (110cd)",        "category":"nac", "ma":252, "abbr":"9101-110", "mfr":"Autocall", "part":"A4906-9101"},
    "ac_9127_15":  {"name":"Autocall A4906-9127 A/V Wall (15cd)",        "category":"nac", "ma":75,  "abbr":"9127-15",  "mfr":"Autocall", "part":"A4906-9127"},
    "ac_9127_30":  {"name":"Autocall A4906-9127 A/V Wall (30cd)",        "category":"nac", "ma":116, "abbr":"9127-30",  "mfr":"Autocall", "part":"A4906-9127"},
    "ac_9127_75":  {"name":"Autocall A4906-9127 A/V Wall (75cd)",        "category":"nac", "ma":221, "abbr":"9127-75",  "mfr":"Autocall", "part":"A4906-9127"},
    "ac_9127_110": {"name":"Autocall A4906-9127 A/V Wall (110cd)",       "category":"nac", "ma":285, "abbr":"9127-110", "mfr":"Autocall", "part":"A4906-9127"},
    "ac_9139_135": {"name":"Autocall A4906-9139 Hi-cd A/V Wall (135cd)", "category":"nac", "ma":333, "abbr":"9139-135", "mfr":"Autocall", "part":"A4906-9139"},
    "ac_9139_177": {"name":"Autocall A4906-9139 Hi-cd A/V Wall (177cd)", "category":"nac", "ma":418, "abbr":"9139-177", "mfr":"Autocall", "part":"A4906-9139"},
    "ac_9139_185": {"name":"Autocall A4906-9139 Hi-cd A/V Wall (185cd)", "category":"nac", "ma":433, "abbr":"9139-185", "mfr":"Autocall", "part":"A4906-9139"},
    "ac_9140_135": {"name":"Autocall A4906-9140 Hi-cd A/V Ceiling (135cd)", "category":"nac", "ma":389, "abbr":"9140-135", "mfr":"Autocall", "part":"A4906-9140"},
    "ac_9140_177": {"name":"Autocall A4906-9140 Hi-cd A/V Ceiling (177cd)", "category":"nac", "ma":456, "abbr":"9140-177", "mfr":"Autocall", "part":"A4906-9140"},
    "ac_9140_185": {"name":"Autocall A4906-9140 Hi-cd A/V Ceiling (185cd)", "category":"nac", "ma":463, "abbr":"9140-185", "mfr":"Autocall", "part":"A4906-9140"},
    "ac_9131_15":  {"name":"Autocall A4906-9131 Weatherproof A/V (15cd)","category":"nac", "ma":91,  "abbr":"9131-15",  "mfr":"Autocall", "part":"A4906-9131"},
    "ac_9131_60":  {"name":"Autocall A4906-9131 Weatherproof A/V (60cd)","category":"nac", "ma":204, "abbr":"9131-60",  "mfr":"Autocall", "part":"A4906-9131"},
    "ac_9131_75":  {"name":"Autocall A4906-9131 Weatherproof A/V (75cd)","category":"nac", "ma":249, "abbr":"9131-75",  "mfr":"Autocall", "part":"A4906-9131"},
}
FA_SLC_DEVICE_KEYS = [k for k,v in FA_DEVICE_TYPES.items() if v["category"]=="slc"]
FA_NAC_DEVICE_KEYS = [k for k,v in FA_DEVICE_TYPES.items() if v["category"]=="nac"]

CIRCUIT_TYPE_INFO = {
    "slc": {"name":"SLC Loop",     "unit":"devices", "default_capacity":159},
    "nac": {"name":"NAC Circuit",  "unit":"mA",       "default_capacity":3000},
}
BOOSTER_CAPACITY_PRESETS_MA = [2500, 3000, 4000, 6000, 6500, 8000]

CIRCUIT_CLASS_INFO = {
    "A": {"label": "Class A", "desc": "Loops back to panel/booster — no EOL needed"},
    "B": {"label": "Class B", "desc": "Single run — ends in an EOL resistor"},
}


def expand_device_ticks(devices):
    """Expand a [{'key','qty'}] tally into one abbreviation string per
    individual device instance, in entry order."""
    ticks = []
    for d in devices:
        abbr = FA_DEVICE_TYPES.get(d["key"], {}).get("abbr", d["key"][:3].upper())
        ticks.extend([abbr] * d["qty"])
    return ticks


def circuit_grid_layout(devices, box_w, cell_w=28, cell_h=16, left_pad=8, right_pad=8):
    """Shared geometry (used by both on-screen paint and PDF export) for laying
    individual device ticks out in a wrapping grid. Returns (ticks, cols, rows)."""
    ticks = expand_device_ticks(devices)
    cols = max(1, int((box_w-left_pad-right_pad)/cell_w))
    rows = math.ceil(len(ticks)/cols) if ticks else 0
    return ticks, cols, rows


# ═══════════════════════════════════════════════════════════════════════════════
#  Graphics items
# ═══════════════════════════════════════════════════════════════════════════════

def _sync_wall_dimension(scene, old_key, wall):
    """Find the DimensionItem paired with `wall` (matched by its coordinates
    before this move/stretch, since dimensions store their own copy of the
    endpoints rather than a live reference) and move it to match."""
    if scene is None:
        return
    for d in scene._dim_items:
        if (d.x1, d.y1, d.x2, d.y2) == old_key:
            d.prepareGeometryChange()
            d.x1, d.y1, d.x2, d.y2 = wall.x1, wall.y1, wall.x2, wall.y2
            d.update()
            return


def _reposition_openings_after_stretch(wall, fixed_x, fixed_y):
    """Keep each door/window on `wall` the same DISTANCE from whichever
    corner of the wall didn't move, after the other corner got dragged —
    otherwise a stretched/shrunk wall would leave its doors floating in the
    wrong spot (or off the end entirely)."""
    wx, wy = wall.x2-wall.x1, wall.y2-wall.y1
    L = math.hypot(wx, wy) or 1.0
    ang = math.degrees(math.atan2(wy, wx))
    sign = 1.0 if (fixed_x, fixed_y) == (wall.x1, wall.y1) else -1.0
    ux, uy = (wx/L)*sign, (wy/L)*sign
    for o in list(wall.opening_items):
        ox_ft, oy_ft = o.pos().x()/PX_PER_FT, o.pos().y()/PX_PER_FT
        dist = max(0.0, min(math.hypot(ox_ft-fixed_x, oy_ft-fixed_y), L))
        o.setPos((fixed_x+ux*dist)*PX_PER_FT, (fixed_y+uy*dist)*PX_PER_FT)
        o.setRotation(ang)


class WallItem(QGraphicsItem):
    """A single straight wall segment, stored in feet, drawn at scene-origin
    (its geometry is baked into its own coordinates rather than using setPos,
    so it can share the "local coords == scene coords" convention every
    other non-setPos item here uses for its own custom-drag mouse handling)."""
    ITEM_TYPE = "wall"

    def __init__(self, x1, y1, x2, y2, thickness_in=4.5, layer=LAYER_WALLS):
        super().__init__()
        self.x1, self.y1, self.x2, self.y2 = x1, y1, x2, y2
        self.thickness_in = thickness_in
        self.layer = layer
        self.opening_items = []   # DoorWindowItem refs attached to this wall
        # +1/-1: which perpendicular side faces a room's interior, set when a
        # wall chain closes into a room — lets doors default to swinging inward.
        self.interior_side = 1
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setZValue(1)
        self._dragging = False
        self._drag_start_ft = None
        self._drag_orig = None   # (x1,y1,x2,y2) at the start of this drag
        self._drag_links = []    # connected walls captured once at drag start

    @property
    def length_ft(self):
        return math.hypot(self.x2-self.x1, self.y2-self.y1)

    def boundingRect(self):
        t = self.thickness_in/12.0 * PX_PER_FT
        x1, y1 = self.x1*PX_PER_FT, self.y1*PX_PER_FT
        x2, y2 = self.x2*PX_PER_FT, self.y2*PX_PER_FT
        pad = t/2 + 4
        return QRectF(min(x1,x2)-pad, min(y1,y2)-pad,
                      abs(x2-x1)+pad*2, abs(y2-y1)+pad*2)

    def shape(self):
        p = QPainterPath(); p.addRect(self.boundingRect()); return p

    def solid_segments_ft(self):
        """Return [(sx1,sy1,sx2,sy2), ...] for the solid parts of this wall,
        with a gap cut out under each attached door/window opening."""
        dx, dy = self.x2-self.x1, self.y2-self.y1
        L = math.hypot(dx, dy)
        if L < 1e-6:
            return [(self.x1,self.y1,self.x2,self.y2)]
        ux, uy = dx/L, dy/L
        intervals = []
        for o in self.opening_items:
            if o.scene() is None:
                continue
            ox, oy = o.pos().x()/PX_PER_FT, o.pos().y()/PX_PER_FT
            off = (ox-self.x1)*ux + (oy-self.y1)*uy
            w = o.width_in/12.0
            intervals.append((max(0.0, off-w/2), min(L, off+w/2)))
        if not intervals:
            return [(self.x1,self.y1,self.x2,self.y2)]
        intervals.sort()
        segs = []; cursor = 0.0
        for s, e in intervals:
            if s > cursor + 1e-6:
                segs.append((cursor, s))
            cursor = max(cursor, e)
        if cursor < L - 1e-6:
            segs.append((cursor, L))
        return [(self.x1+ux*s, self.y1+uy*s, self.x1+ux*e, self.y1+uy*e) for s, e in segs]

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        t = self.thickness_in/12.0 * PX_PER_FT
        sel = self.isSelected()
        col = QColor(LAYER_COLORS[LAYER_WALLS])
        painter.setBrush(QBrush(QColor("#ff7002") if sel else col))
        painter.setPen(QPen(col.darker(160), 1))
        for sx1, sy1, sx2, sy2 in self.solid_segments_ft():
            x1, y1 = sx1*PX_PER_FT, sy1*PX_PER_FT
            x2, y2 = sx2*PX_PER_FT, sy2*PX_PER_FT
            dx, dy = x2-x1, y2-y1
            length = math.hypot(dx, dy) or 1.0
            nx, ny = -dy/length*t/2, dx/length*t/2
            poly = QPolygonF([QPointF(x1+nx,y1+ny), QPointF(x2+nx,y2+ny),
                               QPointF(x2-nx,y2-ny), QPointF(x1-nx,y1-ny)])
            painter.drawPolygon(poly)

    def to_dict(self):
        return {"x1":self.x1,"y1":self.y1,"x2":self.x2,"y2":self.y2,
                "thickness_in":self.thickness_in,"layer":self.layer,
                "interior_side":self.interior_side}

    @classmethod
    def from_dict(cls, d):
        w = cls(d["x1"],d["y1"],d["x2"],d["y2"],
                 d.get("thickness_in",4.5), d.get("layer",LAYER_WALLS))
        w.interior_side = d.get("interior_side", 1)
        return w

    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_a = menu.addAction("Edit Length…")
        del_a = menu.addAction("Delete Wall")
        chosen = menu.exec_(event.screenPos())
        if chosen == edit_a:
            text, ok = QInputDialog.getText(None, "Wall Length",
                f"Length (currently {fmt_feet(self.length_ft)}):", text=fmt_feet(self.length_ft))
            if ok:
                new_len = parse_length(text)
                if new_len and new_len > 0:
                    dx, dy = self.x2-self.x1, self.y2-self.y1
                    L = math.hypot(dx,dy) or 1.0
                    ux, uy = dx/L, dy/L
                    self.x2, self.y2 = self.x1+ux*new_len, self.y1+uy*new_len
                    self.prepareGeometryChange(); self.update()
                    if self.scene(): self.scene().layout_changed.emit()
            return
        if chosen == del_a and self.scene():
            self.scene().remove_item(self)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._dragging = True
            self._drag_start_ft = (event.scenePos().x()/PX_PER_FT, event.scenePos().y()/PX_PER_FT)
            self._drag_orig = (self.x1, self.y1, self.x2, self.y2)
            # Find connected walls ONCE, at drag start, and hold direct
            # references — NOT re-derived by coordinate-matching on every
            # mouse-move frame. Re-matching every frame was the bug: after
            # frame 1 moves a connected wall's shared corner, it no longer
            # sits at THIS wall's *original* corner, so frame 2's re-match
            # would fail and silently drop the connection (walls "stretch a
            # little then disconnect").
            self._drag_links = []   # (wall, "1"/"2" = which end of wall, 1/2 = follows self's corner1/corner2)
            sc = self.scene()
            if sc is not None:
                tol = 0.05
                for w in sc._wall_items:
                    if w is self:
                        continue
                    if math.hypot(w.x1-self.x1, w.y1-self.y1) <= tol:
                        self._drag_links.append((w, "1", 1))
                    elif math.hypot(w.x2-self.x1, w.y2-self.y1) <= tol:
                        self._drag_links.append((w, "2", 1))
                    if math.hypot(w.x1-self.x2, w.y1-self.y2) <= tol:
                        self._drag_links.append((w, "1", 2))
                    elif math.hypot(w.x2-self.x2, w.y2-self.y2) <= tol:
                        self._drag_links.append((w, "2", 2))
            self.setSelected(True)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if not self._dragging:
            super().mouseMoveEvent(event)
            return
        sc = self.scene()
        cur_x_ft, cur_y_ft = event.scenePos().x()/PX_PER_FT, event.scenePos().y()/PX_PER_FT
        ox1, oy1, ox2, oy2 = self._drag_orig
        wdx, wdy = ox2-ox1, oy2-oy1
        L = math.hypot(wdx, wdy) or 1.0
        ux, uy = wdx/L, wdy/L
        nx, ny = -uy, ux   # unit perpendicular to the wall's own (unchanged) direction
        # Only the perpendicular component of the drag moves the wall — it
        # slides sideways parallel to itself, same as pushing a partition
        # wall over in a CAD tool, rather than free 2D dragging.
        push = (cur_x_ft-self._drag_start_ft[0])*nx + (cur_y_ft-self._drag_start_ft[1])*ny
        if sc is not None and getattr(sc, "snap_to_grid", False):
            push = snap_grid(push)
        new_x1, new_y1 = ox1+nx*push, oy1+ny*push
        new_x2, new_y2 = ox2+nx*push, oy2+ny*push

        if sc is not None:
            for w, endpoint, corner in self._drag_links:
                old_key = (w.x1, w.y1, w.x2, w.y2)
                target = (new_x1, new_y1) if corner == 1 else (new_x2, new_y2)
                w.prepareGeometryChange()
                if endpoint == "1":
                    w.x1, w.y1 = target
                    fixed_pt = (w.x2, w.y2)
                else:
                    w.x2, w.y2 = target
                    fixed_pt = (w.x1, w.y1)
                w.update()
                _sync_wall_dimension(sc, old_key, w)
                _reposition_openings_after_stretch(w, fixed_pt[0], fixed_pt[1])

            # This wall's own doors/windows ride along rigidly — the wall
            # itself only translates (its length/angle don't change).
            ddx, ddy = new_x1-ox1, new_y1-oy1
            for o in list(self.opening_items):
                o.setPos(o.pos().x()+ddx*PX_PER_FT, o.pos().y()+ddy*PX_PER_FT)

        old_self_key = (self.x1, self.y1, self.x2, self.y2)
        self.prepareGeometryChange()
        self.x1, self.y1, self.x2, self.y2 = new_x1, new_y1, new_x2, new_y2
        _sync_wall_dimension(sc, old_self_key, self)
        self.update()
        event.accept()

    def mouseReleaseEvent(self, event):
        if self._dragging:
            self._dragging = False
            self._drag_orig = None
            self._drag_links = []
            if self.scene(): self.scene().layout_changed.emit()
            event.accept()
            return
        super().mouseReleaseEvent(event)


class DimensionItem(QGraphicsItem):
    """Auto-generated dimension line offset from a wall, showing its length."""
    ITEM_TYPE = "dimension"

    def __init__(self, x1, y1, x2, y2, offset_ft=1.5, layer=LAYER_DIMS):
        super().__init__()
        self.x1, self.y1, self.x2, self.y2 = x1, y1, x2, y2
        self.offset_ft = offset_ft
        self.layer = layer
        self.setZValue(2)

    def boundingRect(self):
        pad = 30
        x1, y1 = self.x1*PX_PER_FT, self.y1*PX_PER_FT
        x2, y2 = self.x2*PX_PER_FT, self.y2*PX_PER_FT
        return QRectF(min(x1,x2)-pad, min(y1,y2)-pad,
                      abs(x2-x1)+pad*2, abs(y2-y1)+pad*2)

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        x1, y1 = self.x1, self.y1
        x2, y2 = self.x2, self.y2
        dx, dy = x2-x1, y2-y1
        length = math.hypot(dx, dy)
        if length < 0.01:
            return
        ux, uy = dx/length, dy/length
        nx, ny = -uy, ux   # unit normal
        off = self.offset_ft
        p1 = QPointF((x1+nx*off)*PX_PER_FT, (y1+ny*off)*PX_PER_FT)
        p2 = QPointF((x2+nx*off)*PX_PER_FT, (y2+ny*off)*PX_PER_FT)
        col = QColor(LAYER_COLORS[LAYER_DIMS])
        pen = QPen(col, 1)
        painter.setPen(pen)
        # extension lines
        painter.drawLine(QPointF(x1*PX_PER_FT,y1*PX_PER_FT), p1)
        painter.drawLine(QPointF(x2*PX_PER_FT,y2*PX_PER_FT), p2)
        # dimension line
        painter.drawLine(p1, p2)
        # tick marks
        for pt in (p1, p2):
            painter.drawLine(QPointF(pt.x()-3,pt.y()-3), QPointF(pt.x()+3,pt.y()+3))
        # label
        mid = QPointF((p1.x()+p2.x())/2, (p1.y()+p2.y())/2)
        ang = math.degrees(math.atan2(p2.y()-p1.y(), p2.x()-p1.x()))
        painter.save()
        painter.translate(mid)
        if 90 < abs(ang) <= 270:
            ang += 180
        painter.rotate(ang)
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        txt = fmt_feet(length)
        painter.setPen(col)
        painter.drawText(QRectF(-40,-16,80,14), Qt.AlignCenter, txt)
        painter.restore()

    def to_dict(self):
        return {"x1":self.x1,"y1":self.y1,"x2":self.x2,"y2":self.y2,
                "offset_ft":self.offset_ft,"layer":self.layer}

    @classmethod
    def from_dict(cls, d):
        return cls(d["x1"],d["y1"],d["x2"],d["y2"],d.get("offset_ft",1.5),d.get("layer",LAYER_DIMS))


class RoomLabelItem(QGraphicsItem):
    """Centered room name / area / ceiling height label, placed inside a closed
    wall loop (or anywhere the user drops it)."""
    ITEM_TYPE = "room"

    def __init__(self, x, y, name="Room", area_sqft=0.0, ceiling_in=108.0, layer=LAYER_ROOMS,
                 ceiling_type="T-Bar / Lay-in Tile", floor_type="Carpet Tile"):
        super().__init__()
        self.setPos(x*PX_PER_FT, y*PX_PER_FT)
        self.name = name
        self.area_sqft = area_sqft
        self.ceiling_in = ceiling_in   # inches (default 9'-0")
        self.ceiling_type = ceiling_type
        self.floor_type = floor_type
        self.layer = layer
        self.setFlag(QGraphicsItem.ItemIsMovable)
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        # Needed for itemChange(ItemPositionChange) to fire at all (off by
        # default) — that's what lets dragging a rectangle-room label also
        # drag its 4 walls along with it, instead of just the label text.
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges)
        self.setZValue(2)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionChange and hasattr(self, "_rect_walls"):
            sc = self.scene()
            new_pos = value
            if sc is not None and getattr(sc, "snap_to_grid", False):
                new_pos = QPointF(snap_grid(new_pos.x()/PX_PER_FT)*PX_PER_FT,
                                   snap_grid(new_pos.y()/PX_PER_FT)*PX_PER_FT)
            old_pos = self.pos()
            dx_ft = (new_pos.x()-old_pos.x())/PX_PER_FT
            dy_ft = (new_pos.y()-old_pos.y())/PX_PER_FT
            if dx_ft or dy_ft:
                # Drag the room's own 4 walls (and their paired auto-
                # dimensions, and any doors/windows attached to them) by the
                # same delta — otherwise only the text label would move,
                # leaving it detached from the walls it's naming.
                for w in self._rect_walls:
                    old_key = (w.x1, w.y1, w.x2, w.y2)
                    w.prepareGeometryChange()
                    w.x1 += dx_ft; w.y1 += dy_ft; w.x2 += dx_ft; w.y2 += dy_ft
                    for o in list(w.opening_items):
                        o.setPos(o.pos().x()+dx_ft*PX_PER_FT, o.pos().y()+dy_ft*PX_PER_FT)
                    if sc is not None:
                        for d in sc._dim_items:
                            if (d.x1, d.y1, d.x2, d.y2) == old_key:
                                d.prepareGeometryChange()
                                d.x1 += dx_ft; d.y1 += dy_ft; d.x2 += dx_ft; d.y2 += dy_ft
                                d.update()
                                break
                    w.update()
                ox, oy = self._rect_origin
                self._rect_origin = (ox+dx_ft, oy+dy_ft)
            return new_pos
        return super().itemChange(change, value)

    def boundingRect(self):
        return QRectF(-80, -22, 160, 58)

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        accent = QColor("#ff7002") if self.isSelected() else QColor(LAYER_COLORS[LAYER_ROOMS])
        text_col = QColor("#232323")   # dark, readable regardless of layer accent color
        sc = self.scene()
        if sc is not None and point_in_any_zone(getattr(sc, "_zone_items", []),
                                                 self.pos().x()/PX_PER_FT, self.pos().y()/PX_PER_FT):
            painter.setPen(Qt.NoPen)
            painter.setBrush(QBrush(QColor(255, 255, 255, 230)))
            painter.drawRoundedRect(QRectF(-82, -21, 164, 50), 4, 4)
        painter.setPen(accent if self.isSelected() else text_col)
        painter.setFont(QFont("Arial", 9, QFont.Bold))
        painter.drawText(QRectF(-80,-20,160,16), Qt.AlignCenter, self.name)
        painter.setFont(QFont("Arial", 7))
        painter.setPen(text_col)
        ceil_ft = self.ceiling_in/12.0
        painter.drawText(QRectF(-80,-2,160,14), Qt.AlignCenter,
                          f"{self.area_sqft:.0f} SF   ·   {fmt_feet(ceil_ft)} AFF")
        painter.setPen(QColor("#555555"))
        painter.setFont(QFont("Arial", 7))
        painter.drawText(QRectF(-80,12,160,14), Qt.AlignCenter,
                          f"Clg: {self.ceiling_type}   ·   Flr: {self.floor_type}")

    def to_dict(self):
        p = self.pos()
        return {"x":p.x()/PX_PER_FT,"y":p.y()/PX_PER_FT,"name":self.name,
                "area_sqft":self.area_sqft,"ceiling_in":self.ceiling_in,"layer":self.layer,
                "ceiling_type":self.ceiling_type,"floor_type":self.floor_type}

    @classmethod
    def from_dict(cls, d):
        return cls(d["x"],d["y"],d.get("name","Room"),d.get("area_sqft",0.0),
                    d.get("ceiling_in",108.0), d.get("layer",LAYER_ROOMS),
                    d.get("ceiling_type","T-Bar / Lay-in Tile"), d.get("floor_type","Carpet Tile"))

    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_a = menu.addAction("Edit Room…")
        resize_a = None
        copy_a = None
        if hasattr(self, "_rect_walls"):
            resize_a = menu.addAction("Resize Rectangle Room…")
            copy_a = menu.addAction("Copy Room")
        del_a = menu.addAction("Delete Room Label")
        chosen = menu.exec_(event.screenPos())
        if chosen == edit_a:
            dlg = RoomPropertiesDialog(self.name, self.area_sqft, self.ceiling_in,
                                        self.ceiling_type, self.floor_type)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                self.name = v["name"]; self.ceiling_in = v["ceiling_in"]
                self.ceiling_type = v["ceiling_type"]; self.floor_type = v["floor_type"]
                self.update()
                if self.scene(): self.scene().layout_changed.emit()
            return
        if chosen == resize_a:
            sc = self.scene()
            if not sc: return
            w_ft, l_ft = self._rect_size
            dlg = RectRoomDialog(w_ft, l_ft, self.name, self.ceiling_in,
                                  self.ceiling_type, self.floor_type)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                x0, y0 = self._rect_origin
                old_walls = list(self._rect_walls)
                sc.remove_item(self)
                for w in old_walls:
                    sc.remove_item(w)
                sc.add_rect_room(x0, y0, v["width_ft"], v["length_ft"], v["name"],
                                  v["ceiling_in"], v["ceiling_type"], v["floor_type"])
            return
        if copy_a is not None and chosen == copy_a:
            sc = self.scene()
            if sc:
                self.setSelected(True)
                sc.copy_selected()
            return
        if chosen == del_a and self.scene():
            self.scene().remove_item(self)


class FireZoneItem(QGraphicsItem):
    """A closed, colored, labeled fire-zone / separation-boundary polygon
    (e.g. for a Fire Alarm Zone Drawing) — geometry baked into instance
    coords like WallItem, since it's a multi-point shape rather than a
    single draggable point."""
    ITEM_TYPE = "zone"

    def __init__(self, points, name="Zone 1", color="#c0392b", layer=LAYER_FIRE, pattern="solid",
                 label_offset=(0.0, 0.0)):
        super().__init__()
        self.points = list(points)   # [(x_ft,y_ft), ...] — closed implicitly
        self.name = name
        self.color = color
        self.pattern = pattern
        self.layer = layer
        # Label position relative to the zone's centroid, in feet. (0,0) —
        # the default — means "centered in the zone". Drag the label (its
        # own hit area, separate from the zone polygon) to pull it clear
        # when the zone is too small/oddly shaped to write the name inside;
        # a leader line back to the zone follows automatically.
        self.label_offset_ft = tuple(label_offset)
        self._dragging_label = False
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setZValue(0.5)   # under walls/symbols so it reads as a background tint

    def centroid_ft(self):
        return polygon_centroid(self.points)

    def area_sqft(self):
        return shoelace_area(self.points)

    def label_pos_ft(self):
        cx, cy = self.centroid_ft()
        return (cx + self.label_offset_ft[0], cy + self.label_offset_ft[1])

    def _label_rect_px(self):
        lx, ly = self.label_pos_ft()
        return QRectF(lx*PX_PER_FT-82, ly*PX_PER_FT-17, 164, 26)

    def boundingRect(self):
        if not self.points:
            return QRectF(0, 0, 0, 0)
        xs = [p[0]*PX_PER_FT for p in self.points]
        ys = [p[1]*PX_PER_FT for p in self.points]
        pad = 20
        poly_rect = QRectF(min(xs)-pad, min(ys)-pad, max(xs)-min(xs)+pad*2, max(ys)-min(ys)+pad*2)
        return poly_rect.united(self._label_rect_px().adjusted(-4, -4, 4, 4))

    def shape(self):
        p = QPainterPath()
        p.addPolygon(QPolygonF([QPointF(x*PX_PER_FT, y*PX_PER_FT) for x, y in self.points]))
        p.addRect(self._label_rect_px())
        return p

    def paint(self, painter, option, widget=None):
        if len(self.points) < 3:
            return
        painter.setRenderHint(QPainter.Antialiasing)
        sel = self.isSelected()
        base = QColor("#ff7002") if sel else QColor(self.color)
        poly = QPolygonF([QPointF(x*PX_PER_FT, y*PX_PER_FT) for x, y in self.points])
        if self.pattern == "solid":
            fill = QColor(base); fill.setAlpha(45)
            painter.setBrush(QBrush(fill))
        else:
            # Qt hatch-pattern brushes draw the pattern in full accent color
            # over a transparent background (no separate alpha wash needed)
            # — same look as the reference sheet: white background, colored
            # hatch lines, so a repeated color still reads as a different
            # zone by texture alone.
            painter.setBrush(QBrush(base, ZONE_QT_PATTERNS.get(self.pattern, Qt.SolidPattern)))
        painter.setPen(QPen(base, 2, Qt.DashLine))
        painter.drawPolygon(poly)

        cx, cy = self.centroid_ft()
        cpx, cpy = cx*PX_PER_FT, cy*PX_PER_FT
        label_rect = self._label_rect_px()
        lcx, lcy = label_rect.center().x(), label_rect.center().y()
        moved = self.label_offset_ft != (0.0, 0.0)
        if moved:
            # Leader line + anchor dot back to the zone, same convention as
            # the reference Fire Zone Plan's pulled-out zone callouts.
            painter.setPen(QPen(base, 1.2))
            painter.drawLine(QPointF(cpx, cpy), QPointF(lcx, lcy))
            painter.setBrush(QBrush(base))
            painter.drawEllipse(QPointF(cpx, cpy), 3, 3)

        # Clear halo behind the zone name — always "inside itself" by
        # construction, so always halo it, same as any hatch pattern would
        # otherwise run right through the text.
        painter.setPen(Qt.NoPen)
        painter.setBrush(QBrush(QColor(255, 255, 255, 235)))
        painter.drawRoundedRect(label_rect, 4, 4)
        painter.setPen(base)
        painter.setFont(QFont("Arial", 10, QFont.Bold))
        painter.drawText(label_rect, Qt.AlignCenter, self.name)

    def to_dict(self):
        return {"points": self.points, "name": self.name, "color": self.color,
                "layer": self.layer, "pattern": self.pattern,
                "label_offset": list(self.label_offset_ft)}

    @classmethod
    def from_dict(cls, d):
        return cls(d.get("points", []), d.get("name", "Zone"),
                    d.get("color", "#c0392b"), d.get("layer", LAYER_FIRE),
                    d.get("pattern", "solid"), tuple(d.get("label_offset", (0.0, 0.0))))

    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_a = menu.addAction("Edit Zone…")
        reset_a = menu.addAction("Reset Label Position") if self.label_offset_ft != (0.0, 0.0) else None
        del_a = menu.addAction("Delete Zone")
        chosen = menu.exec_(event.screenPos())
        if chosen == edit_a:
            dlg = ZonePropertiesDialog(self.name, self.color, self.pattern)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                self.name = v["name"]; self.color = v["color"]; self.pattern = v["pattern"]
                self.prepareGeometryChange(); self.update()
                if self.scene(): self.scene().layout_changed.emit()
            return
        if reset_a is not None and chosen == reset_a:
            self.prepareGeometryChange()
            self.label_offset_ft = (0.0, 0.0)
            self.update()
            if self.scene(): self.scene().layout_changed.emit()
            return
        if chosen == del_a and self.scene():
            self.scene().remove_item(self)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._label_rect_px().contains(event.pos()):
            self._dragging_label = True
            self.setSelected(True)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging_label:
            cx, cy = self.centroid_ft()
            pos = event.pos()
            self.prepareGeometryChange()
            self.label_offset_ft = (pos.x()/PX_PER_FT - cx, pos.y()/PX_PER_FT - cy)
            self.update()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._dragging_label:
            self._dragging_label = False
            if self.scene(): self.scene().layout_changed.emit()
            event.accept()
            return
        super().mouseReleaseEvent(event)


class SymbolItem(QGraphicsItem):
    """Generic electrical / low-voltage / furniture symbol, rendered from
    SYMBOL_DEFS by 'shape' key."""
    ITEM_TYPE = "symbol"
    R = 12

    SIZE_EDITABLE_SHAPES = ("rect", "circle", "cabinet", "stairs")

    def __init__(self, kind, x=0.0, y=0.0, rotation=0.0, label="", size_override=None):
        super().__init__()
        self.kind = kind
        self.defn = ALL_SYMBOL_DEFS.get(kind, {"name":kind,"layer":LAYER_NOTES,"shape":"rect","tag":""})
        self.layer = self.defn["layer"]
        self.label = label
        # (w_in, d_in) override, or None to use the catalog default footprint
        # — lets a specific desk/table/stairs instance be resized without
        # changing every other symbol of that kind.
        self.size_override = tuple(size_override) if size_override else None
        self.setPos(x, y)
        self.setRotation(rotation)
        self.setFlag(QGraphicsItem.ItemIsMovable)
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        # Without this flag Qt never calls itemChange() for position changes
        # at all (it's off by default for performance) — meaning dragging an
        # already-placed symbol had NO snap logic wired up, even though
        # initial placement did. This is what makes dragging snap too.
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges)
        self.setZValue(3)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionChange:
            sc = self.scene()
            if sc is not None:
                return sc.snap_symbol_point(value)
        return super().itemChange(change, value)

    def effective_wd(self):
        """(w_in, d_in) actually used for drawing/dimensions — the instance
        override if the user resized this one, else the catalog default."""
        if self.size_override:
            return self.size_override
        return (self.defn.get("w", 24), self.defn.get("d", 24))

    def boundingRect(self):
        w_in, d_in = self.effective_wd()
        w = w_in/12.0*PX_PER_FT
        d = d_in/12.0*PX_PER_FT
        R = max(self.R, w/2, d/2) + 6
        shape = self.defn.get("shape")
        extra_bottom = 26 if shape in self.SIZE_EDITABLE_SHAPES else (
            18 if shape in TAG_BELOW_ICON_SHAPES else 0)
        extra_top = 14 if shape in TAG_ABOVE_ICON_SHAPES else 0
        return QRectF(-R, -R-14-extra_top, R*2, R*2+14+extra_top+extra_bottom)

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        shape = self.defn.get("shape","rect")
        base_color = self.defn.get("color") or LAYER_COLORS.get(self.layer,"#333")
        col = QColor("#ff7002") if self.isSelected() else QColor(base_color)
        R = self.R
        tag = self.defn.get("tag","")

        # Clear halo when placed inside a fire zone — a hatch pattern
        # otherwise runs right through the icon and fights with it for
        # attention.
        sc = self.scene()
        if sc is not None and point_in_any_zone(getattr(sc, "_zone_items", []),
                                                 self.pos().x()/PX_PER_FT, self.pos().y()/PX_PER_FT):
            painter.setPen(Qt.NoPen)
            painter.setBrush(QBrush(QColor(255, 255, 255, 235)))
            painter.drawEllipse(QPointF(0, 0), R*1.6, R*1.6)

        painter.setPen(QPen(col, 1.6)); painter.setBrush(QBrush(col.lighter(175)))

        if shape == "outlet":
            painter.drawEllipse(QPointF(0,0), R, R)
            painter.drawLine(QPointF(-4,-6), QPointF(-4,6))
            painter.drawLine(QPointF(4,-6), QPointF(4,6))
        elif shape == "outlet4":
            painter.drawEllipse(QPointF(0,0), R, R)
            for xo in (-5,-1.6,1.6,5):
                painter.drawLine(QPointF(xo,-6), QPointF(xo,6))
        elif shape == "outlet_sq":
            painter.drawRect(QRectF(-R,-R,R*2,R*2))
            painter.drawLine(QPointF(-4,-6), QPointF(-4,6))
            painter.drawLine(QPointF(4,-6), QPointF(4,6))
        elif shape == "switch":
            painter.drawEllipse(QPointF(0,0), R, R)
        elif shape == "light":
            painter.drawEllipse(QPointF(0,0), R, R)
            painter.drawLine(QPointF(-R,-R), QPointF(R,R))
            painter.drawLine(QPointF(-R,R), QPointF(R,-R))
        elif shape == "light_sm":
            painter.setBrush(QBrush(col))
            painter.drawEllipse(QPointF(0,0), R*0.6, R*0.6)
        elif shape == "panel":
            painter.drawRect(QRectF(-R,-R*0.7,R*2,R*1.4))
            painter.setPen(QPen(col.darker(140),1))
            for i in range(1,4):
                xx = -R + i*R*2/4
                painter.drawLine(QPointF(xx,-R*0.7), QPointF(xx,R*0.7))
        elif shape == "hex":
            poly = QPolygonF([QPointF(R*math.cos(math.radians(a)), R*math.sin(math.radians(a)))
                               for a in range(0,360,60)])
            painter.drawPolygon(poly)
        elif shape == "wap":
            painter.setBrush(Qt.NoBrush)
            painter.drawEllipse(QPointF(0,4), R*0.35, R*0.35)
            for rad in (R*0.6, R):
                painter.drawArc(QRectF(-rad,4-rad,rad*2,rad*2), 30*16, 120*16)
        elif shape == "camera":
            poly = QPolygonF([QPointF(-R,-R*0.5), QPointF(R*0.3,-R*0.5),
                               QPointF(R,-R), QPointF(R,R), QPointF(R*0.3,R*0.5),
                               QPointF(-R,R*0.5)])
            painter.drawPolygon(poly)
        elif shape == "rrect":
            painter.drawRoundedRect(QRectF(-R,-R*0.65,R*2,R*1.3), 3, 3)
        elif shape == "speaker":
            painter.drawEllipse(QPointF(0,0), R, R)
            painter.setBrush(QBrush(col))
            painter.drawEllipse(QPointF(0,0), R*0.3, R*0.3)
        elif shape == "circle":
            w_in, _ = self.effective_wd()
            w = w_in/12.0*PX_PER_FT
            painter.drawEllipse(QPointF(0,0), w/2, w/2)
            furn_half_h = w/2
        elif shape == "rect":
            w_in, d_in = self.effective_wd()
            w = w_in/12.0*PX_PER_FT; d = d_in/12.0*PX_PER_FT
            painter.drawRect(QRectF(-w/2,-d/2,w,d))
            furn_half_h = d/2
        elif shape == "cabinet":
            w_in, d_in = self.effective_wd()
            w = w_in/12.0*PX_PER_FT; d = d_in/12.0*PX_PER_FT
            painter.drawRect(QRectF(-w/2,-d/2,w,d))
            painter.drawLine(QPointF(-w/2,0), QPointF(w/2,0))
            furn_half_h = d/2
        elif shape == "stairs":
            w_in, d_in = self.effective_wd()
            w = w_in/12.0*PX_PER_FT; d = d_in/12.0*PX_PER_FT
            painter.drawRect(QRectF(-w/2,-d/2,w,d))
            n_treads = 10
            for i in range(1, n_treads):
                yy = -d/2 + i*d/n_treads
                painter.drawLine(QPointF(-w/2,yy), QPointF(w/2,yy))
            going_up = self.defn.get("dir","up") == "up"
            tip_y, tail_y = (-d/2+6, d/2-6) if going_up else (d/2-6, -d/2+6)
            painter.drawLine(QPointF(0,tail_y), QPointF(0,tip_y))
            ah = 7
            sign = -1 if going_up else 1
            painter.drawLine(QPointF(0,tip_y), QPointF(-ah/2,tip_y-sign*ah))
            painter.drawLine(QPointF(0,tip_y), QPointF(ah/2,tip_y-sign*ah))
            furn_half_h = d/2
        # ── Fire Safety Plan symbols (Calgary Fire Department set) ──────────
        elif shape == "arrow":
            pts = [QPointF(-R,0), QPointF(R*0.4,-R*0.8), QPointF(R*0.4,-R*0.3),
                   QPointF(R,-R*0.3), QPointF(R,R*0.3), QPointF(R*0.4,R*0.3), QPointF(R*0.4,R*0.8)]
            painter.drawPolygon(QPolygonF(pts))
        elif shape == "pull_hand":
            # Mounting box + a hooked hand/lever line — Calgary's Fire Pull
            # Station icon (a hand pulling a lever), simplified to a box with
            # a hook curve since the literal hand illustration doesn't read
            # at icon scale.
            painter.drawRect(QRectF(-R*0.9,-R,R*0.4,R*2))
            path = QPainterPath(); path.moveTo(-R*0.5,-R*0.3)
            for a in range(200, 341, 20):
                rad = math.radians(a)
                path.lineTo(-R*0.1+R*0.35*math.cos(rad), -R*0.1+R*0.35*math.sin(rad))
            path.lineTo(R*0.6, R*0.5)
            painter.setBrush(Qt.NoBrush)
            painter.strokePath(path, QPen(col, 1.4))
        elif shape == "flag":
            painter.drawRect(QRectF(-R*0.75,-R,R*0.3,R*2))
            pts = [QPointF(-R*0.45,-R*0.9), QPointF(R*0.65,-R*0.15),
                   QPointF(R*0.1,R*0.35), QPointF(-R*0.45,R*0.55)]
            painter.drawPolygon(QPolygonF(pts))
        elif shape == "cylinder":
            # Fire extinguisher: red body + a black nozzle/handle assembly on
            # top (a pull-pin handle bar plus a short hose hint), which is
            # the part that actually reads as "extinguisher" at icon scale —
            # a plain colored cap looked like a generic canister.
            painter.drawRoundedRect(QRectF(-R*0.5,-R*0.85,R,R*1.85), 3, 3)
            painter.setPen(QPen(QColor("#1a1a1a"), 1.4))
            painter.setBrush(QBrush(QColor("#1a1a1a")))
            painter.drawRect(QRectF(-R*0.4,-R*1.05,R*0.8,R*0.18))
            painter.drawLine(QPointF(-R*0.55,-R*0.87), QPointF(-R*0.85,-R*0.55))
            painter.setPen(QPen(col, 1.6))
        elif shape == "standpipe":
            painter.setBrush(Qt.NoBrush)
            painter.drawRoundedRect(QRectF(-R*0.85,-R,R*0.7,R*1.9), 2, 2)
            painter.drawRect(QRectF(-R*0.15,-R*0.35,R*0.55,R*0.5))
            painter.drawRect(QRectF(R*0.4,-R*0.25,R*0.35,R*0.3))
            painter.drawLine(QPointF(R*0.75,-R*0.5), QPointF(R*0.75,R*0.1))
        elif shape == "gate_valve":
            for xo in (-R*0.75, R*0.75):
                painter.drawRect(QRectF(xo-R*0.25,-R*0.6,R*0.5,R*1.2))
            painter.drawRect(QRectF(-R*0.75,-R*0.22,R*1.5,R*0.44))
            painter.setBrush(QBrush(col.lighter(175)))
            painter.drawEllipse(QPointF(0,0), R*0.3, R*0.3)
            if self.defn.get("pedestal"):
                painter.drawLine(QPointF(0,R*0.6), QPointF(0,R*0.95))
        elif shape == "elec_vault":
            painter.setBrush(QBrush(QColor("#f1c40f")))
            painter.setPen(QPen(QColor("#1a1a1a"), 1.4))
            painter.drawRoundedRect(QRectF(-R,-R,R*2,R*2), 3, 3)
            bolt = QPolygonF([QPointF(R*0.15,-R*0.9),QPointF(-R*0.3,R*0.05),QPointF(R*0.0,R*0.05),
                               QPointF(-R*0.15,R*0.9),QPointF(R*0.3,-R*0.05),QPointF(R*0.0,-R*0.05)])
            painter.setBrush(QBrush(QColor("#1a1a1a")))
            painter.drawPolygon(bolt)
            painter.setBrush(QBrush(QColor("#f1c40f")))
            painter.drawEllipse(QPointF(0,0), R*0.35, R*0.35)
        elif shape == "line_valve":
            painter.drawLine(QPointF(-R*1.3,0), QPointF(-R*0.9,0))
            painter.drawLine(QPointF(R*0.9,0), QPointF(R*1.3,0))
            painter.drawEllipse(QPointF(0,0), R*0.9, R*0.9)
        elif shape == "spray_head":
            painter.drawRect(QRectF(-R*0.25,-R,R*0.5,R*0.7))
            painter.setBrush(Qt.NoBrush)
            for ang in (-30,-10,10,30):
                rad = math.radians(ang)
                painter.drawLine(QPointF(0,-R*0.3), QPointF(R*0.9*math.sin(rad), R*0.8))
        elif shape in ("diamond_text","triangle_text","circle_text","textbox","rect_outline"):
            if shape == "diamond_text":
                painter.drawPolygon(QPolygonF([QPointF(0,-R),QPointF(R,0),QPointF(0,R),QPointF(-R,0)]))
            elif shape == "triangle_text":
                painter.drawPolygon(QPolygonF([QPointF(0,-R),QPointF(R,R*0.8),QPointF(-R,R*0.8)]))
            elif shape == "circle_text":
                painter.drawEllipse(QPointF(0,0), R, R)
            elif shape == "textbox":
                painter.drawRoundedRect(QRectF(-R*1.3,-R*0.8,R*2.6,R*1.6), 3, 3)
            else:   # rect_outline — plain outlined box (e.g. Handicap Symbol placeholder)
                painter.setBrush(Qt.NoBrush)
                painter.drawRect(QRectF(-R*1.1,-R*0.7,R*2.2,R*1.4))
        elif shape == "gas_cyl_diamond":
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF([QPointF(0,-R),QPointF(R,0),QPointF(0,R),QPointF(-R,0)]))
            painter.setBrush(QBrush(QColor("white")))
            painter.drawRoundedRect(QRectF(-R*0.16,-R*0.55,R*0.32,R*1.05), 2, 2)
        elif shape == "gen_box":
            painter.setBrush(QBrush(QColor("#f1c40f")))
            painter.setPen(QPen(QColor("#1a1a1a"), 1.4))
            painter.drawRect(QRectF(-R*1.3,-R*0.8,R*2.6,R*1.6))
        elif shape == "bolt":
            painter.drawPolygon(QPolygonF([QPointF(R*0.15,-R),QPointF(-R*0.4,R*0.1),QPointF(0,R*0.1),
                                            QPointF(-R*0.15,R),QPointF(R*0.4,-R*0.1),QPointF(0,-R*0.1)]))
            painter.drawEllipse(QPointF(0,0), R*0.32, R*0.32)
        elif shape == "trefoil":
            painter.setPen(QPen(QColor("#1a1a1a"), 1.0))
            painter.setBrush(QBrush(QColor("#1a1a1a")))
            painter.drawRect(QRectF(-R,-R,R*2,R*2))
            painter.setBrush(QBrush(QColor("white")))
            painter.drawEllipse(QPointF(0,0), R*0.95, R*0.95)
            for ang in (90, 210, 330):
                painter.save(); painter.rotate(ang)
                painter.setBrush(QBrush(QColor("#1a1a1a")))
                painter.drawPie(QRectF(-R*0.75,-R*0.75,R*1.5,R*1.5), 60*16, 60*16)
                painter.restore()
            painter.setBrush(QBrush(QColor("#1a1a1a")))
            painter.drawEllipse(QPointF(0,0), R*0.18, R*0.18)
        elif shape == "horn":
            painter.drawRect(QRectF(-R*0.9,-R*0.35,R*0.3,R*0.7))
            painter.setBrush(Qt.NoBrush)
            for rad in (R*0.35, R*0.7, R*1.05):
                path = QPainterPath()
                first = True
                for a in range(-50, 51, 10):
                    ar = math.radians(a)
                    x, y = -R*0.5+rad*math.cos(ar), rad*math.sin(ar)
                    (path.moveTo if first else path.lineTo)(x, y); first = False
                painter.strokePath(path, QPen(col, 1.1))
        elif shape == "wall_break":
            painter.drawLine(QPointF(-R,0), QPointF(-R*0.3,0))
            painter.drawLine(QPointF(R*0.3,0), QPointF(R,0))
            painter.drawArc(QRectF(-R*0.3,-R*0.3,R*0.6,R*0.6), 0, 180*16)
        elif shape == "hose_reel":
            if self.defn.get("cabinet"):
                painter.setBrush(Qt.NoBrush)
                painter.drawRoundedRect(QRectF(-R,-R*0.75,R*2,R*1.5), 2, 2)
            path = QPainterPath(); path.moveTo(-R*0.8,R*0.4)
            for i in range(1,6):
                yx = R*0.4 if i % 2 == 0 else -R*0.35
                path.lineTo(-R*0.8+i*R*0.28, yx)
            painter.strokePath(path, QPen(col, 1.6))
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF([QPointF(R*0.55,-R*0.5), QPointF(R*0.85,-R*0.15), QPointF(R*0.55,0)]))
        elif shape == "wye":
            painter.drawLine(QPointF(0,R), QPointF(0,0))
            painter.drawLine(QPointF(0,0), QPointF(-R*0.7,-R))
            painter.drawLine(QPointF(0,0), QPointF(R*0.7,-R))
        elif shape == "hydrant":
            cap_color = self.defn.get("cap_color")
            painter.setBrush(QBrush(col))
            painter.drawRoundedRect(QRectF(-R*0.4,-R*0.2,R*0.8,R*1.2), 2, 2)
            painter.drawLine(QPointF(-R*0.6,0), QPointF(-R*0.9,0))
            painter.drawLine(QPointF(R*0.6,0), QPointF(R*0.9,0))
            if cap_color:
                painter.setBrush(QBrush(QColor(cap_color)))
            painter.drawEllipse(QPointF(0,-R*0.5), R*0.35, R*0.3)
        elif shape == "compass":
            painter.setPen(QPen(col.lighter(160), 1.0))
            painter.setBrush(Qt.NoBrush)
            painter.drawEllipse(QPointF(0,R*0.1), R*0.75, R*0.75)
            painter.drawLine(QPointF(-R*0.75,R*0.1), QPointF(R*0.75,R*0.1))
            painter.setPen(QPen(col, 1.4))
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF([QPointF(0,-R*0.85),QPointF(R*0.22,R*0.35),
                                            QPointF(0,R*0.15),QPointF(-R*0.22,R*0.35)]))
            painter.drawLine(QPointF(0,R*0.35), QPointF(0,R*0.95))
        elif shape == "cross_box":
            painter.setBrush(QBrush(col))
            painter.drawRect(QRectF(-R*0.15,-R*0.6,R*0.3,R*1.2))
            painter.drawRect(QRectF(-R*0.6,-R*0.15,R*1.2,R*0.3))
        elif shape == "phone":
            pts = [QPointF(-R*0.7,-R*0.9), QPointF(-R*0.2,-R*0.5), QPointF(-R*0.35,-R*0.25),
                   QPointF(R*0.05,R*0.15), QPointF(R*0.35,-R*0.05), QPointF(R*0.75,R*0.35),
                   QPointF(R*0.45,R*0.75), QPointF(R*0.0,R*0.4), QPointF(-R*0.55,-R*0.15),
                   QPointF(-R*0.9,-R*0.55)]
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF(pts))
        elif shape == "gas_meter":
            painter.drawRect(QRectF(-R*0.8,-R*0.55,R*1.6,R*1.1))
            painter.drawLine(QPointF(-R*0.6,R*0.55), QPointF(-R*0.6,R))
            painter.drawLine(QPointF(R*0.6,R*0.55), QPointF(R*0.6,R))
            painter.drawLine(QPointF(-R*0.3,-R*0.55), QPointF(R*0.15,-R*1.15))
        elif shape == "keybox":
            painter.setPen(QPen(col.lighter(180), 0.8))
            painter.setBrush(Qt.NoBrush)
            painter.drawEllipse(QPointF(0,0), R, R)
        elif shape == "helmet":
            pts = [QPointF(-R,R*0.3), QPointF(-R*0.9,-R*0.1), QPointF(-R*0.5,-R*0.6), QPointF(0,-R*0.75),
                   QPointF(R*0.5,-R*0.5), QPointF(R*0.85,0), QPointF(R,R*0.3),
                   QPointF(R*0.6,R*0.45), QPointF(-R*0.6,R*0.45)]
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF(pts))
            painter.setPen(QPen(QColor("#555555"), 1.0))
            painter.setBrush(QBrush(QColor("#d5d5d5")))
            painter.drawEllipse(QPointF(0,-R*0.05), R*0.4, R*0.4)
        elif shape == "you_are_here":
            # Text drawn by the generic tag mechanism below (tag is set to
            # "YOU ARE\nHERE" in SYMBOL_DEFS) rather than baked in here, so
            # screen and PDF export render it identically instead of two
            # separately-maintained text placements.
            painter.setBrush(QBrush(col))
            painter.drawPolygon(QPolygonF([QPointF(0,-R*1.3),QPointF(R*0.9,R*0.2),QPointF(R*0.4,R*0.2),
                                            QPointF(R*0.4,R*1.1),QPointF(-R*0.4,R*1.1),
                                            QPointF(-R*0.4,R*0.2),QPointF(-R*0.9,R*0.2)]))
        else:
            painter.drawRect(QRectF(-R,-R,R*2,R*2))

        # Tag/label text always dark for readability — the layer color is
        # already carried by the symbol's outline/fill, so the text doesn't
        # need to repeat it (and white-on-light-tint was too low-contrast,
        # including for "panel" which used to force white text here).
        text_col = QColor("#ff7002") if self.isSelected() else QColor("#1a1a1a")
        if tag:
            painter.setFont(QFont("Arial", 6, QFont.Bold))
            painter.setPen(text_col)
            if shape in TAG_BELOW_ICON_SHAPES:
                painter.drawText(QRectF(-60, R+3, 120, 20), Qt.AlignHCenter | Qt.AlignTop, tag)
            elif shape in TAG_ABOVE_ICON_SHAPES:
                painter.drawText(QRectF(-60, -R-17, 120, 14), Qt.AlignHCenter | Qt.AlignBottom, tag)
            else:
                painter.drawText(QRectF(-R,-R,R*2,R*2), Qt.AlignCenter, tag)

        # Furniture (and stairs) footprint dimensions (W x D), printed below
        # the icon — so you know at a glance whether a desk/table will
        # actually fit where you dropped it, without opening a dialog.
        _dim_shapes = self.SIZE_EDITABLE_SHAPES
        if shape in _dim_shapes:
            w_in, d_in = self.effective_wd()
            dim_text = f'{fmt_feet(w_in/12.0)} x {fmt_feet(d_in/12.0)}'
            painter.setFont(QFont("Arial", 6))
            painter.setPen(QColor("#666666"))
            painter.drawText(QRectF(-50, furn_half_h+2, 100, 11), Qt.AlignCenter, dim_text)

        if self.label:
            label_y = (furn_half_h+13) if shape in _dim_shapes else (R+2)
            painter.setFont(QFont("Arial", 7, QFont.Bold))
            painter.setPen(text_col)
            painter.drawText(QRectF(-40,label_y,80,12), Qt.AlignCenter, self.label)

    def to_dict(self):
        d = {"kind":self.kind, "x":self.pos().x(), "y":self.pos().y(),
             "rotation":self.rotation(), "label":self.label}
        if self.size_override:
            d["size_override"] = list(self.size_override)
        return d

    @classmethod
    def from_dict(cls, d):
        so = d.get("size_override")
        return cls(d["kind"], d.get("x",0.0), d.get("y",0.0),
                    d.get("rotation",0.0), d.get("label",""),
                    size_override=tuple(so) if so else None)

    def contextMenuEvent(self, event):
        menu = QMenu()
        rot_l = menu.addAction("Rotate 90° ↺")
        rot_r = menu.addAction("Rotate 90° ↻")
        edit_a = menu.addAction("Edit Label…")
        size_a = None
        if self.defn.get("shape") in self.SIZE_EDITABLE_SHAPES:
            size_a = menu.addAction("Edit Size…")
        menu.addSeparator()
        del_a = menu.addAction("Delete")
        chosen = menu.exec_(event.screenPos())
        if chosen == rot_l:
            self.setRotation(self.rotation()-90); return
        if chosen == rot_r:
            self.setRotation(self.rotation()+90); return
        if chosen == edit_a:
            text, ok = QInputDialog.getText(None, "Symbol Label", "Label:", text=self.label)
            if ok:
                self.label = text.strip(); self.prepareGeometryChange(); self.update()
            return
        if chosen == size_a:
            w_in, d_in = self.effective_wd()
            dlg = SymbolSizeDialog(self.defn["name"], w_in, d_in)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                default_w, default_d = self.defn.get("w",24), self.defn.get("d",24)
                if abs(v["w_in"]-default_w) < 0.05 and abs(v["d_in"]-default_d) < 0.05:
                    self.size_override = None
                else:
                    self.size_override = (v["w_in"], v["d_in"])
                self.prepareGeometryChange(); self.update()
                if self.scene(): self.scene().layout_changed.emit()
            return
        if chosen == del_a and self.scene():
            self.scene().remove_item(self)


def arc_points(hinge, p_start, p_end, steps=14):
    """Points tracing the *shorter* arc from p_start to p_end, both assumed to
    be the same radius from hinge — used for door swing arcs so the sweep
    direction is always correct regardless of coordinate/rotation convention
    (avoids QPainter/PDF arc-angle-sign pitfalls entirely)."""
    hx, hy = hinge
    r = math.hypot(p_start[0]-hx, p_start[1]-hy) or 1e-6
    a0 = math.atan2(p_start[1]-hy, p_start[0]-hx)
    a1 = math.atan2(p_end[1]-hy, p_end[0]-hx)
    d = a1 - a0
    while d > math.pi: d -= 2*math.pi
    while d < -math.pi: d += 2*math.pi
    return [(hx+r*math.cos(a0+d*i/steps), hy+r*math.sin(a0+d*i/steps)) for i in range(steps+1)]


class DoorWindowItem(QGraphicsItem):
    """Door / window placed on a wall — cuts a real gap in the host wall
    (see WallItem.solid_segments_ft) and draws the swing/glazing symbol in
    that gap. host_wall is re-linked on load by re-snapping to the nearest
    wall at the saved position."""
    ITEM_TYPE = "opening"

    def __init__(self, kind, x=0.0, y=0.0, rotation=0.0, width_in=None, flip=False):
        super().__init__()
        self.kind = kind
        self.defn = DOOR_WINDOW_DEFS.get(kind, DOOR_WINDOW_DEFS["window"])
        self.layer = self.defn["layer"]
        self.width_in = width_in or self.defn.get("w_in", 36)
        self.flip = flip
        self.host_wall = None
        # Other walls drawn exactly on top of host_wall (e.g. the shared wall
        # between two independently-drawn adjoining rooms) that also need
        # their gap cut so the opening doesn't get visually covered by the
        # duplicate wall segment. See FloorPlanScene._coincident_walls.
        self.extra_host_walls = []
        self.setPos(x, y)
        self.setRotation(rotation)
        self.setFlag(QGraphicsItem.ItemIsMovable)
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setZValue(2.5)

    @property
    def w_px(self):
        return self.width_in/12.0*PX_PER_FT

    def boundingRect(self):
        w = self.w_px
        return QRectF(-w/2-4, -w-8, w+8, w+16)

    def _leaves(self, w):
        """List of (hinge, tip, jamb) tuples in local coords, one per door leaf.
        Defaults to swinging toward the host wall's room interior (set when the
        wall's chain closed into a room); "Flip Swing Side" inverts it."""
        base = self.host_wall.interior_side if self.host_wall is not None else 1
        sign = -base if self.flip else base
        shape = self.defn.get("shape")
        if shape == "door_dbl":
            leaf = w/2
            return [((-w/2,0), (-w/2, sign*-leaf), (0.0,0.0)),
                    (( w/2,0), ( w/2, sign*-leaf), (0.0,0.0))]
        swing = self.defn.get("swing","left")
        hinge = (-w/2,0) if swing=="left" else (w/2,0)
        jamb  = (w/2,0) if swing=="left" else (-w/2,0)
        return [(hinge, (hinge[0], sign*-w), jamb)]

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        col = QColor("#ff7002") if self.isSelected() else QColor(LAYER_COLORS[LAYER_OPENING])
        painter.setPen(QPen(col, 1.4)); painter.setBrush(Qt.NoBrush)
        w = self.w_px
        shape = self.defn.get("shape")
        if shape == "window":
            painter.drawLine(QPointF(-w/2,-2), QPointF(w/2,-2))
            painter.drawLine(QPointF(-w/2, 2), QPointF(w/2, 2))
            return
        for hinge, tip, jamb in self._leaves(w):
            painter.drawLine(QPointF(*hinge), QPointF(*tip))
            pts = arc_points(hinge, tip, jamb)
            painter.drawPolyline(QPolygonF([QPointF(x,y) for x,y in pts]))

    def to_dict(self):
        return {"kind":self.kind, "x":self.pos().x(), "y":self.pos().y(),
                "rotation":self.rotation(), "width_in":self.width_in, "flip":self.flip}

    @classmethod
    def from_dict(cls, d):
        return cls(d["kind"], d.get("x",0.0), d.get("y",0.0),
                    d.get("rotation",0.0), d.get("width_in"), d.get("flip", False))

    def contextMenuEvent(self, event):
        menu = QMenu()
        rot_a = menu.addAction("Rotate 90°")
        flip_a = None
        if self.defn.get("shape") in ("door", "door_dbl"):
            flip_a = menu.addAction("Flip Swing Side")
        del_a = menu.addAction("Delete")
        chosen = menu.exec_(event.screenPos())
        if chosen == rot_a:
            self.setRotation(self.rotation()+90); return
        if flip_a is not None and chosen == flip_a:
            self.flip = not self.flip; self.update(); return
        if chosen == del_a and self.scene():
            self.scene().remove_item(self)


# ═══════════════════════════════════════════════════════════════════════════════
#  Preview items — live feedback while drawing a wall chain
# ═══════════════════════════════════════════════════════════════════════════════

class _WallPreviewItem(QGraphicsItem):
    """Non-persistent rubber-band line + live length label shown while the
    wall tool is active (already-placed chain segments + the pending segment
    from the last placed point to the current cursor/typed position)."""
    def __init__(self):
        super().__init__()
        self.points = []       # committed chain points (feet)
        self.preview_pt = None # current pending endpoint (feet), or None
        self.label = ""
        self.setZValue(50)

    def boundingRect(self):
        return QRectF(-100000, -100000, 200000, 200000)

    def paint(self, painter, option, widget=None):
        if not self.points:
            return
        painter.setRenderHint(QPainter.Antialiasing)
        pen = QPen(QColor("#ff7002"), 2, Qt.DashLine)
        painter.setPen(pen)
        pts_px = [QPointF(x*PX_PER_FT, y*PX_PER_FT) for x, y in self.points]
        for i in range(len(pts_px)-1):
            painter.drawLine(pts_px[i], pts_px[i+1])
        for pt in pts_px:
            painter.setBrush(QBrush(QColor("#ff7002")))
            painter.drawEllipse(pt, 3, 3)
        if self.preview_pt is not None:
            last = pts_px[-1]
            prev_px = QPointF(self.preview_pt[0]*PX_PER_FT, self.preview_pt[1]*PX_PER_FT)
            painter.drawLine(last, prev_px)
            if self.label:
                painter.setFont(QFont("Arial", 9, QFont.Bold))
                painter.setPen(QColor("#232728"))
                mid = QPointF((last.x()+prev_px.x())/2, (last.y()+prev_px.y())/2-12)
                painter.drawText(QRectF(mid.x()-50, mid.y()-14, 100, 14), Qt.AlignCenter, self.label)


class _ZonePreviewItem(QGraphicsItem):
    """Non-persistent rubber-band outline shown while the Fire Zone tool is
    active — mirrors _WallPreviewItem but for a closed-polygon boundary
    (dashed committed edges + a dashed closing edge back to the first
    point, so the user can see the shape that will result before clicking)."""
    def __init__(self):
        super().__init__()
        self.points = []       # committed chain points (feet)
        self.preview_pt = None # current pending endpoint (feet), or None
        self.setZValue(50)

    def boundingRect(self):
        return QRectF(-100000, -100000, 200000, 200000)

    def paint(self, painter, option, widget=None):
        if not self.points:
            return
        painter.setRenderHint(QPainter.Antialiasing)
        pen = QPen(QColor("#c0392b"), 2, Qt.DashLine)
        painter.setPen(pen)
        pts_px = [QPointF(x*PX_PER_FT, y*PX_PER_FT) for x, y in self.points]
        for i in range(len(pts_px)-1):
            painter.drawLine(pts_px[i], pts_px[i+1])
        for pt in pts_px:
            painter.setBrush(QBrush(QColor("#c0392b")))
            painter.drawEllipse(pt, 3, 3)
        if self.preview_pt is not None:
            last = pts_px[-1]
            prev_px = QPointF(self.preview_pt[0]*PX_PER_FT, self.preview_pt[1]*PX_PER_FT)
            painter.drawLine(last, prev_px)
            if len(pts_px) >= 2:
                painter.drawLine(prev_px, pts_px[0])


class FloorPlanScene(QGraphicsScene):
    layout_changed = pyqtSignal()
    status_changed = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setSceneRect(-5000, -5000, 10000, 10000)
        self.setBackgroundBrush(QBrush(QColor("white")))
        self._wall_items = []
        self._dim_items = []
        self._room_items = []
        self._symbol_items = []
        self._opening_items = []
        self._zone_items = []
        self.layers_visible = {name: True for name in LAYER_ORDER}
        self.show_grid = True
        self.snap_to_grid = True
        self.default_wall_thickness_in = 4.5
        self._mode = "select"         # select | wall | symbol | zone
        self._pending_symbol = None   # kind string when placing a symbol
        self._chain = []              # committed points of current wall chain (feet)
        self._chain_walls = []        # WallItems created for the current chain
        self._length_buffer = ""
        self._preview = _WallPreviewItem()
        self.addItem(self._preview)
        self._zone_chain = []         # committed points of current fire-zone chain (feet)
        self._zone_preview = _ZonePreviewItem()
        self.addItem(self._zone_preview)
        # Background trace image (e.g. a scanned/exported PDF floor plan) —
        # a tracing aid only: never exported, never saved with the project
        # (a full raster embedded in the save file would bloat it a lot).
        # bg_pos_ft is the image's top-left corner in feet; bg_px_per_ft is
        # how many of the image's own pixels span one real-world foot.
        self.bg_pixmap = None
        self.bg_pos_ft = (0.0, 0.0)
        self.bg_px_per_ft = 10.0
        self.bg_dpi = 150
        self.bg_opacity = 0.55
        self.bg_visible = True
        self._bg_calib_pts = []

    # ── Layers ───────────────────────────────────────────────────────────────

    def set_layer_visible(self, layer, visible):
        self.layers_visible[layer] = visible
        for lst in (self._wall_items, self._dim_items, self._room_items,
                    self._symbol_items, self._opening_items, self._zone_items):
            for it in lst:
                if it.layer == layer:
                    it.setVisible(visible)
        self.update()

    def _apply_layer_visibility(self, item):
        item.setVisible(self.layers_visible.get(item.layer, True))

    # ── Modes ────────────────────────────────────────────────────────────────

    def set_mode_select(self):
        self._mode = "select"
        self._pending_symbol = None
        self._cancel_wall_chain()
        self._cancel_zone_chain()
        self._bg_calib_pts = []

    def set_mode_wall(self, thickness_in=None):
        if thickness_in is not None:
            self.default_wall_thickness_in = thickness_in
        self._cancel_zone_chain()
        self._mode = "wall"
        self._chain = []
        self._chain_walls = []
        self._length_buffer = ""
        self._preview.points = []
        self._preview.preview_pt = None
        self.status_changed.emit(
            "  Wall tool — click to start, click to place each corner, type a length + Enter to "
            "set it exactly, Esc = cancel, double-click / Enter (empty) = finish open wall.")
        self.update()

    def set_mode_symbol(self, kind):
        self._mode = "symbol"
        self._pending_symbol = kind
        self._cancel_wall_chain()
        self._cancel_zone_chain()
        name = ALL_SYMBOL_DEFS.get(kind, {}).get("name", kind)
        self.status_changed.emit(f"  Placing: {name} — click on the canvas (near a wall for doors/windows).")

    def set_mode_zone(self):
        self._cancel_wall_chain()
        self._mode = "zone"
        self._zone_chain = []
        self._zone_preview.points = []
        self._zone_preview.preview_pt = None
        self.status_changed.emit(
            "  Fire Zone tool — click to place each boundary corner, click near the first point "
            "(or double-click / Enter with an empty buffer) to close the zone, Esc = cancel.")
        self.update()

    def set_mode_bg_calibrate(self):
        self._cancel_wall_chain()
        self._cancel_zone_chain()
        self._mode = "bg_calibrate"
        self._bg_calib_pts = []
        self.status_changed.emit(
            "  Calibrate Background Scale — click one end of a distance you know in real life "
            "(e.g. a doorway), then the other end, and type that distance. Esc = cancel.")

    def _cancel_wall_chain(self):
        self._chain = []
        self._chain_walls = []
        self._length_buffer = ""
        self._preview.points = []
        self._preview.preview_pt = None
        self._preview.label = ""
        self.update()

    def _cancel_zone_chain(self):
        self._zone_chain = []
        self._zone_preview.points = []
        self._zone_preview.preview_pt = None
        self.update()

    # ── Grid background ──────────────────────────────────────────────────────

    def drawBackground(self, painter, rect):
        painter.fillRect(rect, QBrush(QColor("white")))
        if self.bg_pixmap is not None and self.bg_visible:
            painter.save()
            painter.setOpacity(self.bg_opacity)
            painter.translate(self.bg_pos_ft[0]*PX_PER_FT, self.bg_pos_ft[1]*PX_PER_FT)
            s = PX_PER_FT/self.bg_px_per_ft
            painter.scale(s, s)
            painter.drawPixmap(0, 0, self.bg_pixmap)
            painter.restore()
        if not self.show_grid:
            return
        step = GRID_FT * PX_PER_FT
        left = int(rect.left()/step)-1
        right = int(rect.right()/step)+1
        top = int(rect.top()/step)-1
        bottom = int(rect.bottom()/step)+1
        minor_pen = QPen(QColor("#eef0f1"), 1)
        major_pen = QPen(QColor("#d8dcde"), 1)
        for i in range(left, right+1):
            x = i*step
            painter.setPen(major_pen if i % GRID_MAJOR_EVERY == 0 else minor_pen)
            painter.drawLine(QPointF(x, rect.top()), QPointF(x, rect.bottom()))
        for j in range(top, bottom+1):
            y = j*step
            painter.setPen(major_pen if j % GRID_MAJOR_EVERY == 0 else minor_pen)
            painter.drawLine(QPointF(rect.left(), y), QPointF(rect.right(), y))
        # origin cross
        painter.setPen(QPen(QColor("#b8bcbe"), 1.4))
        painter.drawLine(QPointF(-20,0), QPointF(20,0))
        painter.drawLine(QPointF(0,-20), QPointF(0,20))

    # ── Wall building helpers ───────────────────────────────────────────────

    def _snap_point(self, x_ft, y_ft, from_pt=None):
        """Snap to grid, and if a from_pt is given, snap the direction to 45°
        — both gated on Snap being ON. With Snap OFF this returns the raw
        point untouched, for genuinely freehand (any angle, any position)
        wall drawing."""
        if not self.snap_to_grid:
            return x_ft, y_ft
        if from_pt is not None:
            dx, dy = x_ft-from_pt[0], y_ft-from_pt[1]
            dx, dy = snap_angle(dx, dy, 45.0)
            x_ft, y_ft = from_pt[0]+dx, from_pt[1]+dy
        x_ft, y_ft = snap_grid(x_ft), snap_grid(y_ft)
        return x_ft, y_ft

    def _add_wall(self, x1, y1, x2, y2, thickness_in=None, layer=LAYER_WALLS):
        w = WallItem(x1, y1, x2, y2, thickness_in or self.default_wall_thickness_in, layer)
        self.addItem(w)
        self._apply_layer_visibility(w)
        self._wall_items.append(w)
        if self._mode == "wall":
            self._chain_walls.append(w)
        length = w.length_ft
        if length > 0.05:
            d = DimensionItem(x1, y1, x2, y2, offset_ft=1.5, layer=LAYER_DIMS)
            self.addItem(d); self._apply_layer_visibility(d)
            self._dim_items.append(d)
        return w

    def _finish_chain(self, closed):
        pts = self._chain
        if closed and len(pts) >= 3:
            area = shoelace_area(pts)
            cx, cy = polygon_centroid(pts[:-1] if pts[0] == pts[-1] else pts)
            # Determine which perpendicular side of each new wall faces the room
            # interior (toward the centroid) so doors default to swinging inward.
            for wall in self._chain_walls:
                dx, dy = wall.x2-wall.x1, wall.y2-wall.y1
                L = math.hypot(dx,dy) or 1.0
                ux, uy = dx/L, dy/L
                mx, my = (wall.x1+wall.x2)/2, (wall.y1+wall.y2)/2
                eps = 0.5
                d_left  = math.hypot((mx-uy*eps)-cx, (my+ux*eps)-cy)
                d_right = math.hypot((mx+uy*eps)-cx, (my-ux*eps)-cy)
                wall.interior_side = -1 if d_left < d_right else 1
            dlg = RoomPropertiesDialog(name=f"Room {len(self._room_items)+1}", area_sqft=area)
            if dlg.exec_() == QDialog.Accepted:
                vals = dlg.values()
                room = RoomLabelItem(cx, cy, vals["name"], area, vals["ceiling_in"],
                                      ceiling_type=vals["ceiling_type"], floor_type=vals["floor_type"])
                self.addItem(room); self._apply_layer_visibility(room)
                self._room_items.append(room)
        self._cancel_wall_chain()
        self.set_mode_select()
        self.layout_changed.emit()

    def _finish_zone_chain(self):
        pts = list(self._zone_chain)
        if len(pts) >= 3:
            n = len(self._zone_items)
            # Color and pattern cycle independently (8 colors × 7 patterns =
            # 56 unique combinations before an exact repeat) so zones stay
            # distinguishable far longer than color alone would allow —
            # matters most for a zone number reused across floors.
            dlg = ZonePropertiesDialog(
                name=f"Zone {n+1}",
                color=ZONE_COLOR_CHOICES[n % len(ZONE_COLOR_CHOICES)],
                pattern=ZONE_PATTERN_CHOICES[n % len(ZONE_PATTERN_CHOICES)])
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                zone = FireZoneItem(pts, v["name"], v["color"], pattern=v["pattern"])
                self.addItem(zone); self._apply_layer_visibility(zone)
                self._zone_items.append(zone)
        self._cancel_zone_chain()
        self.set_mode_select()
        self.layout_changed.emit()

    def _finish_bg_calibration(self):
        """Rescale the background trace image so the two just-clicked points
        end up the real-world distance apart that the user types in —
        keeping the FIRST clicked point fixed in place so already-traced
        walls near it stay aligned."""
        p1, p2 = self._bg_calib_pts
        apparent_ft = math.hypot(p2.x()-p1.x(), p2.y()-p1.y())/PX_PER_FT
        self._bg_calib_pts = []
        self.set_mode_select()
        if apparent_ft < 1e-6:
            return
        text, ok = QInputDialog.getText(None, "Calibrate Background Scale",
            "Real-world distance between the two points you clicked:", text="10'")
        if not ok:
            return
        real_ft = parse_length(text)
        if not real_ft or real_ft <= 0:
            QMessageBox.information(None, "Calibrate Background Scale",
                                     "Couldn't read that as a length (try e.g. 10', 3'6\", 42\").")
            return
        old_px_per_ft = self.bg_px_per_ft
        new_px_per_ft = old_px_per_ft * apparent_ft / real_ft
        ax_ft, ay_ft = p1.x()/PX_PER_FT, p1.y()/PX_PER_FT
        img_x = (ax_ft - self.bg_pos_ft[0]) * old_px_per_ft
        img_y = (ay_ft - self.bg_pos_ft[1]) * old_px_per_ft
        self.bg_px_per_ft = new_px_per_ft
        self.bg_pos_ft = (ax_ft - img_x/new_px_per_ft, ay_ft - img_y/new_px_per_ft)
        self.update()
        self.status_changed.emit(f"  Background calibrated — {apparent_ft:.1f}' now reads as {real_ft:.1f}'.")

    def add_rect_room(self, x0_ft, y0_ft, width_ft, length_ft, name="Room", ceiling_in=108.0,
                       ceiling_type="T-Bar / Lay-in Tile", floor_type="Carpet Tile", thickness_in=None):
        """Build a simple rectangular room in one shot — 4 walls (each with
        its own auto dimension, same as manual wall drawing) plus a room
        label, instead of clicking out each corner by hand. thickness_in
        defaults to the scene's current default (None) — pasting a copied
        room passes the original's actual thickness explicitly instead."""
        if self.snap_to_grid:
            x0_ft, y0_ft = snap_grid(x0_ft), snap_grid(y0_ft)
        corners = [(x0_ft,y0_ft), (x0_ft+width_ft,y0_ft),
                   (x0_ft+width_ft,y0_ft+length_ft), (x0_ft,y0_ft+length_ft)]
        walls = [self._add_wall(*corners[i], *corners[(i+1) % 4], thickness_in=thickness_in)
                 for i in range(4)]
        cx, cy = x0_ft+width_ft/2, y0_ft+length_ft/2
        for wall in walls:
            dx, dy = wall.x2-wall.x1, wall.y2-wall.y1
            L = math.hypot(dx,dy) or 1.0
            ux, uy = dx/L, dy/L
            mx, my = (wall.x1+wall.x2)/2, (wall.y1+wall.y2)/2
            eps = 0.5
            d_left  = math.hypot((mx-uy*eps)-cx, (my+ux*eps)-cy)
            d_right = math.hypot((mx+uy*eps)-cx, (my-ux*eps)-cy)
            wall.interior_side = -1 if d_left < d_right else 1
        room = RoomLabelItem(cx, cy, name, width_ft*length_ft, ceiling_in,
                              ceiling_type=ceiling_type, floor_type=floor_type)
        # Session-only convenience (not persisted) — lets "Resize Rectangle
        # Room…" on the room label regenerate these same 4 walls in place.
        room._rect_walls = walls
        room._rect_origin = (x0_ft, y0_ft)
        room._rect_size = (width_ft, length_ft)
        self.addItem(room); self._apply_layer_visibility(room)
        self._room_items.append(room)
        self.layout_changed.emit()
        return room

    # ── Copy / paste ─────────────────────────────────────────────────────────

    def copy_selected(self):
        """Copy a selected rectangle room (one placed via 'Add Rectangle
        Room…', or pasted from a previous copy) so Paste can stamp out a
        duplicate — walls, dimensions and label together, not just the
        label text. Freeform (wall-chain-drawn) rooms aren't a single bundled
        object the same way, so copy only supports rectangle rooms for now."""
        for it in self.selectedItems():
            if isinstance(it, RoomLabelItem) and hasattr(it, "_rect_walls"):
                self._clipboard_room = {
                    "name": it.name, "width_ft": it._rect_size[0], "length_ft": it._rect_size[1],
                    "ceiling_in": it.ceiling_in, "ceiling_type": it.ceiling_type,
                    "floor_type": it.floor_type,
                    "thickness_in": (it._rect_walls[0].thickness_in if it._rect_walls
                                      else self.default_wall_thickness_in),
                    "next_origin": (it._rect_origin[0]+4.0, it._rect_origin[1]+4.0),
                }
                self.status_changed.emit(f"  Copied \"{it.name}\" — Ctrl+V to paste a duplicate "
                                          "(drag it into place afterward).")
                return
        self.status_changed.emit("  Select a rectangle room to copy (Ctrl+C) — "
                                  "one placed via 'Add Rectangle Room…'.")

    def paste_clipboard(self):
        c = getattr(self, "_clipboard_room", None)
        if not c:
            self.status_changed.emit("  Nothing to paste — copy a rectangle room first (Ctrl+C).")
            return
        x0, y0 = c["next_origin"]
        room = self.add_rect_room(x0, y0, c["width_ft"], c["length_ft"], f'{c["name"]} (Copy)',
                                   c["ceiling_in"], c["ceiling_type"], c["floor_type"],
                                   thickness_in=c["thickness_in"])
        for it in self.selectedItems():
            it.setSelected(False)
        room.setSelected(True)
        c["next_origin"] = (room._rect_origin[0]+4.0, room._rect_origin[1]+4.0)
        self.status_changed.emit(f"  Pasted \"{room.name}\" — drag it to reposition.")

    # ── Mouse / keyboard ─────────────────────────────────────────────────────

    def mousePressEvent(self, event):
        if self._mode == "wall" and event.button() == Qt.LeftButton:
            raw = event.scenePos()
            x_ft, y_ft = raw.x()/PX_PER_FT, raw.y()/PX_PER_FT
            from_pt = self._chain[-1] if self._chain else None
            x_ft, y_ft = self._snap_point(x_ft, y_ft, from_pt)
            if self._length_buffer and from_pt is not None:
                length = parse_length(self._length_buffer)
                if length is not None:
                    dx, dy = snap_angle(x_ft-from_pt[0], y_ft-from_pt[1], 45.0)
                    ang = math.atan2(dy, dx)
                    x_ft = from_pt[0] + length*math.cos(ang)
                    y_ft = from_pt[1] + length*math.sin(ang)
            self._length_buffer = ""
            if not self._chain:
                self._chain = [(x_ft, y_ft)]
                self._preview.points = list(self._chain)
                self.update()
                return
            # closing the loop?
            first = self._chain[0]
            if len(self._chain) >= 3 and math.hypot(x_ft-first[0], y_ft-first[1]) <= SNAP_TOL_FT:
                last = self._chain[-1]
                self._add_wall(last[0], last[1], first[0], first[1])
                self._chain.append(first)
                self._finish_chain(closed=True)
                return
            last = self._chain[-1]
            self._add_wall(last[0], last[1], x_ft, y_ft)
            self._chain.append((x_ft, y_ft))
            self._preview.points = list(self._chain)
            self.update()
            return
        if self._mode == "symbol" and event.button() == Qt.LeftButton:
            self._place_symbol(event.scenePos())
            return
        if self._mode == "zone" and event.button() == Qt.RightButton:
            self._cancel_zone_chain()
            return
        if self._mode == "zone" and event.button() == Qt.LeftButton:
            raw = event.scenePos()
            x_ft, y_ft = raw.x()/PX_PER_FT, raw.y()/PX_PER_FT
            if self.snap_to_grid:
                x_ft, y_ft = snap_grid(x_ft), snap_grid(y_ft)
            if not self._zone_chain:
                self._zone_chain = [(x_ft, y_ft)]
                self._zone_preview.points = list(self._zone_chain)
                self.update()
                return
            first = self._zone_chain[0]
            if len(self._zone_chain) >= 3 and math.hypot(x_ft-first[0], y_ft-first[1]) <= SNAP_TOL_FT:
                self._finish_zone_chain()
                return
            self._zone_chain.append((x_ft, y_ft))
            self._zone_preview.points = list(self._zone_chain)
            self.update()
            return
        if self._mode == "bg_calibrate" and event.button() == Qt.LeftButton:
            self._bg_calib_pts.append(event.scenePos())
            if len(self._bg_calib_pts) == 2:
                self._finish_bg_calibration()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._mode == "wall" and self._chain:
            raw = event.scenePos()
            x_ft, y_ft = raw.x()/PX_PER_FT, raw.y()/PX_PER_FT
            from_pt = self._chain[-1]
            x_ft, y_ft = self._snap_point(x_ft, y_ft, from_pt)
            if self._length_buffer:
                length = parse_length(self._length_buffer)
                if length is not None:
                    dx, dy = snap_angle(x_ft-from_pt[0], y_ft-from_pt[1], 45.0)
                    ang = math.atan2(dy, dx)
                    x_ft = from_pt[0] + length*math.cos(ang)
                    y_ft = from_pt[1] + length*math.sin(ang)
                self._preview.label = self._length_buffer
            else:
                seg_len = math.hypot(x_ft-from_pt[0], y_ft-from_pt[1])
                self._preview.label = fmt_feet(seg_len)
            self._preview.preview_pt = (x_ft, y_ft)
            self.update()
            return
        if self._mode == "zone" and self._zone_chain:
            raw = event.scenePos()
            x_ft, y_ft = raw.x()/PX_PER_FT, raw.y()/PX_PER_FT
            if self.snap_to_grid:
                x_ft, y_ft = snap_grid(x_ft), snap_grid(y_ft)
            self._zone_preview.preview_pt = (x_ft, y_ft)
            self.update()
            return
        super().mouseMoveEvent(event)

    def mouseDoubleClickEvent(self, event):
        if self._mode == "wall" and self._chain:
            self._finish_chain(closed=False)
            return
        if self._mode == "zone" and len(self._zone_chain) >= 3:
            self._finish_zone_chain()
            return
        super().mouseDoubleClickEvent(event)

    def keyPressEvent(self, event):
        if self._mode == "wall":
            t = event.text()
            if event.key() == Qt.Key_Escape:
                self._cancel_wall_chain(); self.set_mode_select(); return
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                if self._length_buffer and self._chain and self._preview.preview_pt is not None:
                    # commit the typed length as the next point
                    from_pt = self._chain[-1]
                    length = parse_length(self._length_buffer)
                    if length is not None:
                        px, py = self._preview.preview_pt
                        dx, dy = snap_angle(px-from_pt[0], py-from_pt[1], 45.0)
                        ang = math.atan2(dy, dx)
                        x_ft = from_pt[0] + length*math.cos(ang)
                        y_ft = from_pt[1] + length*math.sin(ang)
                        first = self._chain[0]
                        if len(self._chain) >= 3 and math.hypot(x_ft-first[0], y_ft-first[1]) <= SNAP_TOL_FT:
                            self._add_wall(from_pt[0], from_pt[1], first[0], first[1])
                            self._chain.append(first)
                            self._finish_chain(closed=True)
                            return
                        self._add_wall(from_pt[0], from_pt[1], x_ft, y_ft)
                        self._chain.append((x_ft, y_ft))
                        self._preview.points = list(self._chain)
                        self._length_buffer = ""
                        self.update()
                        return
                self._finish_chain(closed=False)
                return
            if event.key() == Qt.Key_Backspace:
                self._length_buffer = self._length_buffer[:-1]
                self.update(); return
            if t and (t.isdigit() or t in ".'\""):
                self._length_buffer += t
                self.update(); return
        if self._mode == "zone":
            if event.key() == Qt.Key_Escape:
                self._cancel_zone_chain(); self.set_mode_select(); return
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                if len(self._zone_chain) >= 3:
                    self._finish_zone_chain()
                return
        if event.key() in (Qt.Key_Delete, Qt.Key_Backspace) and self.selectedItems():
            self.delete_selected(); return
        if event.key() == Qt.Key_C and (event.modifiers() & Qt.ControlModifier):
            self.copy_selected(); return
        if event.key() == Qt.Key_V and (event.modifiers() & Qt.ControlModifier):
            self.paste_clipboard(); return
        if event.key() == Qt.Key_Escape:
            self.set_mode_select(); return
        super().keyPressEvent(event)

    # ── Symbol placement ────────────────────────────────────────────────────

    def _nearest_wall(self, pt, max_dist_ft=2.0):
        """Return (wall, t, dist_ft) for the closest wall to scene point pt,
        where t is the 0..1 fraction along the wall of the closest point."""
        best = None; best_d = max_dist_ft
        for w in self._wall_items:
            x1,y1,x2,y2 = w.x1,w.y1,w.x2,w.y2
            dx,dy = x2-x1, y2-y1
            L2 = dx*dx+dy*dy
            px, py = pt.x()/PX_PER_FT, pt.y()/PX_PER_FT
            if L2 < 1e-9:
                continue
            t = max(0.0, min(1.0, ((px-x1)*dx+(py-y1)*dy)/L2))
            cx, cy = x1+t*dx, y1+t*dy
            d = math.hypot(px-cx, py-cy)
            if d < best_d:
                best_d = d; best = (w, t, d)
        return best

    def _coincident_walls(self, x_ft, y_ft, exclude, tol=0.1):
        """Other walls whose segment passes through (x_ft,y_ft) within tol —
        catches the shared wall between two independently-drawn adjoining
        rooms (each room's chain draws its own wall segment along the same
        line), so a door/window placed there can cut a gap in ALL of them
        instead of only the one _nearest_wall happened to pick — otherwise
        the untouched duplicate wall renders solid right over the opening."""
        found = []
        for w in self._wall_items:
            if w is exclude:
                continue
            dx, dy = w.x2-w.x1, w.y2-w.y1
            L2 = dx*dx+dy*dy
            if L2 < 1e-9:
                continue
            t = ((x_ft-w.x1)*dx + (y_ft-w.y1)*dy)/L2
            if t < -0.05 or t > 1.05:
                continue
            cx, cy = w.x1+t*dx, w.y1+t*dy
            if math.hypot(x_ft-cx, y_ft-cy) <= tol:
                found.append(w)
        return found

    def _place_symbol(self, scene_pt):
        kind = self._pending_symbol
        if not kind:
            return
        is_opening = kind in DOOR_WINDOW_DEFS
        if is_opening:
            hit = self._nearest_wall(scene_pt, max_dist_ft=3.0)
            if hit:
                w, t, _ = hit
                x = w.x1 + t*(w.x2-w.x1); y = w.y1 + t*(w.y2-w.y1)
                ang = math.degrees(math.atan2(w.y2-w.y1, w.x2-w.x1))
                item = DoorWindowItem(kind, x*PX_PER_FT, y*PX_PER_FT, ang)
                item.host_wall = w
                w.opening_items.append(item)
                w.prepareGeometryChange(); w.update()
                for ow in self._coincident_walls(x, y, exclude=w):
                    item.extra_host_walls.append(ow)
                    ow.opening_items.append(item)
                    ow.prepareGeometryChange(); ow.update()
            else:
                item = DoorWindowItem(kind, scene_pt.x(), scene_pt.y(), 0.0)
                self.status_changed.emit("  ⚠ No wall found nearby — placed unattached; drag onto a wall.")
            self.addItem(item); self._apply_layer_visibility(item)
            self._opening_items.append(item)
        else:
            snapped = self.snap_symbol_point(scene_pt)
            item = SymbolItem(kind, snapped.x(), snapped.y())
            self.addItem(item); self._apply_layer_visibility(item)
            self._symbol_items.append(item)
        self.layout_changed.emit()

    def snap_symbol_point(self, scene_pt):
        """Snap a symbol's proposed scene position — onto the nearest wall
        face (with along-wall grid snap, so a row of them lines up evenly)
        if one is close by, else plain grid snap. Shared by both the initial
        placement click (_place_symbol) and live dragging of an existing
        symbol (SymbolItem.itemChange), so both behave identically.

        With Snap OFF this is a no-op — full manual placement, including no
        wall-proximity snap — since that wall snap used to fire regardless
        of the toggle, which meant "Snap OFF" was never actually off for
        anything placed near a wall."""
        x_ft, y_ft = scene_pt.x()/PX_PER_FT, scene_pt.y()/PX_PER_FT
        if not self.snap_to_grid:
            return QPointF(x_ft*PX_PER_FT, y_ft*PX_PER_FT)
        hit = self._nearest_wall(scene_pt, max_dist_ft=3.0)
        if hit:
            # Snap onto the wall face: project along the wall (and snap that
            # position to the grid, so repeated placements — e.g. a row of
            # outlets — land at even intervals instead of drifting to
            # wherever the mouse happened to be), then hold a fixed standoff
            # off the wall face so they all sit on one clean line.
            w, t, _ = hit
            dx, dy = w.x2-w.x1, w.y2-w.y1
            L = math.hypot(dx, dy) or 1.0
            ux, uy = dx/L, dy/L
            nx, ny = -uy, ux
            along = max(0.0, min(L, snap_grid(t*L)))
            wx, wy = w.x1+ux*along, w.y1+uy*along
            side = getattr(w, "interior_side", 1) or 1
            standoff_ft = w.thickness_in/24.0 + 0.4
            x_ft, y_ft = wx+nx*standoff_ft*side, wy+ny*standoff_ft*side
        else:
            x_ft, y_ft = snap_grid(x_ft), snap_grid(y_ft)
        return QPointF(x_ft*PX_PER_FT, y_ft*PX_PER_FT)

    # ── Bulk ops ─────────────────────────────────────────────────────────────

    def all_items_bbox_ft(self):
        """Bounding box (min_x,min_y,max_x,max_y) in feet across all content."""
        xs, ys = [], []
        for w in self._wall_items:
            xs += [w.x1, w.x2]; ys += [w.y1, w.y2]
        for it in self._symbol_items + self._opening_items:
            xs.append(it.pos().x()/PX_PER_FT); ys.append(it.pos().y()/PX_PER_FT)
        for it in self._room_items:
            xs.append(it.pos().x()/PX_PER_FT); ys.append(it.pos().y()/PX_PER_FT)
        for z in self._zone_items:
            for x, y in z.points:
                xs.append(x); ys.append(y)
        if not xs:
            return (0.0, 0.0, 10.0, 10.0)
        return (min(xs), min(ys), max(xs), max(ys))

    def clear_all(self):
        for lst in (self._wall_items, self._dim_items, self._room_items,
                    self._symbol_items, self._opening_items, self._zone_items):
            for it in list(lst):
                self.removeItem(it)
            lst.clear()
        self.layout_changed.emit()

    def remove_item(self, it):
        """Remove one item and keep all bookkeeping (paired dimension, door/window
        <-> host-wall opening lists) consistent. Safe to call more than once."""
        if it.scene() is None and it not in (self._wall_items+self._dim_items+self._room_items
                                              +self._symbol_items+self._opening_items+self._zone_items):
            return
        item_type = getattr(it, "ITEM_TYPE", "")
        if item_type == "wall":
            # cascade-delete attached doors/windows (their opening no longer has a wall)
            for o in list(it.opening_items):
                self.remove_item(o)
            # remove its paired auto-dimension
            for d in list(self._dim_items):
                if (d.x1,d.y1,d.x2,d.y2) == (it.x1,it.y1,it.x2,it.y2):
                    self.removeItem(d)
                    if d in self._dim_items: self._dim_items.remove(d)
        elif item_type == "opening":
            for hw in [it.host_wall] + list(it.extra_host_walls):
                if hw is not None and it in hw.opening_items:
                    hw.opening_items.remove(it)
                    hw.update()
        for lst in (self._wall_items, self._dim_items, self._room_items,
                    self._symbol_items, self._opening_items, self._zone_items):
            if it in lst:
                lst.remove(it)
        if it.scene() is not None:
            self.removeItem(it)
        self.layout_changed.emit()

    def delete_selected(self):
        for it in list(self.selectedItems()):
            self.remove_item(it)

    # ── Save / load ──────────────────────────────────────────────────────────

    def to_dict(self):
        return {
            "walls":   [w.to_dict() for w in self._wall_items],
            "dims":    [d.to_dict() for d in self._dim_items],
            "rooms":   [r.to_dict() for r in self._room_items],
            "symbols": [s.to_dict() for s in self._symbol_items],
            "openings":[o.to_dict() for o in self._opening_items],
            "zones":   [z.to_dict() for z in self._zone_items],
            "layers_visible": self.layers_visible,
        }

    def load_dict(self, d):
        self.clear_all()
        for wd in d.get("walls", []):
            w = WallItem.from_dict(wd); self.addItem(w); self._wall_items.append(w)
        for dd in d.get("dims", []):
            dm = DimensionItem.from_dict(dd); self.addItem(dm); self._dim_items.append(dm)
        for rd in d.get("rooms", []):
            r = RoomLabelItem.from_dict(rd); self.addItem(r); self._room_items.append(r)
        for sd in d.get("symbols", []):
            s = SymbolItem.from_dict(sd); self.addItem(s); self._symbol_items.append(s)
        for od in d.get("openings", []):
            o = DoorWindowItem.from_dict(od); self.addItem(o); self._opening_items.append(o)
            # host_wall isn't serialized (not JSON-able) — re-snap to whichever
            # wall the door/window was saved sitting on.
            hit = self._nearest_wall(o.pos(), max_dist_ft=1.0)
            if hit:
                w, _t, _d = hit
                o.host_wall = w
                w.opening_items.append(o)
                x_ft, y_ft = o.pos().x()/PX_PER_FT, o.pos().y()/PX_PER_FT
                for ow in self._coincident_walls(x_ft, y_ft, exclude=w):
                    o.extra_host_walls.append(ow)
                    ow.opening_items.append(o)
        for zd in d.get("zones", []):
            z = FireZoneItem.from_dict(zd); self.addItem(z); self._zone_items.append(z)
        self.layers_visible.update(d.get("layers_visible", {}))
        for w in self._wall_items:
            w.prepareGeometryChange(); w.update()
        for lst in (self._wall_items, self._dim_items, self._room_items,
                    self._symbol_items, self._opening_items, self._zone_items):
            for it in lst:
                self._apply_layer_visibility(it)
        self.layout_changed.emit()


# ═══════════════════════════════════════════════════════════════════════════════
#  Canvas (view)
# ═══════════════════════════════════════════════════════════════════════════════

class FloorPlanCanvas(QGraphicsView):
    def __init__(self, scene):
        super().__init__(scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setDragMode(QGraphicsView.RubberBandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self.setMouseTracking(True)
        self._panning = False
        self._pan_start = None

    def wheelEvent(self, e):
        f = 1.15 if e.angleDelta().y() > 0 else 1/1.15
        self.scale(f, f)

    def mousePressEvent(self, e):
        if e.button() == Qt.MiddleButton or (
                e.button() == Qt.LeftButton and (e.modifiers() & Qt.ControlModifier)):
            self._panning = True; self._pan_start = e.pos()
            self.setCursor(Qt.ClosedHandCursor); return
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self._panning and self._pan_start:
            d = e.pos()-self._pan_start; self._pan_start = e.pos()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value()-d.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value()-d.y()); return
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        if self._panning:
            self._panning = False; self.setCursor(Qt.ArrowCursor)
        super().mouseReleaseEvent(e)

    def fit_all(self):
        sc = self.scene()
        items = [i for i in sc.items() if i is not getattr(sc, "_preview", None)
                 and i is not getattr(sc, "_zone_preview", None)]
        if not items:
            self.fitInView(QRectF(0, 0, 30*PX_PER_FT, 20*PX_PER_FT), Qt.KeepAspectRatio)
            return
        r = items[0].sceneBoundingRect()
        for i in items[1:]:
            r = r.united(i.sceneBoundingRect())
        self.fitInView(r.adjusted(-60,-60,60,60), Qt.KeepAspectRatio)


# ═══════════════════════════════════════════════════════════════════════════════
#  Layers panel
# ═══════════════════════════════════════════════════════════════════════════════

class LayerPanel(QWidget):
    visibility_changed = pyqtSignal(str, bool)

    def __init__(self):
        super().__init__()
        self.setMinimumWidth(170); self.setMaximumWidth(210)
        l = QVBoxLayout(self); l.setContentsMargins(6,6,6,6); l.setSpacing(3)
        lbl = QLabel("Layers"); lbl.setStyleSheet("font-weight:bold;font-size:12px;padding:2px;")
        l.addWidget(lbl)
        self._checks = {}
        for name in LAYER_ORDER:
            row = QWidget(); rl = QHBoxLayout(row); rl.setContentsMargins(2,2,2,2); rl.setSpacing(6)
            swatch = QLabel(); swatch.setFixedSize(12,12)
            swatch.setStyleSheet(f"background:{LAYER_COLORS[name]};border-radius:2px;")
            cb = QCheckBox(name); cb.setChecked(True)
            cb.toggled.connect(lambda checked, n=name: self.visibility_changed.emit(n, checked))
            rl.addWidget(swatch); rl.addWidget(cb); rl.addStretch()
            l.addWidget(row)
            self._checks[name] = cb
        l.addStretch()

    def set_checked(self, name, checked):
        if name in self._checks:
            self._checks[name].blockSignals(True)
            self._checks[name].setChecked(checked)
            self._checks[name].blockSignals(False)


# ═══════════════════════════════════════════════════════════════════════════════
#  Symbol palette
# ═══════════════════════════════════════════════════════════════════════════════

class SymbolPalette(QWidget):
    symbol_clicked = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setMinimumWidth(190); self.setMaximumWidth(230)
        outer = QVBoxLayout(self); outer.setContentsMargins(4,4,4,4); outer.setSpacing(4)
        lbl = QLabel("Symbols"); lbl.setStyleSheet("font-weight:bold;font-size:12px;padding:2px;")
        outer.addWidget(lbl)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        inner = QWidget(); il = QVBoxLayout(inner); il.setContentsMargins(2,2,2,2); il.setSpacing(6)
        self._btns = {}
        self._group_kinds = []
        for cat_name, kinds in SYMBOL_CATEGORIES:
            hdr = QPushButton(cat_name)
            hdr.setStyleSheet("background:#3a3d3e;color:#efe6e1;border:none;border-radius:3px;"
                               "padding:4px 6px;font-size:10px;font-weight:bold;text-align:left;")
            il.addWidget(hdr)
            for k in kinds:
                defn = ALL_SYMBOL_DEFS[k]
                b = QPushButton(defn["name"]); b.setCheckable(True)
                b.setStyleSheet(
                    "QPushButton{background:#f2f0ee;color:#2a2a2a;border:1px solid #ccc;border-radius:3px;"
                    "padding:4px 6px;font-size:10px;text-align:left;}"
                    "QPushButton:hover{background:#e6e2de;}"
                    "QPushButton:checked{background:#ff7002;color:white;border-color:#ff7002;}")
                b.clicked.connect(lambda checked, kk=k: self._on_click(kk, checked))
                il.addWidget(b)
                self._btns[k] = b
        il.addStretch()
        scroll.setWidget(inner)
        outer.addWidget(scroll)

    def _on_click(self, kind, checked):
        for k, b in self._btns.items():
            if k != kind:
                b.setChecked(False)
        self.symbol_clicked.emit(kind if checked else "")

    def clear_all(self):
        for b in self._btns.values():
            b.setChecked(False)


# ═══════════════════════════════════════════════════════════════════════════════
#  Dialogs
# ═══════════════════════════════════════════════════════════════════════════════

class RoomPropertiesDialog(QDialog):
    def __init__(self, name="Room", area_sqft=0.0, ceiling_in=108.0,
                 ceiling_type="T-Bar / Lay-in Tile", floor_type="Carpet Tile", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Room Properties"); self.setMinimumWidth(320)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.name_edit = QLineEdit(name)
        self.ceil_ft = QDoubleSpinBox(); self.ceil_ft.setRange(6.0, 30.0); self.ceil_ft.setSingleStep(0.5)
        self.ceil_ft.setValue(ceiling_in/12.0); self.ceil_ft.setSuffix("  ft AFF")
        self.ceiling_type_combo = QComboBox(); self.ceiling_type_combo.setEditable(True)
        self.ceiling_type_combo.addItems(CEILING_TYPES)
        idx = self.ceiling_type_combo.findText(ceiling_type)
        self.ceiling_type_combo.setCurrentIndex(idx if idx >= 0 else 0)
        if idx < 0: self.ceiling_type_combo.setEditText(ceiling_type)
        self.floor_type_combo = QComboBox(); self.floor_type_combo.setEditable(True)
        self.floor_type_combo.addItems(FLOOR_TYPES)
        idx = self.floor_type_combo.findText(floor_type)
        self.floor_type_combo.setCurrentIndex(idx if idx >= 0 else 0)
        if idx < 0: self.floor_type_combo.setEditText(floor_type)
        l.addRow("Room name:", self.name_edit)
        l.addRow("Ceiling height (AFF):", self.ceil_ft)
        l.addRow("Ceiling type:", self.ceiling_type_combo)
        l.addRow("Floor type:", self.floor_type_combo)
        l.addRow(QLabel(f"Area (auto): {area_sqft:.0f} SF"))
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        return {"name": self.name_edit.text().strip() or "Room",
                "ceiling_in": self.ceil_ft.value()*12.0,
                "ceiling_type": self.ceiling_type_combo.currentText().strip() or "Drywall",
                "floor_type": self.floor_type_combo.currentText().strip() or "Carpet Tile"}


ZONE_COLOR_NAMES = {
    "#c0392b": "Red", "#2980b9": "Blue", "#27ae60": "Green", "#e67e22": "Orange",
    "#8e44ad": "Purple", "#16a085": "Teal", "#2c3e50": "Charcoal", "#d35400": "Rust",
}


def _swatch_icon(hex_color, size=16):
    """Small solid-color square icon for a color combo entry, so choices show
    as an actual swatch instead of a bare hex code the user has to decode."""
    pm = QPixmap(size, size)
    pm.fill(QColor(hex_color))
    return QIcon(pm)


def _pattern_icon(hex_color, pattern, size=28):
    """Small preview icon showing the actual hatch pattern in the given
    color, so the pattern combo shows real texture instead of a name —
    same reasoning as _swatch_icon for colors."""
    pm = QPixmap(size, size)
    pm.fill(QColor("white"))
    painter = QPainter(pm)
    painter.setRenderHint(QPainter.Antialiasing)
    col = QColor(hex_color)
    painter.setPen(QPen(col.darker(130), 1))
    if pattern == "solid":
        fill = QColor(col); fill.setAlpha(60)
        painter.setBrush(QBrush(fill))
    else:
        painter.setBrush(QBrush(col, ZONE_QT_PATTERNS.get(pattern, Qt.SolidPattern)))
    painter.drawRect(1, 1, size-2, size-2)
    painter.end()
    return QIcon(pm)


class ZonePropertiesDialog(QDialog):
    def __init__(self, name="Zone 1", color="#c0392b", pattern="solid", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Fire Zone Properties"); self.setMinimumWidth(300)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.name_edit = QLineEdit(name)
        self.color_combo = QComboBox()
        self.color_combo.setIconSize(QSize(16, 16))
        for hex_c in ZONE_COLOR_CHOICES:
            label = ZONE_COLOR_NAMES.get(hex_c, hex_c)
            self.color_combo.addItem(_swatch_icon(hex_c), label, hex_c)
        idx = self.color_combo.findData(color)
        if idx < 0:
            self.color_combo.addItem(_swatch_icon(color), ZONE_COLOR_NAMES.get(color, color), color)
            idx = self.color_combo.count()-1
        self.color_combo.setCurrentIndex(idx)
        self.color_combo.currentIndexChanged.connect(self._refresh_pattern_icons)

        self.pattern_combo = QComboBox()
        self.pattern_combo.setIconSize(QSize(28, 28))
        for key in ZONE_PATTERN_CHOICES:
            self.pattern_combo.addItem(_pattern_icon(color, key), ZONE_PATTERN_NAMES[key], key)
        pidx = self.pattern_combo.findData(pattern)
        self.pattern_combo.setCurrentIndex(pidx if pidx >= 0 else 0)

        l.addRow("Zone name:", self.name_edit)
        l.addRow("Boundary color:", self.color_combo)
        l.addRow("Fill pattern:", self.pattern_combo)
        note = QLabel("A distinct pattern keeps a zone that reuses a color on another "
                       "floor visually distinguishable.")
        note.setWordWrap(True); note.setStyleSheet("color:#888;font-size:10px;")
        l.addRow(note)
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def _refresh_pattern_icons(self):
        """Repaint the pattern combo's icons in the newly-picked color so the
        preview always matches what the zone will actually look like."""
        col = self.color_combo.currentData() or "#c0392b"
        for i, key in enumerate(ZONE_PATTERN_CHOICES):
            self.pattern_combo.setItemIcon(i, _pattern_icon(col, key))

    def values(self):
        return {"name": self.name_edit.text().strip() or "Zone",
                "color": self.color_combo.currentData() or "#c0392b",
                "pattern": self.pattern_combo.currentData() or "solid"}


class RectRoomDialog(QDialog):
    """Quick square/rectangular room: type a width and length and it builds
    the 4 walls (with dimensions) and a room label automatically — no
    manual wall-chain drawing needed."""
    def __init__(self, width_ft=12.0, length_ft=12.0, name="Room", ceiling_in=108.0,
                 ceiling_type="T-Bar / Lay-in Tile", floor_type="Carpet Tile", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Rectangle Room"); self.setMinimumWidth(320)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.name_edit = QLineEdit(name)
        self.w_spin = QDoubleSpinBox(); self.w_spin.setRange(1.0, 500.0); self.w_spin.setDecimals(1)
        self.w_spin.setValue(width_ft); self.w_spin.setSuffix("  ft")
        self.l_spin = QDoubleSpinBox(); self.l_spin.setRange(1.0, 500.0); self.l_spin.setDecimals(1)
        self.l_spin.setValue(length_ft); self.l_spin.setSuffix("  ft")
        self.ceil_ft = QDoubleSpinBox(); self.ceil_ft.setRange(6.0, 30.0); self.ceil_ft.setSingleStep(0.5)
        self.ceil_ft.setValue(ceiling_in/12.0); self.ceil_ft.setSuffix("  ft AFF")
        self.ceiling_type_combo = QComboBox(); self.ceiling_type_combo.setEditable(True)
        self.ceiling_type_combo.addItems(CEILING_TYPES)
        idx = self.ceiling_type_combo.findText(ceiling_type)
        self.ceiling_type_combo.setCurrentIndex(idx if idx >= 0 else 0)
        if idx < 0: self.ceiling_type_combo.setEditText(ceiling_type)
        self.floor_type_combo = QComboBox(); self.floor_type_combo.setEditable(True)
        self.floor_type_combo.addItems(FLOOR_TYPES)
        idx = self.floor_type_combo.findText(floor_type)
        self.floor_type_combo.setCurrentIndex(idx if idx >= 0 else 0)
        if idx < 0: self.floor_type_combo.setEditText(floor_type)
        l.addRow("Room name:", self.name_edit)
        l.addRow("Width:", self.w_spin)
        l.addRow("Length:", self.l_spin)
        l.addRow("Ceiling height (AFF):", self.ceil_ft)
        l.addRow("Ceiling type:", self.ceiling_type_combo)
        l.addRow("Floor type:", self.floor_type_combo)
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        return {"name": self.name_edit.text().strip() or "Room",
                "width_ft": self.w_spin.value(), "length_ft": self.l_spin.value(),
                "ceiling_in": self.ceil_ft.value()*12.0,
                "ceiling_type": self.ceiling_type_combo.currentText().strip() or "Drywall",
                "floor_type": self.floor_type_combo.currentText().strip() or "Carpet Tile"}


class SymbolSizeDialog(QDialog):
    """Resize one placed furniture/stairs symbol instance — other symbols of
    the same kind keep their catalog default footprint."""
    def __init__(self, name, w_in=24.0, d_in=24.0, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Edit Size — {name}"); self.setMinimumWidth(300)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.w_spin = QDoubleSpinBox(); self.w_spin.setRange(1.0, 600.0); self.w_spin.setDecimals(1)
        self.w_spin.setValue(w_in); self.w_spin.setSuffix("  in")
        self.d_spin = QDoubleSpinBox(); self.d_spin.setRange(1.0, 600.0); self.d_spin.setDecimals(1)
        self.d_spin.setValue(d_in); self.d_spin.setSuffix("  in")
        l.addRow("Width:", self.w_spin)
        l.addRow("Depth:", self.d_spin)
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        return {"w_in": self.w_spin.value(), "d_in": self.d_spin.value()}


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF export — auto (or manual) scale-to-fit-page
# ═══════════════════════════════════════════════════════════════════════════════

TITLE_BLOCK_H_IN = 1.1   # reserved strip at the bottom of the sheet
MARGIN_IN = 0.5

def export_floor_plan_pdf(scene, path, paper_key, orientation, feet_per_inch,
                           region_bbox=None, project_meta=None, sheet_title="FLOOR PLAN"):
    """Render the floor plan to a single PDF page sized to paper_key/orientation
    at the given feet_per_inch scale, centered within the printable area."""
    project_meta = project_meta or {}
    pw_in, ph_in = PAPER_SIZES_IN[paper_key]
    if orientation == "landscape" and pw_in < ph_in:
        pw_in, ph_in = ph_in, pw_in
    elif orientation == "portrait" and pw_in > ph_in:
        pw_in, ph_in = ph_in, pw_in

    bbox = region_bbox or scene.all_items_bbox_ft()
    min_x, min_y, max_x, max_y = bbox
    width_ft = max(0.1, max_x-min_x); height_ft = max(0.1, max_y-min_y)

    printable_w_in = pw_in - 2*MARGIN_IN
    printable_h_in = ph_in - MARGIN_IN - TITLE_BLOCK_H_IN
    px_per_ft = 72.0/feet_per_inch   # PDF points per foot

    content_w_pt = width_ft*px_per_ft
    content_h_pt = height_ft*px_per_ft
    origin_x = MARGIN_IN*72 + max(0.0, (printable_w_in*72 - content_w_pt)/2)
    origin_y = MARGIN_IN*72 + max(0.0, (printable_h_in*72 - content_h_pt)/2)

    def tx(x_ft, y_ft):
        return fitz.Point(origin_x + (x_ft-min_x)*px_per_ft, origin_y + (y_ft-min_y)*px_per_ft)

    doc = fitz.open()
    page = doc.new_page(width=pw_in*72, height=ph_in*72)
    shape = page.new_shape()

    def visible(it):
        return scene.layers_visible.get(it.layer, True)

    # Fire zones — translucent background tint + dashed border, drawn first
    # so walls/symbols/dimensions print on top of it. Name label text
    # deferred to the text pass below, same reason as dimension labels.
    zone_labels = []
    if visible_layer_any(scene, LAYER_FIRE):
        for z in scene._zone_items:
            if not visible(z) or len(z.points) < 3:
                continue
            col_hex = z.color
            col = tuple(int(col_hex[i:i+2],16)/255.0 for i in (1,3,5))
            pts = [tx(x, y) for x, y in z.points]
            shape.draw_polyline(pts+[pts[0]])
            shape.finish(color=col, fill=col, width=1.4, dashes="[4 3] 0", fill_opacity=0.15)
            cx, cy = z.centroid_ft()
            lx, ly = z.label_pos_ft()
            label_pt = tx(lx, ly)
            if (lx, ly) != (cx, cy):
                anchor_pt = tx(cx, cy)
                shape.draw_line(anchor_pt, label_pt)
                shape.finish(color=col, width=1.0)
                shape.draw_circle(anchor_pt, 2.2)
                shape.finish(color=col, fill=col, width=0.5)
            zone_labels.append((label_pt, z.name, col))

    # Walls — filled polygon per solid segment (gaps cut for doors/windows)
    for w in scene._wall_items:
        if not visible(w): continue
        t = w.thickness_in/12.0
        for sx1, sy1, sx2, sy2 in w.solid_segments_ft():
            dx, dy = sx2-sx1, sy2-sy1
            L = math.hypot(dx, dy) or 1.0
            nx, ny = -dy/L*t/2, dx/L*t/2
            pts = [tx(sx1+nx, sy1+ny), tx(sx2+nx, sy2+ny),
                   tx(sx2-nx, sy2-ny), tx(sx1-nx, sy1-ny)]
            shape.draw_polyline(pts+[pts[0]])
            shape.finish(color=(0.23,0.24,0.24), fill=(0.75,0.76,0.77), width=0.6)

    # Dimensions — lines now, length labels deferred to the text pass below
    # (text inserted before shape.commit() ends up UNDER whatever the shape
    # draws, e.g. a wall's fill, since commit() flushes everything the Shape
    # accumulated as one later content-stream chunk regardless of call order).
    dim_labels = []
    if visible_layer_any(scene, LAYER_DIMS):
        for d in scene._dim_items:
            if not visible(d): continue
            dx, dy = d.x2-d.x1, d.y2-d.y1
            L = math.hypot(dx, dy)
            if L < 0.01: continue
            ux, uy = dx/L, dy/L; nx, ny = -uy, ux
            p1 = tx(d.x1+nx*d.offset_ft, d.y1+ny*d.offset_ft)
            p2 = tx(d.x2+nx*d.offset_ft, d.y2+ny*d.offset_ft)
            shape.draw_line(tx(d.x1,d.y1), p1); shape.finish(color=(0.75,0.22,0.17), width=0.4)
            shape.draw_line(tx(d.x2,d.y2), p2); shape.finish(color=(0.75,0.22,0.17), width=0.4)
            shape.draw_line(p1, p2); shape.finish(color=(0.75,0.22,0.17), width=0.6)
            mid = fitz.Point((p1.x+p2.x)/2, (p1.y+p2.y)/2)
            dim_labels.append((mid, fmt_feet(L)))

    # Doors / windows
    for o in scene._opening_items:
        if not visible(o): continue
        _draw_opening_pdf(shape, page, o, tx)

    # Symbols — actual per-kind icon shapes (mirrors SymbolItem.paint()),
    # not a generic circle, so different device types read as different
    # marks on the printed plan. Tag/dimension/label text deferred, same
    # reason as the dimension lines above.
    symbol_labels = []
    for s in scene._symbol_items:
        if not visible(s): continue
        p = tx(s.pos().x()/PX_PER_FT, s.pos().y()/PX_PER_FT)
        col_hex = s.defn.get("color") or LAYER_COLORS.get(s.layer, "#333333")
        col = tuple(int(col_hex[i:i+2],16)/255.0 for i in (1,3,5))
        tag = s.defn.get("tag") or "".join(w[0] for w in s.defn["name"].split()[:2]).upper()
        defn_eff = s.defn
        if s.size_override:
            defn_eff = dict(s.defn); defn_eff["w"], defn_eff["d"] = s.size_override
        half_h = _draw_symbol_icon_pdf(shape, defn_eff, p.x, p.y, col, px_per_ft, s.rotation())
        symbol_labels.append((s, p, col, tag, half_h, defn_eff))

    shape.commit()

    # Text pass — everything below is safely on top of the committed shapes.
    for mid, text in dim_labels:
        page.insert_text(fitz.Point(mid.x-22, mid.y-4), text,
                          fontsize=9, fontname="helv", color=(0.75,0.22,0.17))

    for p, name, col in zone_labels:
        page.insert_text(fitz.Point(p.x-50, p.y), name, fontsize=10,
                          fontname="helv", color=col)

    # Room labels
    for r in scene._room_items:
        if not visible(r): continue
        p = tx(r.pos().x()/PX_PER_FT, r.pos().y()/PX_PER_FT)
        page.insert_text(fitz.Point(p.x-50, p.y-4), r.name, fontsize=11,
                          fontname="helv", color=(0.1,0.1,0.1))
        ceil_ft = r.ceiling_in/12.0
        page.insert_text(fitz.Point(p.x-50, p.y+12),
                          f"{r.area_sqft:.0f} SF   ·   {fmt_feet(ceil_ft)} AFF",
                          fontsize=9, color=(0.3,0.3,0.3))
        page.insert_text(fitz.Point(p.x-50, p.y+24),
                          f"Clg: {r.ceiling_type}   ·   Flr: {r.floor_type}",
                          fontsize=8, color=(0.45,0.45,0.45))

    # Tag position mirrors the official Calgary symbol sheet: on/inside the
    # icon for shapes that are literal text-carrying glyphs there (HAZ
    # diamond, FAAP-style boxes, "E"/"W"/"O2" circles, generator box,
    # elevator badge), above the icon for the north-arrow compass, and below
    # for every other (purely pictorial) icon — where some icons (wap,
    # camera, furniture) aren't a simple filled blob a short abbreviation
    # can sit inside of legibly.
    for s, p, col, tag, half_h, defn_eff in symbol_labels:
        shp_key = defn_eff.get("shape")
        below_y = p.y + half_h + 9
        if tag and shp_key in TAG_ON_ICON_SHAPES:
            tag_col = _legible_tag_color(col)
            lines = tag.split("\n")
            fs = 6.5
            start_y = p.y - (fs*1.15*(len(lines)-1))/2 + fs*0.35
            for i, line in enumerate(lines):
                tw = fitz.get_text_length(line, fontname="hebo", fontsize=fs)
                page.insert_text(fitz.Point(p.x-tw/2, start_y+i*fs*1.15), line,
                                  fontsize=fs, fontname="hebo", color=tag_col)
        elif tag and shp_key in TAG_ABOVE_ICON_SHAPES:
            tw = fitz.get_text_length(tag, fontname="hebo", fontsize=6)
            page.insert_text(fitz.Point(p.x-tw/2, p.y-half_h-4), tag, fontsize=6,
                              fontname="hebo", color=_legible_tag_color(col))
        elif tag:
            tw = fitz.get_text_length(tag, fontname="hebo", fontsize=6)
            page.insert_text(fitz.Point(p.x-tw/2, below_y), tag, fontsize=6,
                              fontname="hebo", color=_legible_tag_color(col))
            below_y += 11
        if defn_eff.get("shape") in ("rect", "circle", "cabinet", "stairs"):
            dim_text = f'{fmt_feet(defn_eff.get("w",24)/12.0)} x {fmt_feet(defn_eff.get("d",24)/12.0)}'
            dtw = fitz.get_text_length(dim_text, fontname="helv", fontsize=6.5)
            page.insert_text(fitz.Point(p.x-dtw/2, below_y), dim_text, fontsize=6.5, color=(0.4,0.4,0.4))
            below_y += 10
        if s.label:
            page.insert_text(fitz.Point(p.x-20, below_y), s.label, fontsize=7.5, color=(0.3,0.3,0.3))

    # Title block
    tb_y = (ph_in-TITLE_BLOCK_H_IN)*72
    page.draw_line(fitz.Point(MARGIN_IN*72, tb_y), fitz.Point((pw_in-MARGIN_IN)*72, tb_y),
                    color=(0,0,0), width=1.0)
    page.insert_text(fitz.Point(MARGIN_IN*72, tb_y+16), sheet_title,
                      fontsize=13, fontname="helv", color=(0,0,0))
    scale_label = next((lbl for lbl, fpi in SCALE_OPTIONS if abs(fpi-feet_per_inch) < 1e-6),
                        f'1" = {feet_per_inch:.1f}\'')
    info_lines = [
        f"Scale: {scale_label}   ·   Paper: {paper_key}",
        f"Project: {project_meta.get('customer','')}   {project_meta.get('job_number','')}",
        f"Date: {datetime.date.today().strftime('%b %d, %Y')}",
    ]
    for i, line in enumerate(info_lines):
        page.insert_text(fitz.Point(MARGIN_IN*72, tb_y+32+i*13), line, fontsize=8, color=(0.2,0.2,0.2))

    _append_symbol_legend_page(doc, scene, project_meta, sheet_title, visible)

    doc.save(path)
    doc.close()


FSP_LEGEND_COL_W_IN = 2.6
FSP_LEGEND_GAP_IN = 0.2


def export_fire_safety_plan_pdf(scene, path, paper_key, orientation, project_meta=None,
                                 sheet_title="FIRE SAFETY PLAN", symbol_scale=2.0, text_pt=12):
    """Dedicated single-page export for the Fire Safety Plan / 'You Are Here'
    evacuation diagram — deliberately NOT a copy of export_floor_plan_pdf's
    behavior in three ways a plan posted on a wall actually needs:
      1. Always one page: the on-page legend column's width is subtracted
         from the printable area BEFORE the auto-fit scale is computed, so
         there's always room for it instead of a second legend page.
      2. Walls print as a plain black centerline, not a filled/shaded
         parallelogram — reads more clearly at a glance than construction
         wall-thickness detail nobody evacuating a building needs.
      3. Symbol icons and label text are drawn at caller-supplied sizes
         (symbol_scale / text_pt) well above the regular export's fixed,
         print-at-a-desk sizing, since this sheet needs to be legible from
         a few feet away, not zoomed into on screen.
    """
    project_meta = project_meta or {}
    pw_in, ph_in = PAPER_SIZES_IN[paper_key]
    if orientation == "landscape" and pw_in < ph_in:
        pw_in, ph_in = ph_in, pw_in
    elif orientation == "portrait" and pw_in > ph_in:
        pw_in, ph_in = ph_in, pw_in

    bbox = scene.all_items_bbox_ft()
    min_x, min_y, max_x, max_y = bbox
    width_ft = max(0.1, max_x-min_x); height_ft = max(0.1, max_y-min_y)

    def visible(it):
        return scene.layers_visible.get(it.layer, True)

    used_kinds = sorted(
        {s.kind for s in scene._symbol_items
         if visible(s) and ALL_SYMBOL_DEFS.get(s.kind, {}).get("layer") == LAYER_FIRE},
        key=lambda k: ALL_SYMBOL_DEFS.get(k, {}).get("name", k))
    legend_zones = [z for z in scene._zone_items if visible(z) and len(z.points) >= 3]
    has_legend = bool(used_kinds) or bool(legend_zones)
    legend_w_in = FSP_LEGEND_COL_W_IN if has_legend else 0.0
    legend_gap_in = FSP_LEGEND_GAP_IN if has_legend else 0.0

    printable_w_in = pw_in - 2*MARGIN_IN - legend_w_in - legend_gap_in
    printable_h_in = ph_in - MARGIN_IN - TITLE_BLOCK_H_IN
    feet_per_inch = best_fit_scale(width_ft, height_ft, printable_w_in, printable_h_in, margin_in=0.0)
    px_per_ft = 72.0/feet_per_inch

    content_w_pt = width_ft*px_per_ft
    content_h_pt = height_ft*px_per_ft
    origin_x = MARGIN_IN*72 + max(0.0, (printable_w_in*72 - content_w_pt)/2)
    origin_y = MARGIN_IN*72 + max(0.0, (printable_h_in*72 - content_h_pt)/2)

    def tx(x_ft, y_ft):
        return fitz.Point(origin_x + (x_ft-min_x)*px_per_ft, origin_y + (y_ft-min_y)*px_per_ft)

    def zone_check(x_ft, y_ft):
        """Whether (x_ft,y_ft) falls inside any VISIBLE drawn zone — used to
        give symbols/text a clear white halo there so a hatch pattern
        doesn't run behind and fight with them for readability."""
        return point_in_any_zone([z for z in scene._zone_items if visible(z)], x_ft, y_ft)

    doc = fitz.open()
    page = doc.new_page(width=pw_in*72, height=ph_in*72)
    shape = page.new_shape()

    # Fire zones — dashed border + either a translucent tint (solid pattern)
    # or a hand-drawn hatch texture (fitz has no pattern-brush equivalent to
    # Qt's), drawn first so walls/symbols print on top. The hatch is what
    # keeps a zone number that reuses a color on another floor visually
    # distinct — same idea as the reference Fire Zone drawing.
    zone_labels = []
    for z in scene._zone_items:
        if not visible(z) or len(z.points) < 3:
            continue
        col_hex = z.color
        col = tuple(int(col_hex[i:i+2],16)/255.0 for i in (1,3,5))
        pts = [tx(x, y) for x, y in z.points]
        pattern = getattr(z, "pattern", "solid")
        if pattern == "solid":
            shape.draw_polyline(pts+[pts[0]])
            shape.finish(color=col, fill=col, width=1.6, dashes="[4 3] 0", fill_opacity=0.15)
        else:
            shape.draw_polyline(pts+[pts[0]])
            shape.finish(color=col, width=1.6, dashes="[4 3] 0")
            pts_xy = [(p.x, p.y) for p in pts]
            for hang, hspacing in ZONE_HATCH_SPECS.get(pattern, []):
                for (x1, y1), (x2, y2) in hatch_lines_in_polygon(pts_xy, hang, hspacing):
                    shape.draw_line(fitz.Point(x1, y1), fitz.Point(x2, y2))
                    shape.finish(color=col, width=0.7)
        cx, cy = z.centroid_ft()
        lx, ly = z.label_pos_ft()
        label_pt = tx(lx, ly)
        if (lx, ly) != (cx, cy):
            # Leader line + anchor dot back to the zone — the label was
            # dragged clear because the zone itself is too small/oddly
            # shaped to write the name inside, same convention as the
            # reference Fire Zone Plan's pulled-out callouts.
            anchor_pt = tx(cx, cy)
            shape.draw_line(anchor_pt, label_pt)
            shape.finish(color=col, width=1.2)
            shape.draw_circle(anchor_pt, 2.5)
            shape.finish(color=col, fill=col, width=0.5)
        zone_labels.append((label_pt, z.name, col))

    # Walls — plain black centerline (no thickness fill) per the "we don't
    # need the wall detail" feedback; a posted evacuation diagram reads
    # better as a clean line drawing than a shaded construction plan.
    for w in scene._wall_items:
        if not visible(w): continue
        for sx1, sy1, sx2, sy2 in w.solid_segments_ft():
            shape.draw_line(tx(sx1, sy1), tx(sx2, sy2))
            shape.finish(color=(0,0,0), width=2.2)

    # Doors / windows
    for o in scene._opening_items:
        if not visible(o): continue
        _draw_opening_pdf(shape, page, o, tx)

    # Symbols — drawn at symbol_scale (default well above the regular
    # export) so an exit arrow etc. reads from across a room, not just
    # zoomed-in on screen.
    symbol_labels = []
    for s in scene._symbol_items:
        if not visible(s): continue
        sx_ft, sy_ft = s.pos().x()/PX_PER_FT, s.pos().y()/PX_PER_FT
        p = tx(sx_ft, sy_ft)
        col_hex = s.defn.get("color") or LAYER_COLORS.get(s.layer, "#333333")
        col = tuple(int(col_hex[i:i+2],16)/255.0 for i in (1,3,5))
        tag = s.defn.get("tag") or "".join(w2[0] for w2 in s.defn["name"].split()[:2]).upper()
        defn_eff = s.defn
        if s.size_override:
            defn_eff = dict(s.defn); defn_eff["w"], defn_eff["d"] = s.size_override
        in_zone = zone_check(sx_ft, sy_ft)
        if in_zone:
            # Clear halo behind the icon so a zone's hatch pattern doesn't
            # run right through it — sized generously since icon extents
            # vary by shape (pin, cylinder, diamond, ...).
            halo_r = SYMBOL_ICON_R_PT * symbol_scale * 1.6
            shape.draw_circle(p, halo_r)
            shape.finish(color=None, fill=(1,1,1), width=0, fill_opacity=0.9)
        half_h = _draw_symbol_icon_pdf(shape, defn_eff, p.x, p.y, col, px_per_ft, s.rotation(),
                                       scale=symbol_scale)
        symbol_labels.append((p, col, tag, half_h, defn_eff.get("shape"), in_zone))

    shape.commit()

    def halo_behind(x0, y_baseline, text, fontsize, fontname="helv", pad=2.0):
        """White backing rect sized to `text`, drawn immediately before it's
        inserted, so a zone's hatch pattern doesn't run through readable
        text. Returns the text width so callers already computing it for
        centering don't need get_text_length twice."""
        tw = fitz.get_text_length(text, fontname=fontname, fontsize=fontsize)
        page.draw_rect(fitz.Rect(x0-pad, y_baseline-fontsize*0.82-pad, x0+tw+pad, y_baseline+fontsize*0.28+pad),
                        color=None, fill=(1,1,1), fill_opacity=0.9)
        return tw

    # Text pass — everything below is safely on top of the committed shapes.
    # All text on this sheet is forced uppercase per the Fire Safety Plan
    # text-style requirement ("must be entirely in uppercase for clear
    # visibility and compliance"). Zone name labels always get a halo — at
    # the default centroid position that's always "inside itself"; when
    # dragged off to the side (see the leader line drawn above) it isn't,
    # but the halo is just as needed there.
    for p, name, col in zone_labels:
        name_u = name.upper()
        halo_behind(p.x-60, p.y, name_u, text_pt, "helv")
        page.insert_text(fitz.Point(p.x-60, p.y), name_u, fontsize=text_pt,
                          fontname="helv", color=col)

    for r in scene._room_items:
        if not visible(r): continue
        rx_ft, ry_ft = r.pos().x()/PX_PER_FT, r.pos().y()/PX_PER_FT
        p = tx(rx_ft, ry_ft)
        name_u = r.name.upper()
        if zone_check(rx_ft, ry_ft):
            halo_behind(p.x-60, p.y-4, name_u, text_pt, "helv")
        page.insert_text(fitz.Point(p.x-60, p.y-4), name_u, fontsize=text_pt,
                          fontname="helv", color=(0.1,0.1,0.1))

    for p, col, tag, half_h, shp_key, in_zone in symbol_labels:
        if not tag:
            continue
        tag = tag.upper()
        tag_col = _legible_tag_color(col)
        if shp_key in TAG_ON_ICON_SHAPES:
            # Font size still follows the user's text_pt setting but is
            # capped to the icon's own size so a multi-line tag (e.g.
            # "F.D.\nKEY") can't blow out past a small icon like the key box.
            # No halo here — it already sits on the icon's own white/light
            # fill, and the icon itself already got a zone halo if needed.
            lines = tag.split("\n")
            fs = min(text_pt, max(8, half_h*1.1))
            start_y = p.y - (fs*1.15*(len(lines)-1))/2 + fs*0.35
            for i, line in enumerate(lines):
                tw = fitz.get_text_length(line, fontname="hebo", fontsize=fs)
                page.insert_text(fitz.Point(p.x-tw/2, start_y+i*fs*1.15), line,
                                  fontsize=fs, fontname="hebo", color=tag_col)
        elif shp_key in TAG_ABOVE_ICON_SHAPES:
            tw = fitz.get_text_length(tag, fontname="hebo", fontsize=text_pt)
            above_y = p.y-half_h-text_pt
            if in_zone:
                halo_behind(p.x-tw/2, above_y, tag, text_pt, "hebo")
            page.insert_text(fitz.Point(p.x-tw/2, above_y), tag, fontsize=text_pt,
                              fontname="hebo", color=tag_col)
        else:
            below_y = p.y + half_h + text_pt + 2
            tw = fitz.get_text_length(tag, fontname="hebo", fontsize=text_pt)
            if in_zone:
                halo_behind(p.x-tw/2, below_y, tag, text_pt, "hebo")
            page.insert_text(fitz.Point(p.x-tw/2, below_y), tag, fontsize=text_pt,
                              fontname="hebo", color=tag_col)

    # Title block
    tb_y = (ph_in-TITLE_BLOCK_H_IN)*72
    page.draw_line(fitz.Point(MARGIN_IN*72, tb_y), fitz.Point((pw_in-MARGIN_IN)*72, tb_y),
                    color=(0,0,0), width=1.0)
    page.insert_text(fitz.Point(MARGIN_IN*72, tb_y+18), sheet_title.upper(),
                      fontsize=15, fontname="helv", color=(0,0,0))
    info_lines = [
        f"PROJECT: {(project_meta.get('customer') or '').upper()}   {(project_meta.get('job_number') or '').upper()}",
        f"DATE: {datetime.date.today().strftime('%b %d, %Y').upper()}",
    ]
    for i, line in enumerate(info_lines):
        page.insert_text(fitz.Point(MARGIN_IN*72, tb_y+36+i*14), line, fontsize=9, color=(0.2,0.2,0.2))

    # On-page legend column (right side) — only Fire Safety Plan symbols
    # actually used, PLUS every fire zone actually drawn (colored swatch +
    # whatever the zone was named), arranged to fit whatever vertical room
    # is available rather than spilling onto a second page.
    if has_legend:
        col_x = (pw_in-MARGIN_IN-legend_w_in)*72
        divider_x = col_x - legend_gap_in*72/2
        div_shape = page.new_shape()
        div_shape.draw_line(fitz.Point(divider_x, MARGIN_IN*72), fitz.Point(divider_x, tb_y))
        div_shape.finish(color=(0.6,0.6,0.6), width=0.8)

        header_h = text_pt + 16
        row_y0 = MARGIN_IN*72 + header_h
        available_h = tb_y - row_y0 - 6
        # +1 extra row reserved for the "ZONES" sub-header when both symbol
        # and zone entries are present.
        n_rows = len(used_kinds) + len(legend_zones) + (1 if used_kinds and legend_zones else 0)
        row_h = max(text_pt+14, min(text_pt+26, available_h/max(1, n_rows)))

        legend_rows = []   # (row_y, tag_or_None, label_text, swatch_color)
        row_y = row_y0
        for k in used_kinds:
            defn = ALL_SYMBOL_DEFS[k]
            col_hex = defn.get("color") or LAYER_COLORS.get(defn["layer"], "#333333")
            lcol = tuple(int(col_hex[i:i+2],16)/255.0 for i in (1,3,5))
            icon_cx, icon_cy = col_x+14, row_y-2
            _draw_symbol_icon_pdf(div_shape, defn, icon_cx, icon_cy, lcol, px_per_ft=4.0, scale=1.3)
            tag = defn.get("tag") or "".join(w2[0] for w2 in defn["name"].split()[:2]).upper()
            legend_rows.append((row_y, tag.upper(), defn["name"].upper(), None))
            row_y += row_h

        zone_header_y = None
        if legend_zones:
            if used_kinds:
                zone_header_y = row_y
                row_y += row_h
            for z in legend_zones:
                zcol_hex = z.color
                zcol = tuple(int(zcol_hex[i:i+2],16)/255.0 for i in (1,3,5))
                sw = text_pt*0.9
                swatch_rect = fitz.Rect(col_x+14-sw/2, row_y-2-sw/2, col_x+14+sw/2, row_y-2+sw/2)
                z_pattern = getattr(z, "pattern", "solid")
                if z_pattern == "solid":
                    div_shape.draw_rect(swatch_rect)
                    div_shape.finish(color=zcol, fill=zcol, width=0.8)
                else:
                    # Same swatch-plus-hatch rendering as the main plan, at
                    # legend scale, so a color reused across zones still
                    # reads as different textures here too.
                    div_shape.draw_rect(swatch_rect)
                    div_shape.finish(color=zcol, width=0.8)
                    sw_pts = [(swatch_rect.x0,swatch_rect.y0),(swatch_rect.x1,swatch_rect.y0),
                              (swatch_rect.x1,swatch_rect.y1),(swatch_rect.x0,swatch_rect.y1)]
                    for hang, hspacing in ZONE_HATCH_SPECS.get(z_pattern, []):
                        for (x1,y1),(x2,y2) in hatch_lines_in_polygon(sw_pts, hang, max(2.5, hspacing*0.35)):
                            div_shape.draw_line(fitz.Point(x1,y1), fitz.Point(x2,y2))
                            div_shape.finish(color=zcol, width=0.6)
                legend_rows.append((row_y, None, z.name.upper(), zcol))
                row_y += row_h
        div_shape.commit()

        page.insert_text(fitz.Point(col_x, MARGIN_IN*72+text_pt+2), "LEGEND",
                          fontsize=text_pt+2, fontname="helv", color=(0,0,0))
        if zone_header_y is not None:
            page.insert_text(fitz.Point(col_x, zone_header_y), "ZONES",
                              fontsize=text_pt, fontname="helv", color=(0,0,0))
        legend_text_pt = max(8, text_pt*0.85)
        for row_y, tag, name, zcol in legend_rows:
            if tag is not None:
                page.insert_text(fitz.Point(col_x+30, row_y), f"{tag} — {name}",
                                  fontsize=legend_text_pt, color=(0.1,0.1,0.1))
            else:
                page.insert_text(fitz.Point(col_x+30, row_y), name,
                                  fontsize=legend_text_pt, color=zcol)

    doc.save(path)
    doc.close()


def _legible_tag_color(col):
    """A symbol's own accent color doubles as its PDF tag-text color, which
    reads fine for every color in the catalog except yellow (Electrical
    Vault/Hazard, Generator) — yellow text on the white page background is
    nearly invisible. Darken only that case; everything else passes through."""
    r, g, b = col
    if r > 0.85 and g > 0.65 and b < 0.3:
        return (0.25, 0.19, 0.0)
    return col


# Where a symbol's tag/label text is positioned relative to its icon —
# matches the official Calgary Fire Department symbol sheet exactly rather
# than a single generic convention: shapes that are literal text-carrying
# glyphs there (HAZ diamond, R triangle, FAAP/FACP/... boxes, the "E"/"W"/
# "O2" inside a circle or badge, "Gen" filling the generator box) get their
# tag centered ON the icon; the compass gets "N" printed above it; every
# other (purely pictorial) icon keeps the tag below it.
TAG_ON_ICON_SHAPES = {"diamond_text", "triangle_text", "textbox", "line_valve",
                       "bolt", "elec_vault", "gen_box", "helmet", "keybox"}
TAG_ABOVE_ICON_SHAPES = {"compass"}
TAG_BELOW_ICON_SHAPES = {"arrow", "pull_hand", "flag", "cylinder", "standpipe", "gate_valve",
                          "spray_head", "wall_break", "hose_reel", "wye", "hydrant", "cross_box",
                          "phone", "gas_meter", "gas_cyl_diamond", "horn", "you_are_here"}


SYMBOL_ICON_R_PT = 7.0   # fixed on-page radius for non-furniture symbol icons


def _draw_symbol_icon_pdf(shp, defn, cx, cy, col, px_per_ft, rotation_deg=0.0, scale=1.0):
    """Mirrors SymbolItem.paint()'s per-shape drawing (same shape cases,
    same relative proportions) so the exported plan shows the actual icon —
    outlet ticks, switch dot, light X, camera trapezoid, etc — instead of a
    generic circle. Furniture (rect/circle/cabinet) draws at true scale
    using px_per_ft, exactly like on screen with PX_PER_FT.

    `scale` uniformly enlarges the icon (both R and px_per_ft, which every
    shape case below is expressed in terms of) — used by the Fire Safety
    Plan export to make icons legible from a distance without touching the
    regular floor-plan export's sizing.

    Returns the icon's (unrotated) half-height in points, so the caller can
    place a tag/label below it without guessing the icon's extent."""
    shape = defn.get("shape", "rect")
    R = SYMBOL_ICON_R_PT * scale
    px_per_ft = px_per_ft * scale
    white = (1, 1, 1)
    ang = math.radians(rotation_deg)
    ca, sa = math.cos(ang), math.sin(ang)
    half_h = R
    if shape in ("rect", "cabinet", "stairs"):
        half_h = defn.get("d",24)/12.0*px_per_ft/2
    elif shape == "circle":
        half_h = defn.get("w",24)/12.0*px_per_ft/2
    elif shape == "panel":
        half_h = R*0.7
    elif shape == "rrect":
        half_h = R*0.65
    elif shape in ("textbox", "gen_box"):
        half_h = R*0.8
    elif shape == "rect_outline":
        half_h = R*0.7
    elif shape == "triangle_text":
        half_h = R*0.8
    elif shape == "wap":
        half_h = R + 4   # dot sits at +4, arcs extend to radius R above it
    elif shape == "gate_valve":
        half_h = R*0.95 if defn.get("pedestal") else R*0.6
    elif shape == "horn":
        half_h = R*1.05
    elif shape == "hose_reel":
        half_h = R*0.75 if defn.get("cabinet") else R*0.5
    elif shape == "compass":
        half_h = R*0.95
    elif shape == "gas_meter":
        half_h = R*1.15
    elif shape == "cylinder":
        half_h = R*1.05

    def pt(dx, dy):
        return fitz.Point(cx + dx*ca - dy*sa, cy + dx*sa + dy*ca)

    def poly_rect(w, d):
        pts = [pt(-w/2,-d/2), pt(w/2,-d/2), pt(w/2,d/2), pt(-w/2,d/2)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=white, width=1.0)

    if shape == "outlet":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
        shp.draw_line(pt(-4,-6), pt(-4,6)); shp.finish(color=col, width=1.0)
        shp.draw_line(pt(4,-6), pt(4,6)); shp.finish(color=col, width=1.0)
    elif shape == "outlet4":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
        for xo in (-5,-1.6,1.6,5):
            shp.draw_line(pt(xo,-6), pt(xo,6)); shp.finish(color=col, width=1.0)
    elif shape == "outlet_sq":
        poly_rect(R*2, R*2)
        shp.draw_line(pt(-4,-6), pt(-4,6)); shp.finish(color=col, width=1.0)
        shp.draw_line(pt(4,-6), pt(4,6)); shp.finish(color=col, width=1.0)
    elif shape == "switch":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
    elif shape == "light":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
        shp.draw_line(pt(-R,-R), pt(R,R)); shp.finish(color=col, width=1.0)
        shp.draw_line(pt(-R,R), pt(R,-R)); shp.finish(color=col, width=1.0)
    elif shape == "light_sm":
        shp.draw_circle(pt(0,0), R*0.6); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "panel":
        poly_rect(R*2, R*1.4)
        for i in range(1,4):
            xx = -R + i*R*2/4
            shp.draw_line(pt(xx,-R*0.7), pt(xx,R*0.7)); shp.finish(color=col, width=0.7)
    elif shape == "hex":
        pts = [pt(R*math.cos(math.radians(a)), R*math.sin(math.radians(a))) for a in range(0,360,60)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=white, width=1.0)
    elif shape == "wap":
        shp.draw_circle(pt(0,4), R*0.35); shp.finish(color=col, fill=col, width=1.0)
        # Two upward "signal" arcs above the dot — sampled as short line
        # segments rather than relying on draw_sector's angle convention.
        for rad in (R*0.6, R):
            pts = [pt(rad*math.sin(math.radians(a)), 4-rad*math.cos(math.radians(a)))
                   for a in range(-60, 61, 12)]
            shp.draw_polyline(pts); shp.finish(color=col, width=1.0)
    elif shape == "camera":
        pts = [pt(-R,-R*0.5), pt(R*0.3,-R*0.5), pt(R,-R), pt(R,R), pt(R*0.3,R*0.5), pt(-R,R*0.5)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=white, width=1.0)
    elif shape == "rrect":
        poly_rect(R*2, R*1.3)
    elif shape == "speaker":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
        shp.draw_circle(pt(0,0), R*0.3); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "circle":
        w = defn.get("w",24)/12.0*px_per_ft
        shp.draw_circle(pt(0,0), w/2); shp.finish(color=col, fill=white, width=1.0)
    elif shape == "rect":
        w = defn.get("w",24)/12.0*px_per_ft
        d = defn.get("d",24)/12.0*px_per_ft
        poly_rect(w, d)
    elif shape == "cabinet":
        w = defn.get("w",18)/12.0*px_per_ft
        d = defn.get("d",24)/12.0*px_per_ft
        poly_rect(w, d)
        shp.draw_line(pt(-w/2,0), pt(w/2,0)); shp.finish(color=col, width=1.0)
    elif shape == "stairs":
        w = defn.get("w",36)/12.0*px_per_ft
        d = defn.get("d",132)/12.0*px_per_ft
        poly_rect(w, d)
        n_treads = 10
        for i in range(1, n_treads):
            yy = -d/2 + i*d/n_treads
            shp.draw_line(pt(-w/2,yy), pt(w/2,yy)); shp.finish(color=col, width=0.6)
        going_up = defn.get("dir","up") == "up"
        tip_y, tail_y = (-d/2+6, d/2-6) if going_up else (d/2-6, -d/2+6)
        shp.draw_line(pt(0,tail_y), pt(0,tip_y)); shp.finish(color=col, width=1.0)
        ah = 7
        sign = -1 if going_up else 1
        shp.draw_line(pt(0,tip_y), pt(-ah/2,tip_y-sign*ah)); shp.finish(color=col, width=1.0)
        shp.draw_line(pt(0,tip_y), pt(ah/2,tip_y-sign*ah)); shp.finish(color=col, width=1.0)
    # ── Fire Safety Plan symbols (Calgary Fire Department set) ──────────────
    elif shape == "arrow":
        pts = [pt(-R,0), pt(R*0.4,-R*0.8), pt(R*0.4,-R*0.3), pt(R,-R*0.3),
               pt(R,R*0.3), pt(R*0.4,R*0.3), pt(R*0.4,R*0.8)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "pull_hand":
        box = [pt(-R*0.9,-R),pt(-R*0.5,-R),pt(-R*0.5,R),pt(-R*0.9,R)]
        shp.draw_polyline(box+[box[0]]); shp.finish(color=col, fill=col, width=1.0)
        pts = [pt(-R*0.5,-R*0.3)]
        for a in range(200, 341, 20):
            rad = math.radians(a)
            pts.append(pt(-R*0.1+R*0.35*math.cos(rad), -R*0.1+R*0.35*math.sin(rad)))
        pts.append(pt(R*0.6, R*0.5))
        shp.draw_polyline(pts); shp.finish(color=col, width=1.3)
    elif shape == "flag":
        bar = [pt(-R*0.75,-R),pt(-R*0.45,-R),pt(-R*0.45,R),pt(-R*0.75,R)]
        shp.draw_polyline(bar+[bar[0]]); shp.finish(color=col, fill=col, width=1.0)
        pennant = [pt(-R*0.45,-R*0.9),pt(R*0.65,-R*0.15),pt(R*0.1,R*0.35),pt(-R*0.45,R*0.55)]
        shp.draw_polyline(pennant+[pennant[0]]); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "cylinder":
        # Fire extinguisher: red body + black nozzle/handle assembly on top —
        # mirrors SymbolItem.paint()'s "cylinder" case.
        body = [pt(-R*0.5,-R*0.85),pt(R*0.5,-R*0.85),pt(R*0.5,R),pt(-R*0.5,R)]
        shp.draw_polyline(body+[body[0]]); shp.finish(color=col, fill=col, width=1.0)
        handle = [pt(-R*0.4,-R*1.05),pt(R*0.4,-R*1.05),pt(R*0.4,-R*0.87),pt(-R*0.4,-R*0.87)]
        shp.draw_polyline(handle+[handle[0]]); shp.finish(color=(0.1,0.1,0.1), fill=(0.1,0.1,0.1), width=1.0)
        shp.draw_line(pt(-R*0.55,-R*0.87), pt(-R*0.85,-R*0.55))
        shp.finish(color=(0.1,0.1,0.1), width=1.4)
    elif shape == "standpipe":
        body = [pt(-R*0.85,-R),pt(-R*0.15,-R),pt(-R*0.15,R*0.9),pt(-R*0.85,R*0.9)]
        shp.draw_polyline(body+[body[0]]); shp.finish(color=col, width=1.0)
        port = [pt(-R*0.15,-R*0.35),pt(R*0.4,-R*0.35),pt(R*0.4,R*0.15),pt(-R*0.15,R*0.15)]
        shp.draw_polyline(port+[port[0]]); shp.finish(color=col, width=1.0)
        shp.draw_line(pt(R*0.75,-R*0.5), pt(R*0.75,R*0.1)); shp.finish(color=col, width=1.6)
    elif shape == "gate_valve":
        for xo in (-R*0.75, R*0.75):
            vb = [pt(xo-R*0.25,-R*0.6),pt(xo+R*0.25,-R*0.6),pt(xo+R*0.25,R*0.6),pt(xo-R*0.25,R*0.6)]
            shp.draw_polyline(vb+[vb[0]]); shp.finish(color=col, fill=col, width=1.0)
        hb = [pt(-R*0.75,-R*0.22),pt(R*0.75,-R*0.22),pt(R*0.75,R*0.22),pt(-R*0.75,R*0.22)]
        shp.draw_polyline(hb+[hb[0]]); shp.finish(color=col, fill=col, width=1.0)
        shp.draw_circle(pt(0,0), R*0.3); shp.finish(color=col, fill=white, width=1.2)
        if defn.get("pedestal"):
            shp.draw_line(pt(0,R*0.6), pt(0,R*0.95)); shp.finish(color=col, width=2.0)
    elif shape == "elec_vault":
        sq = [pt(-R,-R),pt(R,-R),pt(R,R),pt(-R,R)]
        shp.draw_polyline(sq+[sq[0]]); shp.finish(color=(0.1,0.1,0.1), fill=(0.965,0.769,0.059), width=1.2)
        bolt = [pt(R*0.15,-R*0.9),pt(-R*0.3,R*0.05),pt(R*0.0,R*0.05),pt(-R*0.15,R*0.9),
                pt(R*0.3,-R*0.05),pt(R*0.0,-R*0.05)]
        shp.draw_polyline(bolt+[bolt[0]]); shp.finish(color=(0.1,0.1,0.1), fill=(0.1,0.1,0.1), width=0.5)
        shp.draw_circle(pt(0,0), R*0.35); shp.finish(color=(0.1,0.1,0.1), fill=(0.965,0.769,0.059), width=1.0)
    elif shape == "line_valve":
        shp.draw_line(pt(-R*1.3,0), pt(-R*0.9,0)); shp.finish(color=col, width=2.2)
        shp.draw_line(pt(R*0.9,0), pt(R*1.3,0)); shp.finish(color=col, width=2.2)
        shp.draw_circle(pt(0,0), R*0.9); shp.finish(color=col, fill=white, width=1.3)
    elif shape == "spray_head":
        nb = [pt(-R*0.25,-R),pt(R*0.25,-R),pt(R*0.25,-R*0.3),pt(-R*0.25,-R*0.3)]
        shp.draw_polyline(nb+[nb[0]]); shp.finish(color=col, fill=col, width=1.0)
        for a in (-30,-10,10,30):
            rad = math.radians(a)
            shp.draw_line(pt(0,-R*0.3), pt(R*0.9*math.sin(rad), R*0.8))
            shp.finish(color=col, width=1.0, dashes="[2 2] 0")
    elif shape in ("diamond_text","triangle_text","circle_text","textbox","rect_outline"):
        if shape == "diamond_text":
            pts = [pt(0,-R),pt(R,0),pt(0,R),pt(-R,0)]
            shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=white, width=1.0)
        elif shape == "triangle_text":
            pts = [pt(0,-R),pt(R,R*0.8),pt(-R,R*0.8)]
            shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=white, width=1.0)
        elif shape == "circle_text":
            shp.draw_circle(pt(0,0), R); shp.finish(color=col, fill=white, width=1.0)
        elif shape == "textbox":
            poly_rect(R*2.6, R*1.6)
        else:  # rect_outline
            pts = [pt(-R*1.1,-R*0.7),pt(R*1.1,-R*0.7),pt(R*1.1,R*0.7),pt(-R*1.1,R*0.7)]
            shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, width=1.0)
    elif shape == "gas_cyl_diamond":
        pts = [pt(0,-R),pt(R,0),pt(0,R),pt(-R,0)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
        cyl = [pt(-R*0.16,-R*0.55),pt(R*0.16,-R*0.55),pt(R*0.16,R*0.5),pt(-R*0.16,R*0.5)]
        shp.draw_polyline(cyl+[cyl[0]]); shp.finish(color=col, fill=white, width=0.8)
    elif shape == "gen_box":
        rect = [pt(-R*1.3,-R*0.8),pt(R*1.3,-R*0.8),pt(R*1.3,R*0.8),pt(-R*1.3,R*0.8)]
        shp.draw_polyline(rect+[rect[0]]); shp.finish(color=(0.1,0.1,0.1), fill=(0.965,0.769,0.059), width=1.2)
    elif shape == "bolt":
        pts = [pt(R*0.15,-R),pt(-R*0.4,R*0.1),pt(0,R*0.1),pt(-R*0.15,R),pt(R*0.4,-R*0.1),pt(0,-R*0.1)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
        shp.draw_circle(pt(0,0), R*0.32); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "trefoil":
        sq = [pt(-R,-R),pt(R,-R),pt(R,R),pt(-R,R)]
        shp.draw_polyline(sq+[sq[0]]); shp.finish(color=(0.1,0.1,0.1), fill=(0.1,0.1,0.1), width=1.0)
        shp.draw_circle(pt(0,0), R*0.95); shp.finish(color=(0.1,0.1,0.1), fill=white, width=0.8)
        for base_ang in (90, 210, 330):
            wedge = [pt(0,0)]
            for a in range(base_ang-30, base_ang+31, 10):
                rad = math.radians(a)
                wedge.append(pt(R*0.75*math.cos(rad), R*0.75*math.sin(rad)))
            wedge.append(pt(0,0))
            shp.draw_polyline(wedge); shp.finish(color=(0.1,0.1,0.1), fill=(0.1,0.1,0.1), width=0.5)
        shp.draw_circle(pt(0,0), R*0.18); shp.finish(color=(0.1,0.1,0.1), fill=(0.1,0.1,0.1), width=1.0)
    elif shape == "horn":
        bar = [pt(-R*0.9,-R*0.35),pt(-R*0.6,-R*0.35),pt(-R*0.6,R*0.35),pt(-R*0.9,R*0.35)]
        shp.draw_polyline(bar+[bar[0]]); shp.finish(color=col, fill=col, width=1.0)
        for rad in (R*0.35, R*0.7, R*1.05):
            pts = [pt(-R*0.5+rad*math.cos(math.radians(a)), rad*math.sin(math.radians(a)))
                   for a in range(-50, 51, 10)]
            shp.draw_polyline(pts); shp.finish(color=col, width=1.1)
    elif shape == "wall_break":
        shp.draw_line(pt(-R,0), pt(-R*0.3,0)); shp.finish(color=col, width=1.2)
        shp.draw_line(pt(R*0.3,0), pt(R,0)); shp.finish(color=col, width=1.2)
        arc_pts = [pt(R*0.3*math.cos(math.radians(a)), R*0.3*math.sin(math.radians(a))) for a in range(0,181,20)]
        shp.draw_polyline(arc_pts); shp.finish(color=col, width=1.0)
    elif shape == "hose_reel":
        if defn.get("cabinet"):
            box = [pt(-R,-R*0.75),pt(R,-R*0.75),pt(R,R*0.75),pt(-R,R*0.75)]
            shp.draw_polyline(box+[box[0]]); shp.finish(color=col, width=1.2)
        zig = [pt(-R*0.8,R*0.4)]
        for i in range(1,6):
            yx = R*0.4 if i % 2 == 0 else -R*0.35
            zig.append(pt(-R*0.8+i*R*0.28, yx))
        shp.draw_polyline(zig); shp.finish(color=col, width=1.6)
        noz = [pt(R*0.55,-R*0.5),pt(R*0.85,-R*0.15),pt(R*0.55,0)]
        shp.draw_polyline(noz+[noz[0]]); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "wye":
        shp.draw_line(pt(0,R), pt(0,0)); shp.finish(color=col, width=1.4)
        shp.draw_line(pt(0,0), pt(-R*0.7,-R)); shp.finish(color=col, width=1.4)
        shp.draw_line(pt(0,0), pt(R*0.7,-R)); shp.finish(color=col, width=1.4)
    elif shape == "hydrant":
        cap_hex = defn.get("cap_color")
        cap_col = tuple(int(cap_hex[i:i+2],16)/255.0 for i in (1,3,5)) if cap_hex else col
        poly_body = [pt(-R*0.4,-R*0.2),pt(R*0.4,-R*0.2),pt(R*0.4,R),pt(-R*0.4,R)]
        shp.draw_polyline(poly_body+[poly_body[0]])
        shp.finish(color=col, fill=col, width=1.0)
        shp.draw_line(pt(-R*0.6,0), pt(-R*0.9,0)); shp.finish(color=col, width=1.2)
        shp.draw_line(pt(R*0.6,0), pt(R*0.9,0)); shp.finish(color=col, width=1.2)
        shp.draw_circle(pt(0,-R*0.5), R*0.35); shp.finish(color=col, fill=cap_col, width=1.0)
    elif shape == "compass":
        shp.draw_circle(pt(0,R*0.1), R*0.75); shp.finish(color=(0.7,0.7,0.7), width=0.8)
        shp.draw_line(pt(-R*0.75,R*0.1), pt(R*0.75,R*0.1)); shp.finish(color=(0.7,0.7,0.7), width=0.6)
        pts = [pt(0,-R*0.85),pt(R*0.22,R*0.35),pt(0,R*0.15),pt(-R*0.22,R*0.35)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
        shp.draw_line(pt(0,R*0.35), pt(0,R*0.95)); shp.finish(color=col, width=1.5)
    elif shape == "cross_box":
        h1 = [pt(-R*0.15,-R*0.6),pt(R*0.15,-R*0.6),pt(R*0.15,R*0.6),pt(-R*0.15,R*0.6)]
        shp.draw_polyline(h1+[h1[0]]); shp.finish(color=col, fill=col, width=0.5)
        h2 = [pt(-R*0.6,-R*0.15),pt(R*0.6,-R*0.15),pt(R*0.6,R*0.15),pt(-R*0.6,R*0.15)]
        shp.draw_polyline(h2+[h2[0]]); shp.finish(color=col, fill=col, width=0.5)
    elif shape == "phone":
        pts = [pt(-R*0.7,-R*0.9),pt(-R*0.2,-R*0.5),pt(-R*0.35,-R*0.25),pt(R*0.05,R*0.15),
               pt(R*0.35,-R*0.05),pt(R*0.75,R*0.35),pt(R*0.45,R*0.75),pt(R*0.0,R*0.4),
               pt(-R*0.55,-R*0.15),pt(-R*0.9,-R*0.55)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
    elif shape == "gas_meter":
        body = [pt(-R*0.8,-R*0.55),pt(R*0.8,-R*0.55),pt(R*0.8,R*0.55),pt(-R*0.8,R*0.55)]
        shp.draw_polyline(body+[body[0]]); shp.finish(color=col, fill=col, width=1.0)
        shp.draw_line(pt(-R*0.6,R*0.55), pt(-R*0.6,R)); shp.finish(color=col, width=2.0)
        shp.draw_line(pt(R*0.6,R*0.55), pt(R*0.6,R)); shp.finish(color=col, width=2.0)
        shp.draw_line(pt(-R*0.3,-R*0.55), pt(R*0.15,-R*1.15)); shp.finish(color=col, width=2.2)
    elif shape == "keybox":
        shp.draw_circle(pt(0,0), R); shp.finish(color=col, width=0.6)
    elif shape == "helmet":
        pts = [pt(-R,R*0.3),pt(-R*0.9,-R*0.1),pt(-R*0.5,-R*0.6),pt(0,-R*0.75),pt(R*0.5,-R*0.5),
               pt(R*0.85,0),pt(R,R*0.3),pt(R*0.6,R*0.45),pt(-R*0.6,R*0.45)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
        shp.draw_circle(pt(0,-R*0.05), R*0.4); shp.finish(color=(0.35,0.35,0.35), fill=(0.83,0.83,0.83), width=1.0)
    elif shape == "you_are_here":
        pts = [pt(0,-R*1.3),pt(R*0.9,R*0.2),pt(R*0.4,R*0.2),pt(R*0.4,R*1.1),
               pt(-R*0.4,R*1.1),pt(-R*0.4,R*0.2),pt(-R*0.9,R*0.2)]
        shp.draw_polyline(pts+[pts[0]]); shp.finish(color=col, fill=col, width=1.0)
        half_h = R*1.3
    else:
        poly_rect(R*2, R*2)

    return half_h


def _append_symbol_legend_page(doc, scene, project_meta, sheet_title, visible):
    """One or more Letter-size pages listing every symbol kind actually
    placed on the plan (and visible), grouped by category, so a contractor
    can read the tag/circle markers on the plan without this app open."""
    used_kinds = sorted({s.kind for s in scene._symbol_items if visible(s)},
                         key=lambda k: ALL_SYMBOL_DEFS.get(k, {}).get("name", k))
    if not used_kinds:
        return
    used_set = set(used_kinds)

    pw, ph = 8.5*72, 11*72
    margin = 36
    row_h = 20
    state = {"page": None, "y": 0.0}

    def new_page():
        state["page"] = doc.new_page(width=pw, height=ph)
        state["page"].insert_text(fitz.Point(margin, margin+14), "SYMBOL LEGEND",
                                   fontsize=13, fontname="helv", color=(0,0,0))
        state["page"].insert_text(
            fitz.Point(margin, margin+30),
            f"Project: {project_meta.get('customer','')}   {project_meta.get('job_number','')}   "
            f"—   {sheet_title}",
            fontsize=8, color=(0.2,0.2,0.2))
        state["y"] = margin + 54

    new_page()
    for cat_name, kinds in SYMBOL_CATEGORIES:
        cat_kinds = [k for k in kinds if k in used_set]
        if not cat_kinds:
            continue
        if state["y"] > ph - margin - 30:
            new_page()
        state["page"].insert_text(fitz.Point(margin, state["y"]), cat_name,
                                   fontsize=10, fontname="helv", color=(0,0,0))
        state["y"] += row_h
        for k in sorted(cat_kinds, key=lambda kk: ALL_SYMBOL_DEFS[kk]["name"]):
            if state["y"] > ph - margin - 20:
                new_page()
            defn = ALL_SYMBOL_DEFS[k]
            col_hex = defn.get("color") or LAYER_COLORS.get(defn["layer"], "#333333")
            col = tuple(int(col_hex[i:i+2], 16)/255.0 for i in (1,3,5))
            cx, cy = margin + 14, state["y"] - 4
            swatch = state["page"].new_shape()
            # Fixed small representative scale for furniture (rect/circle/
            # cabinet) icons in the legend — a real px_per_ft would make a
            # desk swatch huge and a chair swatch tiny in the same row.
            _draw_symbol_icon_pdf(swatch, defn, cx, cy, col, px_per_ft=3.0)
            swatch.commit()
            tag = defn.get("tag") or "".join(w[0] for w in defn["name"].split()[:2]).upper()
            state["page"].insert_text(fitz.Point(margin+28, state["y"]), tag,
                                       fontsize=7.5, fontname="helv", color=_legible_tag_color(col))
            state["page"].insert_text(fitz.Point(margin+70, state["y"]), defn["name"],
                                       fontsize=8.5, color=(0.1,0.1,0.1))
            # Multi-line tags (e.g. "F.D.\nKEY") need extra row height or the
            # next row's text overlaps the tag's second line.
            state["y"] += row_h + tag.count("\n")*10

        state["y"] += 8


def visible_layer_any(scene, layer):
    return scene.layers_visible.get(layer, True)


def _draw_opening_pdf(shape, page, o, tx):
    """Mirrors DoorWindowItem.paint() exactly (same _leaves()/arc_points()
    geometry, just in feet instead of px) so the PDF matches the canvas."""
    ang = math.radians(o.rotation())
    ca, sa = math.cos(ang), math.sin(ang)
    def rot(lx, ly):
        return (lx*ca-ly*sa, lx*sa+ly*ca)
    ox, oy = o.pos().x()/PX_PER_FT, o.pos().y()/PX_PER_FT
    w = o.width_in/12.0
    col = (0.17,0.24,0.31)

    def P(lx, ly):
        rx, ry = rot(lx, ly)
        return tx(ox+rx, oy+ry)

    shape_key = o.defn.get("shape")
    if shape_key == "window":
        shape.draw_line(P(-w/2,-0.02), P(w/2,-0.02)); shape.finish(color=col, width=0.8)
        shape.draw_line(P(-w/2, 0.02), P(w/2, 0.02)); shape.finish(color=col, width=0.8)
        return
    for hinge, tip, jamb in o._leaves(w):
        shape.draw_line(P(*hinge), P(*tip)); shape.finish(color=col, width=0.8)
        pts = [P(x,y) for x,y in arc_points(hinge, tip, jamb)]
        shape.draw_polyline(pts); shape.finish(color=col, width=0.8)


class PrintExportDialog(QDialog):
    """Choose paper size / orientation / scale (auto-fit or manual override)
    and whole-drawing vs. current-selection region, then export to PDF."""
    def __init__(self, scene, parent=None):
        super().__init__(parent)
        self.scene = scene
        self.setWindowTitle("Export Floor Plan to PDF"); self.setMinimumWidth(380)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)

        # Region
        has_sel = bool(scene.selectedItems())
        self.region_combo = QComboBox()
        self.region_combo.addItem("Whole drawing", "all")
        self.region_combo.addItem("Current selection only" + ("" if has_sel else "  (nothing selected)"), "selection")
        if not has_sel:
            self.region_combo.model().item(1).setEnabled(False)
        l.addRow("Print:", self.region_combo)

        # Paper
        self.paper_combo = QComboBox()
        for k in PAPER_SIZES_IN:
            self.paper_combo.addItem(k, k)
        self.paper_combo.setCurrentText("ARCH D (24x36)" if "ARCH D (24x36)" in PAPER_SIZES_IN else list(PAPER_SIZES_IN)[0])
        l.addRow("Paper size:", self.paper_combo)

        self.orient_combo = QComboBox()
        self.orient_combo.addItem("Auto (best fit)", "auto")
        self.orient_combo.addItem("Landscape", "landscape")
        self.orient_combo.addItem("Portrait", "portrait")
        l.addRow("Orientation:", self.orient_combo)

        # Scale
        self.auto_scale_rb = QRadioButton("Auto-scale to fit the page (Recommended)")
        self.manual_scale_rb = QRadioButton("Use this scale:")
        self.auto_scale_rb.setChecked(True)
        grp = QButtonGroup(self); grp.addButton(self.auto_scale_rb); grp.addButton(self.manual_scale_rb)
        l.addRow(self.auto_scale_rb)
        self.scale_combo = QComboBox()
        for lbl, fpi in SCALE_OPTIONS:
            self.scale_combo.addItem(lbl, fpi)
        self.scale_combo.setEnabled(False)
        self.manual_scale_rb.toggled.connect(self.scale_combo.setEnabled)
        row = QHBoxLayout(); row.addWidget(self.manual_scale_rb); row.addWidget(self.scale_combo)
        l.addRow(row)

        self.fit_note = QLabel(""); self.fit_note.setWordWrap(True)
        self.fit_note.setStyleSheet("color:#888;font-size:10px;")
        l.addRow(self.fit_note)

        for w in (self.region_combo, self.paper_combo, self.orient_combo, self.scale_combo):
            w.currentIndexChanged.connect(self._refresh_note)
        self.auto_scale_rb.toggled.connect(self._refresh_note)
        self._refresh_note()

        br = QHBoxLayout()
        ok = QPushButton("Export PDF")
        ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def _region_bbox(self):
        if self.region_combo.currentData() == "selection":
            items = self.scene.selectedItems()
            if items:
                xs, ys = [], []
                for it in items:
                    if hasattr(it, "x1"):
                        xs += [it.x1, it.x2]; ys += [it.y1, it.y2]
                    else:
                        xs.append(it.pos().x()/PX_PER_FT); ys.append(it.pos().y()/PX_PER_FT)
                if xs:
                    return (min(xs), min(ys), max(xs), max(ys))
        return self.scene.all_items_bbox_ft()

    def _resolve_paper_orientation(self, width_ft, height_ft):
        pw_in, ph_in = PAPER_SIZES_IN[self.paper_combo.currentData()]
        orient = self.orient_combo.currentData()
        if orient == "auto":
            fpi_land = best_fit_scale(width_ft, height_ft, max(pw_in,ph_in), min(pw_in,ph_in))
            fpi_port = best_fit_scale(width_ft, height_ft, min(pw_in,ph_in), max(pw_in,ph_in))
            orient = "landscape" if fpi_land <= fpi_port else "portrait"
        return orient

    def _refresh_note(self):
        try:
            min_x,min_y,max_x,max_y = self._region_bbox()
            width_ft, height_ft = max_x-min_x, max_y-min_y
            paper_key = self.paper_combo.currentData()
            pw_in, ph_in = PAPER_SIZES_IN[paper_key]
            orient = self._resolve_paper_orientation(width_ft, height_ft)
            if orient == "landscape" and pw_in < ph_in: pw_in, ph_in = ph_in, pw_in
            if orient == "portrait" and pw_in > ph_in: pw_in, ph_in = ph_in, pw_in
            if self.auto_scale_rb.isChecked():
                fpi = best_fit_scale(width_ft, height_ft, pw_in, ph_in - (TITLE_BLOCK_H_IN-MARGIN_IN))
                lbl = next((l for l,f in SCALE_OPTIONS if abs(f-fpi)<1e-6), f"1\"={fpi:.1f}'")
                self.fit_note.setText(f"Content is {width_ft:.0f}' × {height_ft:.0f}'. "
                                       f"Best-fit scale: {lbl} ({orient}).")
            else:
                fpi = self.scale_combo.currentData()
                ok = fits_at_scale(width_ft, height_ft, pw_in, ph_in - (TITLE_BLOCK_H_IN-MARGIN_IN), fpi)
                if ok:
                    self.fit_note.setText(f"Fits on one {orient} sheet at this scale.")
                    self.fit_note.setStyleSheet("color:#27ae60;font-size:10px;")
                else:
                    self.fit_note.setText("⚠ Content is larger than the page at this scale — "
                                           "it will be centered and may extend past the margins. "
                                           "Choose a smaller scale, larger paper, or a selection region.")
                    self.fit_note.setStyleSheet("color:#c0392b;font-size:10px;")
                    return
            self.fit_note.setStyleSheet("color:#888;font-size:10px;")
        except Exception:
            pass

    def values(self):
        min_x,min_y,max_x,max_y = self._region_bbox()
        width_ft, height_ft = max_x-min_x, max_y-min_y
        orient = self._resolve_paper_orientation(width_ft, height_ft)
        paper_key = self.paper_combo.currentData()
        pw_in, ph_in = PAPER_SIZES_IN[paper_key]
        if orient == "landscape" and pw_in < ph_in: pw_in, ph_in = ph_in, pw_in
        if orient == "portrait" and pw_in > ph_in: pw_in, ph_in = ph_in, pw_in
        if self.auto_scale_rb.isChecked():
            fpi = best_fit_scale(width_ft, height_ft, pw_in, ph_in - (TITLE_BLOCK_H_IN-MARGIN_IN))
        else:
            fpi = self.scale_combo.currentData()
        region_bbox = (min_x,min_y,max_x,max_y) if self.region_combo.currentData()=="selection" else None
        return {"paper_key": paper_key, "orientation": orient, "feet_per_inch": fpi,
                "region_bbox": region_bbox}


class FireSafetyPlanExportDialog(QDialog):
    """Export settings for the dedicated Fire Safety Plan sheet. Deliberately
    simpler than PrintExportDialog — always one auto-fit page (room for the
    on-page legend is reserved automatically, so there's no scale/region
    choice to make) — but adds the two controls that actually matter for a
    plan posted on a wall: how big the symbol icons and label text print."""
    def __init__(self, scene, parent=None):
        super().__init__(parent)
        self.scene = scene
        self.setWindowTitle("Export Fire Safety Plan"); self.setMinimumWidth(360)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)

        self.paper_combo = QComboBox()
        for k in PAPER_SIZES_IN:
            self.paper_combo.addItem(k, k)
        self.paper_combo.setCurrentText("ARCH D (24x36)" if "ARCH D (24x36)" in PAPER_SIZES_IN
                                         else list(PAPER_SIZES_IN)[0])
        l.addRow("Paper size:", self.paper_combo)

        self.orient_combo = QComboBox()
        self.orient_combo.addItem("Auto (best fit)", "auto")
        self.orient_combo.addItem("Landscape", "landscape")
        self.orient_combo.addItem("Portrait", "portrait")
        l.addRow("Orientation:", self.orient_combo)

        self.icon_scale_spin = QDoubleSpinBox()
        self.icon_scale_spin.setRange(1.0, 4.0); self.icon_scale_spin.setSingleStep(0.1)
        self.icon_scale_spin.setValue(2.0); self.icon_scale_spin.setSuffix("×")
        l.addRow("Symbol icon size:", self.icon_scale_spin)

        self.text_pt_spin = QSpinBox()
        # Floor of 8pt keeps every label at/above the 0.10" (~7.2pt) minimum
        # text height called for on posted Fire Safety Plans; raised the
        # ceiling to 48pt so tags/room names can go large enough to read
        # from across a room, not just up close.
        self.text_pt_spin.setRange(8, 48); self.text_pt_spin.setValue(14)
        self.text_pt_spin.setSuffix(" pt")
        self.text_pt_spin.setToolTip("Code minimum is 0.10\" (~7.2pt) text height — this floors at 8pt.")
        l.addRow("Label text size:", self.text_pt_spin)

        note = QLabel("Always fits on one page — the on-page legend column's width is "
                       "reserved automatically before the plan is scaled to fit.")
        note.setWordWrap(True); note.setStyleSheet("color:#888;font-size:10px;")
        l.addRow(note)

        br = QHBoxLayout()
        ok = QPushButton("Export PDF")
        ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        min_x,min_y,max_x,max_y = self.scene.all_items_bbox_ft()
        width_ft, height_ft = max(0.1,max_x-min_x), max(0.1,max_y-min_y)
        paper_key = self.paper_combo.currentData()
        pw_in, ph_in = PAPER_SIZES_IN[paper_key]
        orient = self.orient_combo.currentData()
        if orient == "auto":
            fpi_land = best_fit_scale(width_ft, height_ft, max(pw_in,ph_in), min(pw_in,ph_in))
            fpi_port = best_fit_scale(width_ft, height_ft, min(pw_in,ph_in), max(pw_in,ph_in))
            orient = "landscape" if fpi_land <= fpi_port else "portrait"
        return {"paper_key": paper_key, "orientation": orient,
                "symbol_scale": self.icon_scale_spin.value(),
                "text_pt": self.text_pt_spin.value()}


# ═══════════════════════════════════════════════════════════════════════════════
#  Wiring Diagram — freeform (not-to-scale) schematic: devices + connection lines
# ═══════════════════════════════════════════════════════════════════════════════

class WireConnectionItem(QGraphicsItem):
    """A connector line between two device symbols, tracking their positions
    live so dragging a device keeps its wires attached."""
    ITEM_TYPE = "wire"

    def __init__(self, a, b, label=""):
        super().__init__()
        self.a = a; self.b = b
        self.label = label
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setZValue(0.5)

    def boundingRect(self):
        pa, pb = self.a.scenePos(), self.b.scenePos()
        pad = 20
        return QRectF(min(pa.x(),pb.x())-pad, min(pa.y(),pb.y())-pad,
                      abs(pb.x()-pa.x())+pad*2, abs(pb.y()-pa.y())+pad*2)

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        pa, pb = self.a.scenePos(), self.b.scenePos()
        col = QColor("#ff7002") if self.isSelected() else QColor("#444")
        painter.setPen(QPen(col, 1.8))
        painter.drawLine(pa, pb)
        if self.label:
            mid = QPointF((pa.x()+pb.x())/2, (pa.y()+pb.y())/2)
            painter.setFont(QFont("Arial", 7, QFont.Bold))
            painter.setPen(col)
            painter.drawText(QRectF(mid.x()-40, mid.y()-14, 80, 14), Qt.AlignCenter, self.label)

    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_a = menu.addAction("Edit Label…")
        del_a = menu.addAction("Delete Wire")
        chosen = menu.exec_(event.screenPos())
        if chosen == edit_a:
            text, ok = QInputDialog.getText(None, "Wire Label", "Label (circuit #, wire type, etc.):", text=self.label)
            if ok:
                self.label = text.strip(); self.update()
        elif chosen == del_a and self.scene():
            self.scene().remove_wire(self)


class WiringScene(QGraphicsScene):
    """Not-to-scale schematic canvas: place device symbols, click-click to
    connect two devices with a labeled wire."""
    layout_changed = pyqtSignal()
    status_changed = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setSceneRect(-4000, -4000, 8000, 8000)
        self.setBackgroundBrush(QBrush(QColor("white")))
        self._devices = []
        self._wires = []
        self._mode = "select"
        self._pending_symbol = None
        self._connect_first = None
        self._uid_counter = 0

    def set_mode_select(self):
        self._mode = "select"; self._pending_symbol = None; self._connect_first = None

    def set_mode_symbol(self, kind):
        self._mode = "symbol"; self._pending_symbol = kind; self._connect_first = None
        name = ALL_SYMBOL_DEFS.get(kind, {}).get("name", kind)
        self.status_changed.emit(f"  Placing device: {name} — click on the canvas.")

    def set_mode_connect(self):
        self._mode = "connect"; self._pending_symbol = None; self._connect_first = None
        self.status_changed.emit("  Connect tool — click a device, then click a second device to wire them.")

    def _next_uid(self):
        self._uid_counter += 1
        return self._uid_counter

    def mousePressEvent(self, event):
        if self._mode == "symbol" and event.button() == Qt.LeftButton:
            kind = self._pending_symbol
            item = SymbolItem(kind, event.scenePos().x(), event.scenePos().y())
            item.uid = self._next_uid()
            self.addItem(item); self._devices.append(item)
            self.layout_changed.emit()
            return
        if self._mode == "connect" and event.button() == Qt.LeftButton:
            # items(QPointF) doesn't need a view/transform, unlike itemAt() —
            # robust even before the canvas has been shown/laid out.
            hits = [i for i in self.items(event.scenePos()) if i in self._devices]
            dev = hits[0] if hits else None
            if dev is None:
                return
            if self._connect_first is None:
                self._connect_first = dev
                self.status_changed.emit("  First device selected — click the device to connect it to.")
            elif dev is not self._connect_first:
                wire = WireConnectionItem(self._connect_first, dev)
                self.addItem(wire); self._wires.append(wire)
                self._connect_first = None
                self.layout_changed.emit()
                self.status_changed.emit("  Connected. Click another device to start a new wire, or Esc to stop.")
            return
        super().mousePressEvent(event)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.set_mode_select(); return
        if event.key() in (Qt.Key_Delete, Qt.Key_Backspace) and self.selectedItems():
            for it in list(self.selectedItems()):
                self.remove_item(it)
            return
        super().keyPressEvent(event)

    def remove_item(self, it):
        """Remove a device (and its wires) or a wire. Mirrors FloorPlanScene's
        remove_item() so SymbolItem's context-menu Delete works in both scenes."""
        if it in self._devices:
            for w in [w for w in self._wires if w.a is it or w.b is it]:
                self.remove_wire(w)
            self._devices.remove(it); self.removeItem(it)
        elif it in self._wires:
            self.remove_wire(it)
        self.layout_changed.emit()

    def remove_wire(self, w):
        if w in self._wires:
            self._wires.remove(w)
        self.removeItem(w)

    def clear_all(self):
        for w in list(self._wires):
            self.removeItem(w)
        self._wires.clear()
        for d in list(self._devices):
            self.removeItem(d)
        self._devices.clear()
        self.layout_changed.emit()

    def to_dict(self):
        return {
            "devices": [dict(d.to_dict(), uid=d.uid) for d in self._devices],
            "wires": [{"a_uid":w.a.uid, "b_uid":w.b.uid, "label":w.label} for w in self._wires],
        }

    def load_dict(self, d):
        self.clear_all()
        uid_map = {}
        for dd in d.get("devices", []):
            item = SymbolItem.from_dict(dd)
            item.uid = dd.get("uid", self._next_uid())
            self._uid_counter = max(self._uid_counter, item.uid)
            self.addItem(item); self._devices.append(item)
            uid_map[item.uid] = item
        for wd in d.get("wires", []):
            a = uid_map.get(wd.get("a_uid")); b = uid_map.get(wd.get("b_uid"))
            if a and b:
                wire = WireConnectionItem(a, b, wd.get("label",""))
                self.addItem(wire); self._wires.append(wire)
        self.layout_changed.emit()


class WiringCanvas(QGraphicsView):
    def __init__(self, scene):
        super().__init__(scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setDragMode(QGraphicsView.RubberBandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self._panning = False; self._pan_start = None

    def wheelEvent(self, e):
        f = 1.15 if e.angleDelta().y() > 0 else 1/1.15
        self.scale(f, f)

    def mousePressEvent(self, e):
        if e.button() == Qt.MiddleButton or (e.button()==Qt.LeftButton and (e.modifiers()&Qt.ControlModifier)):
            self._panning = True; self._pan_start = e.pos(); self.setCursor(Qt.ClosedHandCursor); return
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self._panning and self._pan_start:
            d = e.pos()-self._pan_start; self._pan_start = e.pos()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value()-d.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value()-d.y()); return
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        if self._panning:
            self._panning = False; self.setCursor(Qt.ArrowCursor)
        super().mouseReleaseEvent(e)

    def fit_all(self):
        items = self.scene().items()
        if not items:
            return
        r = items[0].sceneBoundingRect()
        for i in items[1:]:
            r = r.united(i.sceneBoundingRect())
        self.fitInView(r.adjusted(-60,-60,60,60), Qt.KeepAspectRatio)


# ═══════════════════════════════════════════════════════════════════════════════
#  One-Line Diagram — fire alarm panel / SLC loops / NAC circuits / boosters,
#  with live device-load tallying and booster-capacity planning. Auto-laid-out
#  tree (not user-positioned like the floor plan / wiring canvases).
# ═══════════════════════════════════════════════════════════════════════════════

OL_NODE_W = 220
OL_H_PANEL = 66
OL_H_CIRCUIT_BASE = 96   # header + load line + terminus line, before device-grid rows
OL_H_BOOSTER = 66
OL_LEVEL_GAP = 50        # vertical gap between a node's bottom edge and its children's top
OL_SIB_GAP = 24
OL_TICK_W, OL_TICK_H = 28, 16
OL_MIN_W, OL_MIN_H = 140, 50
OL_GRIP = 10             # resize-grip hit zone, bottom-right corner


class OneLineNodeBase(QGraphicsItem):
    """Shared tree-node behavior: children list, parent ref, subtree layout
    width, and free drag/resize. Nodes are laid out automatically only when
    first created (or on an explicit Auto-Arrange); once dragged or resized
    by the user, that override sticks until Auto-Arrange is used again."""
    def __init__(self):
        super().__init__()
        self.children = []
        self.parent_node = None
        self._w = OL_NODE_W
        self._h = OL_H_CIRCUIT_BASE
        self.manual_pos = False
        self.manual_size = False
        self._resizing = False
        self._resize_start = None
        self._press_pos = None
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setFlag(QGraphicsItem.ItemIsMovable)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges)
        self.setAcceptHoverEvents(True)

    def boundingRect(self):
        return QRectF(-4, -4, self._w+8, self._h+8)

    def subtree_width(self):
        if not self.children:
            return self._w
        total = sum(c.subtree_width() for c in self.children) + OL_SIB_GAP*(len(self.children)-1)
        return max(total, self._w)

    def has_visible_parent(self):
        """False when there's nothing drawn upstream to connect a line to —
        either a true standalone node, or one whose parent is hidden."""
        return self.parent_node is not None and not getattr(self.parent_node, "hidden", False)

    def _in_grip(self, pos):
        return self._w-OL_GRIP <= pos.x() <= self._w+4 and self._h-OL_GRIP <= pos.y() <= self._h+4

    def hoverMoveEvent(self, event):
        self.setCursor(Qt.SizeFDiagCursor if self._in_grip(event.pos()) else Qt.ArrowCursor)
        super().hoverMoveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._in_grip(event.pos()):
            self._resizing = True
            self._resize_start = (event.scenePos(), self._w, self._h)
            event.accept()
            return
        self._press_pos = self.pos()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._resizing:
            start_pos, start_w, start_h = self._resize_start
            delta = event.scenePos() - start_pos
            self.prepareGeometryChange()
            self._w = max(OL_MIN_W, start_w + delta.x())
            self._h = max(OL_MIN_H, start_h + delta.y())
            self.manual_size = True
            self.update()
            sc = self.scene()
            if sc:
                sc.update_connectors()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._resizing:
            self._resizing = False
            event.accept()
            return
        super().mouseReleaseEvent(event)
        if self._press_pos is not None and self.pos() != self._press_pos:
            self.manual_pos = True
        self._press_pos = None

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            # Live connector-line follow while dragging/auto-arranging.
            # Does NOT set manual_pos — that's decided at mouseReleaseEvent
            # (an actual user drag) or explicitly by auto_arrange()/load_dict,
            # so a programmatic setPos() here doesn't fight those callers.
            sc = self.scene()
            if sc:
                sc.update_connectors()
        return super().itemChange(change, value)


class PanelNode(OneLineNodeBase):
    ITEM_TYPE = "fa_panel"

    def __init__(self, name="FACP", nac_budget_ma=3000.0):
        super().__init__()
        self.name = name
        self.nac_budget_ma = nac_budget_ma
        self.hidden = False
        self._h = OL_H_PANEL
        self.setZValue(2)

    def direct_nac_load_ma(self):
        return sum(c.total_load() for c in self.children
                   if getattr(c, "circuit_type", None) == "nac")

    def paint(self, painter, option, widget=None):
        if self.hidden:
            return
        painter.setRenderHint(QPainter.Antialiasing)
        load = self.direct_nac_load_ma()
        over = self.nac_budget_ma > 0 and load > self.nac_budget_ma
        border = QColor("#ff7002") if self.isSelected() else (QColor("#c0392b") if over else QColor("#2c3e50"))
        painter.setBrush(QBrush(QColor("#ecf0f1")))
        painter.setPen(QPen(border, 2))
        painter.drawRoundedRect(QRectF(0,0,self._w,self._h), 6, 6)
        painter.setPen(QColor("#1a1a1a")); painter.setFont(QFont("Arial", 10, QFont.Bold))
        painter.drawText(QRectF(4,4,self._w-8,20), Qt.AlignCenter, self.name)
        pct = (load/self.nac_budget_ma*100) if self.nac_budget_ma else 0
        painter.setFont(QFont("Arial", 8))
        painter.setPen(QColor("#c0392b") if over else QColor("#555"))
        painter.drawText(QRectF(4,26,self._w-8,16), Qt.AlignCenter,
                          f"Direct NAC: {load:.0f}/{self.nac_budget_ma:.0f} mA ({pct:.0f}%)")
        if over:
            painter.setFont(QFont("Arial", 8, QFont.Bold))
            painter.drawText(QRectF(4,44,self._w-8,18), Qt.AlignCenter, "⚠ Over budget — add booster(s)")

    def contextMenuEvent(self, event):
        menu = QMenu()
        add_loop = menu.addAction("+ Add SLC Loop")
        add_nac  = menu.addAction("+ Add NAC Circuit")
        add_boost = menu.addAction("+ Add Booster")
        menu.addSeparator()
        edit_a = menu.addAction("Edit Panel…")
        report_a = menu.addAction("Check Loading / Booster Planner…")
        menu.addSeparator()
        hide_a = menu.addAction("Show Panel Box" if self.hidden else "Hide Panel Box")
        chosen = menu.exec_(event.screenPos())
        sc = self.scene()
        if not sc: return
        if chosen == add_loop: sc.add_circuit(self, "slc")
        elif chosen == add_nac: sc.add_circuit(self, "nac")
        elif chosen == add_boost: sc.add_booster(self)
        elif chosen == edit_a:
            dlg = PanelEditDialog(self.name, self.nac_budget_ma)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values(); self.name = v["name"]; self.nac_budget_ma = v["nac_budget_ma"]
                self.update(); sc.layout_changed.emit()
        elif chosen == report_a:
            sc.show_loading_report()
        elif chosen == hide_a:
            self.hidden = not self.hidden
            self.update(); sc.update_connectors(); sc.layout_changed.emit()


class CircuitNode(OneLineNodeBase):
    ITEM_TYPE = "fa_circuit"

    def __init__(self, circuit_type="nac", name=None, capacity=None, devices=None, circuit_class="B",
                 source_label=""):
        super().__init__()
        self.circuit_type = circuit_type   # "slc" | "nac"
        info = CIRCUIT_TYPE_INFO[circuit_type]
        self.name = name or info["name"]
        self.capacity = capacity if capacity is not None else info["default_capacity"]
        self.devices = devices or []       # [{"key":str, "qty":int, "ma":float|None}]
        self.circuit_class = circuit_class if circuit_class in ("A","B") else "B"
        # Only meaningful for a standalone circuit (parent_node is None) —
        # printed on the box in place of a connector line back to its source,
        # e.g. "Panel — NAC 3" or "Booster 2 — NAC 1", since there's nothing
        # upstream drawn on this diagram to connect a line to.
        self.source_label = source_label
        self.setZValue(2)
        self._recompute_height()

    def _recompute_height(self):
        if self.manual_size:
            return
        _, _, rows = circuit_grid_layout(self.devices, self._w, OL_TICK_W, OL_TICK_H)
        grid_h = max(rows, 1) * OL_TICK_H
        self._h = OL_H_CIRCUIT_BASE + grid_h

    def total_load(self):
        if self.circuit_type == "slc":
            return sum(d["qty"] for d in self.devices)
        total = 0.0
        for d in self.devices:
            ma = d.get("ma")
            if ma is None:
                ma = FA_DEVICE_TYPES.get(d["key"], {}).get("ma", 0.0)
            total += d["qty"]*ma
        return total

    def utilization_pct(self):
        return (self.total_load()/self.capacity*100) if self.capacity else 0.0

    def paint(self, painter, option, widget=None):
        painter.setRenderHint(QPainter.Antialiasing)
        info = CIRCUIT_TYPE_INFO[self.circuit_type]
        load = self.total_load()
        pct = self.utilization_pct()
        if pct > 100: fill, border = QColor("#fdecea"), QColor("#c0392b")
        elif pct > 80: fill, border = QColor("#fef5e7"), QColor("#e67e22")
        else:          fill, border = QColor("#eafaf1"), QColor("#27ae60")
        painter.setBrush(QBrush(fill))
        painter.setPen(QPen(QColor("#ff7002") if self.isSelected() else border, 2))
        painter.drawRoundedRect(QRectF(0,0,self._w,self._h), 6, 6)
        painter.setPen(QColor("#1a1a1a")); painter.setFont(QFont("Arial", 9, QFont.Bold))
        badge = "SLC" if self.circuit_type=="slc" else "NAC"
        cls_tag = f" · Class {self.circuit_class}"
        title = self.name
        if not self.has_visible_parent() and self.source_label:
            title = f"{self.name}  —  {self.source_label}"
        painter.drawText(QRectF(4,3,self._w-8,15), Qt.AlignCenter, f"{title}  [{badge}{cls_tag}]")
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.setPen(border.darker(120))
        unit = info["unit"]
        painter.drawText(QRectF(4,19,self._w-8,14), Qt.AlignCenter,
                          f"{load:.0f}/{self.capacity:.0f} {unit}  ({pct:.0f}%)")

        # Individual device ticks in a wrapping grid — not just a tally count
        ticks, cols, rows = circuit_grid_layout(self.devices, self._w, OL_TICK_W, OL_TICK_H)
        grid_top = 36
        painter.setFont(QFont("Arial", 6))
        for i, abbr in enumerate(ticks):
            col, row = i % cols, i // cols
            x = 8 + col*OL_TICK_W
            y = grid_top + row*OL_TICK_H
            painter.setPen(QPen(border.darker(110), 1))
            painter.setBrush(QBrush(QColor("white")))
            painter.drawRect(QRectF(x, y, OL_TICK_W-3, OL_TICK_H-3))
            painter.setPen(QColor("#333"))
            painter.drawText(QRectF(x, y, OL_TICK_W-3, OL_TICK_H-3), Qt.AlignCenter, abbr)
        if not ticks:
            painter.setPen(QColor("#999")); painter.setFont(QFont("Arial", 7))
            painter.drawText(QRectF(6, grid_top, self._w-12, 13), Qt.AlignLeft, "(no devices — right-click to add)")

        # Terminus: Class B ends in an EOL resistor; Class A loops back to its source
        term_y = self._h - 16
        painter.setPen(QPen(QColor("#555"), 1.3))
        if self.circuit_class == "B":
            zx = self._w/2 - 16
            pts = [QPointF(zx+i*4, term_y+6+((-1)**i)*4) for i in range(9)]
            painter.drawPolyline(QPolygonF(pts))
            painter.setFont(QFont("Arial", 6, QFont.Bold)); painter.setPen(QColor("#555"))
            painter.drawText(QRectF(zx+36, term_y-2, 60, 14), Qt.AlignLeft, "EOL")
        else:
            painter.drawArc(QRectF(self._w/2-10, term_y-2, 20, 16), 30*16, 300*16)
            painter.setFont(QFont("Arial", 6, QFont.Bold)); painter.setPen(QColor("#555"))
            painter.drawText(QRectF(self._w/2+12, term_y-2, 100, 14), Qt.AlignLeft, "Class A return")

    def contextMenuEvent(self, event):
        menu = QMenu()
        edit_a = menu.addAction("Edit Devices / Capacity…")
        del_a = menu.addAction("Delete Circuit")
        chosen = menu.exec_(event.screenPos())
        sc = self.scene()
        if chosen == edit_a:
            dlg = CircuitEditDialog(self.circuit_type, self.name, self.capacity, self.devices, self.circuit_class,
                                     source_label=self.source_label)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values()
                self.name = v["name"]; self.capacity = v["capacity"]; self.devices = v["devices"]
                self.circuit_class = v["circuit_class"]; self.source_label = v["source_label"]
                self.prepareGeometryChange()
                self._recompute_height()
                self.update()
                if sc: sc.relayout(); sc.layout_changed.emit()
        elif chosen == del_a and sc:
            sc.remove_node(self)


class BoosterNode(OneLineNodeBase):
    ITEM_TYPE = "fa_booster"

    def __init__(self, name="Booster 1", capacity_ma=3000.0):
        super().__init__()
        self.name = name
        self.capacity_ma = capacity_ma
        self.hidden = False
        self._h = OL_H_BOOSTER
        self.setZValue(2)

    def total_load(self):
        return sum(c.total_load() for c in self.children
                   if getattr(c, "circuit_type", None) == "nac")

    def paint(self, painter, option, widget=None):
        if self.hidden:
            return
        painter.setRenderHint(QPainter.Antialiasing)
        load = self.total_load()
        over = self.capacity_ma > 0 and load > self.capacity_ma
        border = QColor("#ff7002") if self.isSelected() else (QColor("#c0392b") if over else QColor("#7d3c98"))
        painter.setBrush(QBrush(QColor("#f5eefc")))
        painter.setPen(QPen(border, 2))
        painter.drawRoundedRect(QRectF(0,0,self._w,self._h), 6, 6)
        painter.setPen(QColor("#1a1a1a")); painter.setFont(QFont("Arial", 9, QFont.Bold))
        painter.drawText(QRectF(4,4,self._w-8,18), Qt.AlignCenter, f"⚡ {self.name}")
        pct = (load/self.capacity_ma*100) if self.capacity_ma else 0
        painter.setFont(QFont("Arial", 8))
        painter.setPen(QColor("#c0392b") if over else QColor("#555"))
        painter.drawText(QRectF(4,26,self._w-8,16), Qt.AlignCenter,
                          f"{load:.0f}/{self.capacity_ma:.0f} mA ({pct:.0f}%)")

    def contextMenuEvent(self, event):
        menu = QMenu()
        add_nac = menu.addAction("+ Add NAC Circuit")
        menu.addSeparator()
        edit_a = menu.addAction("Edit Booster…")
        del_a = menu.addAction("Delete Booster")
        menu.addSeparator()
        hide_a = menu.addAction("Show Booster Box" if self.hidden else "Hide Booster Box")
        chosen = menu.exec_(event.screenPos())
        sc = self.scene()
        if not sc: return
        if chosen == add_nac:
            sc.add_circuit(self, "nac")
        elif chosen == edit_a:
            dlg = BoosterEditDialog(self.name, self.capacity_ma)
            if dlg.exec_() == QDialog.Accepted:
                v = dlg.values(); self.name = v["name"]; self.capacity_ma = v["capacity_ma"]
                self.update(); sc.layout_changed.emit()
        elif chosen == del_a:
            sc.remove_node(self)
        elif chosen == hide_a:
            self.hidden = not self.hidden
            self.update(); sc.update_connectors(); sc.layout_changed.emit()


class OneLineScene(QGraphicsScene):
    layout_changed = pyqtSignal()
    status_changed = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setSceneRect(-3000, -3000, 6000, 6000)
        self.setBackgroundBrush(QBrush(QColor("white")))
        self.panel = PanelNode()
        self.addItem(self.panel)
        self.standalone = []    # standalone CircuitNodes — no panel/booster parent
        self._connectors = []
        self.auto_arrange()

    def set_mode_select(self):
        """No-op — the one-line diagram has no draw/place tool modes, this
        just gives it the same interface as FloorPlanScene/WiringScene so
        the shared tab-switch/toolbar plumbing can call it unconditionally."""
        self.clearSelection()

    def all_nodes(self):
        out = []
        def walk(n):
            out.append(n)
            for c in n.children: walk(c)
        walk(self.panel)
        for s in self.standalone:
            walk(s)
        return out

    def _place_new_node(self, node):
        """Lightweight default placement for a freshly created node — put it
        near its parent (or, for a standalone node, to the right of existing
        content) without disturbing anything else already on the canvas.
        Auto-Arrange is the full clean-layout escape hatch."""
        if node.parent_node is not None:
            p = node.parent_node
            siblings = [c for c in p.children if c is not node]
            if siblings:
                last = siblings[-1]
                node.setPos(last.pos().x() + last._w + OL_SIB_GAP, last.pos().y())
            else:
                node.setPos(p.pos().x(), p.pos().y() + p._h + OL_LEVEL_GAP)
        else:
            others = [n for n in self.all_nodes() if n is not node]
            if not others:
                node.setPos(0, 0)
            else:
                max_right = max(n.pos().x() + n._w for n in others)
                node.setPos(max_right + OL_SIB_GAP*3, self.panel.pos().y())
        node.manual_pos = False

    def add_circuit(self, parent_node, circuit_type):
        c = CircuitNode(circuit_type)
        c.parent_node = parent_node
        parent_node.children.append(c)
        self.addItem(c)
        self._place_new_node(c)
        self.update_connectors()
        self.layout_changed.emit()
        return c

    def add_standalone_circuit(self, circuit_type):
        c = CircuitNode(circuit_type)
        c.parent_node = None
        self.standalone.append(c)
        self.addItem(c)
        self._place_new_node(c)
        self.update_connectors()
        self.layout_changed.emit()
        return c

    def add_booster(self, parent_node):
        n = sum(1 for nd in self.all_nodes() if isinstance(nd, BoosterNode)) + 1
        b = BoosterNode(name=f"Booster {n}")
        b.parent_node = parent_node
        parent_node.children.append(b)
        self.addItem(b)
        self._place_new_node(b)
        self.update_connectors()
        self.layout_changed.emit()
        return b

    def remove_node(self, node):
        if node is self.panel:
            return
        for c in list(node.children):
            self.remove_node(c)
        if node.parent_node and node in node.parent_node.children:
            node.parent_node.children.remove(node)
        if node in self.standalone:
            self.standalone.remove(node)
        self.removeItem(node)
        self.update_connectors()
        self.layout_changed.emit()

    def clear_all(self):
        for n in list(self.all_nodes()):
            if n is not self.panel:
                self.removeItem(n)
        self.panel.children = []
        self.standalone = []
        for ln in self._connectors:
            self.removeItem(ln)
        self._connectors = []
        self.update_connectors()
        self.layout_changed.emit()

    def relayout(self):
        """Cheap refresh — redraw connector lines from current positions.
        Does NOT reposition any node; see auto_arrange() for that."""
        self.update_connectors()

    def update_connectors(self):
        from PyQt5.QtWidgets import QGraphicsLineItem
        for ln in self._connectors:
            self.removeItem(ln)
        self._connectors = []

        def walk(n):
            for c in n.children:
                if not getattr(n, "hidden", False) and not getattr(c, "hidden", False):
                    line = QGraphicsLineItem(n.pos().x()+n._w/2, n.pos().y()+n._h,
                                              c.pos().x()+c._w/2, c.pos().y())
                    line.setPen(QPen(QColor("#888"), 1.6))
                    line.setZValue(0.5)
                    self.addItem(line)
                    self._connectors.append(line)
                # Keep recursing even when n or c is hidden, so a hidden
                # node's children (e.g. circuits under a hidden booster)
                # still get their own subtrees connected.
                walk(c)
        walk(self.panel)
        for s in self.standalone:
            walk(s)
        self.update()

    def auto_arrange(self):
        """Full clean tree layout — resets every node's position (including
        ones the user dragged) back to the automatic tidy arrangement."""
        for n in self.all_nodes():
            n.manual_pos = False
        self._layout_subtree(self.panel, 0, 0)
        x = max((self.panel.pos().x() + self.panel.subtree_width() + OL_SIB_GAP*3), 0)
        for s in self.standalone:
            w = self._layout_subtree(s, x, self.panel.pos().y())
            x += w + OL_SIB_GAP*2
        self.update_connectors()

    def _layout_subtree(self, node, x, y):
        w = node.subtree_width()
        if not node.children:
            node.setPos(x + (w-node._w)/2, y)
            return w
        child_y = y + node._h + OL_LEVEL_GAP   # dynamic — circuit boxes vary in height
        cx = x
        for c in node.children:
            cw = self._layout_subtree(c, cx, child_y)
            cx += cw + OL_SIB_GAP
        first, last = node.children[0], node.children[-1]
        left = first.pos().x() + first._w/2
        right = last.pos().x() + last._w/2
        node.setPos((left+right)/2 - node._w/2, y)
        return w

    def show_loading_report(self):
        lines = [f"Panel direct NAC load: {self.panel.direct_nac_load_ma():.0f} / "
                 f"{self.panel.nac_budget_ma:.0f} mA"]
        overflow = self.panel.direct_nac_load_ma() - self.panel.nac_budget_ma
        if overflow > 0:
            lines.append(f"\n⚠ Panel NAC budget exceeded by {overflow:.0f} mA.")
            lines.append("Boosters needed at common capacities:")
            for cap in BOOSTER_CAPACITY_PRESETS_MA:
                need = math.ceil(overflow/cap)
                lines.append(f"   • {cap} mA booster(s): {need} needed")
        else:
            lines.append("Panel NAC budget OK.")
        over_circuits = [n for n in self.all_nodes() if isinstance(n, CircuitNode) and n.utilization_pct() > 100]
        if over_circuits:
            lines.append("\n⚠ Circuits individually over capacity:")
            for c in over_circuits:
                lines.append(f"   • {c.name}: {c.total_load():.0f}/{c.capacity:.0f} "
                              f"{CIRCUIT_TYPE_INFO[c.circuit_type]['unit']}")
        over_boosters = [n for n in self.all_nodes() if isinstance(n, BoosterNode)
                         and n.capacity_ma > 0 and n.total_load() > n.capacity_ma]
        if over_boosters:
            lines.append("\n⚠ Boosters individually over capacity:")
            for b in over_boosters:
                lines.append(f"   • {b.name}: {b.total_load():.0f}/{b.capacity_ma:.0f} mA")
        QMessageBox.information(None, "Loading / Booster Planner", "\n".join(lines))

    def to_dict(self):
        def node_dict(n):
            base = {"x": n.pos().x(), "y": n.pos().y(), "w": n._w, "h": n._h,
                    "manual_pos": n.manual_pos, "manual_size": n.manual_size}
            if isinstance(n, PanelNode):
                base.update({"type":"panel", "name":n.name, "nac_budget_ma":n.nac_budget_ma,
                              "hidden": n.hidden})
            elif isinstance(n, BoosterNode):
                base.update({"type":"booster", "name":n.name, "capacity_ma":n.capacity_ma, "hidden":n.hidden})
            elif isinstance(n, CircuitNode):
                base.update({"type":"circuit", "circuit_type":n.circuit_type, "name":n.name,
                              "capacity":n.capacity, "devices":n.devices, "circuit_class":n.circuit_class,
                              "source_label": n.source_label})
            base["children"] = [node_dict(c) for c in n.children]
            return base
        return {"panel": node_dict(self.panel),
                "standalone": [node_dict(s) for s in self.standalone]}

    def _apply_geom(self, node, cd):
        if "x" in cd and "y" in cd:
            node.setPos(cd["x"], cd["y"])
            node.manual_pos = cd.get("manual_pos", True)
        if "w" in cd and "h" in cd:
            node._w = cd["w"]; node._h = cd["h"]
            node.manual_size = cd.get("manual_size", False)

    def load_dict(self, d):
        self.clear_all()
        if not d:
            self.auto_arrange(); return
        # New format is {"panel": {...}, "standalone": [...]}; a bare dict
        # with "type":"panel" at the top level is the old flat, panel-only,
        # no-geometry format from before drag/resize/standalone existed.
        if "panel" in d:
            panel_d, standalone_d = d["panel"], d.get("standalone", [])
        else:
            panel_d, standalone_d = d, []
        self.panel.name = panel_d.get("name","FACP")
        self.panel.nac_budget_ma = panel_d.get("nac_budget_ma", 3000.0)
        self.panel.hidden = panel_d.get("hidden", False)
        self._apply_geom(self.panel, panel_d)

        def build(parent_node, cd):
            t = cd.get("type")
            if t == "circuit":
                node = CircuitNode(cd.get("circuit_type","nac"), cd.get("name"),
                                    cd.get("capacity"), cd.get("devices", []), cd.get("circuit_class", "B"),
                                    source_label=cd.get("source_label", ""))
            elif t == "booster":
                node = BoosterNode(cd.get("name","Booster"), cd.get("capacity_ma",3000.0))
                node.hidden = cd.get("hidden", False)
            else:
                return None
            node.parent_node = parent_node
            if parent_node is not None:
                parent_node.children.append(node)
            self.addItem(node)
            self._apply_geom(node, cd)
            for ccd in cd.get("children", []):
                build(node, ccd)
            return node

        for cd in panel_d.get("children", []):
            build(self.panel, cd)
        for cd in standalone_d:
            node = build(None, cd)
            if node is not None:
                self.standalone.append(node)

        if any(not n.manual_pos for n in self.all_nodes()):
            self.auto_arrange()
        self.update_connectors()
        self.layout_changed.emit()


class OneLineCanvas(QGraphicsView):
    def __init__(self, scene):
        super().__init__(scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setDragMode(QGraphicsView.RubberBandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self._panning = False; self._pan_start = None

    def wheelEvent(self, e):
        f = 1.15 if e.angleDelta().y() > 0 else 1/1.15
        self.scale(f, f)

    def mousePressEvent(self, e):
        if e.button() == Qt.MiddleButton or (e.button()==Qt.LeftButton and (e.modifiers()&Qt.ControlModifier)):
            self._panning = True; self._pan_start = e.pos(); self.setCursor(Qt.ClosedHandCursor); return
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self._panning and self._pan_start:
            d = e.pos()-self._pan_start; self._pan_start = e.pos()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value()-d.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value()-d.y()); return
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        if self._panning:
            self._panning = False; self.setCursor(Qt.ArrowCursor)
        super().mouseReleaseEvent(e)

    def fit_all(self):
        items = self.scene().items()
        if not items:
            return
        r = items[0].sceneBoundingRect()
        for i in items[1:]:
            r = r.united(i.sceneBoundingRect())
        self.fitInView(r.adjusted(-60,-60,60,60), Qt.KeepAspectRatio)


class PanelEditDialog(QDialog):
    def __init__(self, name, nac_budget_ma, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Panel"); self.setMinimumWidth(320)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.name_edit = QLineEdit(name)
        self.budget_spin = QDoubleSpinBox(); self.budget_spin.setRange(0, 100000)
        self.budget_spin.setValue(nac_budget_ma); self.budget_spin.setSuffix("  mA")
        l.addRow("Panel name:", self.name_edit)
        l.addRow("Total NAC power budget:", self.budget_spin)
        l.addRow(QLabel("(Combined mA available across all NAC circuits wired\n"
                         "directly off the panel, before a booster is required.)"))
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        return {"name": self.name_edit.text().strip() or "FACP",
                "nac_budget_ma": self.budget_spin.value()}


class BoosterEditDialog(QDialog):
    def __init__(self, name, capacity_ma, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Booster"); self.setMinimumWidth(320)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.name_edit = QLineEdit(name)
        self.cap_combo = QComboBox(); self.cap_combo.setEditable(True)
        for c in BOOSTER_CAPACITY_PRESETS_MA:
            self.cap_combo.addItem(f"{c} mA", c)
        idx = self.cap_combo.findData(capacity_ma)
        if idx >= 0: self.cap_combo.setCurrentIndex(idx)
        else: self.cap_combo.setEditText(f"{capacity_ma:g}")
        l.addRow("Booster name:", self.name_edit)
        l.addRow("Output capacity:", self.cap_combo)
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        data = self.cap_combo.currentData()
        if data is None:
            try: data = float(re.sub(r"[^\d.]", "", self.cap_combo.currentText()) or 0)
            except ValueError: data = 3000.0
        return {"name": self.name_edit.text().strip() or "Booster", "capacity_ma": float(data)}


class CircuitEditDialog(QDialog):
    """Edit a circuit's name/capacity, its individual device tally, and its
    NFPA 72 wiring class (A = loops back to the panel/booster, no EOL needed;
    B = single run, ends in an end-of-line resistor)."""
    def __init__(self, circuit_type, name, capacity, devices, circuit_class="B", parent=None,
                 source_label=""):
        super().__init__(parent)
        self.circuit_type = circuit_type
        self._keys = FA_NAC_DEVICE_KEYS if circuit_type=="nac" else FA_SLC_DEVICE_KEYS
        info = CIRCUIT_TYPE_INFO[circuit_type]
        self.setWindowTitle(f"Edit {info['name']}"); self.setMinimumWidth(480)
        l = QVBoxLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)

        # Created before any _add_row() calls below — _add_row() calls
        # _refresh_total() as it goes, which needs total_lbl to already exist
        # (it's added to the layout later, at its proper visual position).
        self._rows = []
        self.total_lbl = QLabel(); self.total_lbl.setStyleSheet("font-weight:bold;padding:4px;")

        form = QFormLayout(); form.setSpacing(8)
        self.name_edit = QLineEdit(name)
        self.cap_spin = QDoubleSpinBox(); self.cap_spin.setRange(1, 100000)
        self.cap_spin.setValue(capacity); self.cap_spin.setSuffix(f"  {info['unit']}")
        form.addRow("Name:", self.name_edit)
        form.addRow(f"Capacity ({info['unit']}):", self.cap_spin)

        self.class_a_rb = QRadioButton(f"{CIRCUIT_CLASS_INFO['A']['label']} — {CIRCUIT_CLASS_INFO['A']['desc']}")
        self.class_b_rb = QRadioButton(f"{CIRCUIT_CLASS_INFO['B']['label']} — {CIRCUIT_CLASS_INFO['B']['desc']}")
        (self.class_a_rb if circuit_class=="A" else self.class_b_rb).setChecked(True)
        cls_grp = QButtonGroup(self); cls_grp.addButton(self.class_a_rb); cls_grp.addButton(self.class_b_rb)
        form.addRow("Wiring class:", self.class_a_rb)
        form.addRow("", self.class_b_rb)

        self.source_edit = QLineEdit(source_label)
        self.source_edit.setPlaceholderText("e.g. Panel — NAC 3, or Booster 2 — NAC 1")
        self.source_edit.setToolTip(
            "Only shown on the box when this circuit has no visible parent — i.e. it's "
            "standalone, or its panel/booster box is hidden — in place of a connector line.")
        form.addRow("Source label:", self.source_edit)
        l.addLayout(form)

        self.tbl = QTableWidget(0, 4)
        headers = ["Device", "Qty", "mA ea" if circuit_type=="nac" else "", ""]
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        for d in devices:
            self._add_row(d["key"], d["qty"], d.get("ma"))
        l.addWidget(self.tbl)

        add_btn = QPushButton("+ Add Device Row")
        add_btn.clicked.connect(lambda: self._add_row(self._keys[0], 1, None))
        l.addWidget(add_btn)

        l.addWidget(self.total_lbl)
        self.cap_spin.valueChanged.connect(self._refresh_total)
        self._refresh_total()

        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addLayout(br)

    def _add_row(self, key, qty, ma_override):
        r = self.tbl.rowCount(); self.tbl.insertRow(r)
        combo = QComboBox()
        for k in self._keys:
            combo.addItem(FA_DEVICE_TYPES[k]["name"], k)
        idx = combo.findData(key)
        combo.setCurrentIndex(idx if idx >= 0 else 0)
        qty_spin = QSpinBox(); qty_spin.setRange(1, 999); qty_spin.setValue(qty)
        ma_spin = QDoubleSpinBox(); ma_spin.setRange(0, 2000); ma_spin.setSuffix(" mA")
        default_ma = FA_DEVICE_TYPES.get(key, {}).get("ma", 0.0)
        ma_spin.setValue(ma_override if ma_override is not None else default_ma)
        ma_spin.setEnabled(self.circuit_type == "nac")
        rm_btn = QPushButton("✕"); rm_btn.setFixedWidth(24)

        def _on_type_changed(_i, combo=combo, ma_spin=ma_spin):
            k = combo.currentData()
            ma_spin.setValue(FA_DEVICE_TYPES.get(k, {}).get("ma", 0.0))
        combo.currentIndexChanged.connect(_on_type_changed)
        qty_spin.valueChanged.connect(self._refresh_total)
        ma_spin.valueChanged.connect(self._refresh_total)

        row = {"combo":combo, "qty":qty_spin, "ma":ma_spin}
        rm_btn.clicked.connect(lambda: self._remove_row(row))
        self.tbl.setCellWidget(r, 0, combo); self.tbl.setCellWidget(r, 1, qty_spin)
        self.tbl.setCellWidget(r, 2, ma_spin); self.tbl.setCellWidget(r, 3, rm_btn)
        self._rows.append(row)
        self._refresh_total()

    def _remove_row(self, row):
        i = self._rows.index(row)
        self.tbl.removeRow(i)
        self._rows.pop(i)
        self._refresh_total()

    def _refresh_total(self):
        if self.circuit_type == "nac":
            total = sum(r["qty"].value()*r["ma"].value() for r in self._rows)
            cap = self.cap_spin.value()
            pct = (total/cap*100) if cap else 0
            self.total_lbl.setText(f"Total load: {total:.0f} mA / {cap:.0f} mA  ({pct:.0f}%)"
                                    + ("   ⚠ OVER CAPACITY" if total > cap else ""))
            self.total_lbl.setStyleSheet(
                "font-weight:bold;padding:4px;color:%s;" % ("#c0392b" if total > cap else "#27ae60"))
        else:
            total = sum(r["qty"].value() for r in self._rows)
            cap = self.cap_spin.value()
            pct = (total/cap*100) if cap else 0
            self.total_lbl.setText(f"Total devices: {total:.0f} / {cap:.0f}  ({pct:.0f}%)"
                                    + ("   ⚠ OVER CAPACITY" if total > cap else ""))
            self.total_lbl.setStyleSheet(
                "font-weight:bold;padding:4px;color:%s;" % ("#c0392b" if total > cap else "#27ae60"))

    def values(self):
        devices = []
        for r in self._rows:
            key = r["combo"].currentData()
            qty = r["qty"].value()
            ma = r["ma"].value() if self.circuit_type == "nac" else None
            default_ma = FA_DEVICE_TYPES.get(key, {}).get("ma", 0.0)
            devices.append({"key": key, "qty": qty,
                             "ma": ma if (ma is not None and abs(ma-default_ma) > 1e-6) else None})
        return {"name": self.name_edit.text().strip() or CIRCUIT_TYPE_INFO[self.circuit_type]["name"],
                "capacity": self.cap_spin.value(), "devices": devices,
                "circuit_class": "A" if self.class_a_rb.isChecked() else "B",
                "source_label": self.source_edit.text().strip()}


# ═══════════════════════════════════════════════════════════════════════════════
#  FQQ (Autocall "Fast Quick Quote" .xlsm) import
#
#  The FQQ tool's own NAC sheet embeds a device current-draw reference table
#  (column DI = part number, DM = candela setting, DN = nameplate mA) plus,
#  for whichever part/candela rows the estimator actually used, a non-zero
#  quantity in the designation column ("Floor 1", "Warehouse", etc — named in
#  row 24). We read that live data directly rather than guessing candela/mA,
#  register any device not already in FA_DEVICE_TYPES on the fly, then bin
#  the resulting device list into NAC circuits sized to the app's normal NAC
#  circuit capacity.
# ═══════════════════════════════════════════════════════════════════════════════

def _fqq_register_device(part, candela, ma):
    cd_tag = re.sub(r"[^0-9a-z]+", "", str(candela).lower()) if candela else ""
    key = re.sub(r"[^0-9a-z]+", "_", f"fqq_{part}_{cd_tag}".lower()).strip("_")
    if key not in FA_DEVICE_TYPES:
        label = f"{part} ({candela})" if candela else str(part)
        abbr = str(part).split("-")[-1][-5:] + (f"-{cd_tag}" if cd_tag else "")
        FA_DEVICE_TYPES[key] = {"name": label, "category": "nac", "ma": float(ma),
                                 "abbr": abbr[:8], "mfr": "Autocall", "part": str(part)}
        FA_NAC_DEVICE_KEYS.append(key)
    return key


def import_fqq_xlsm(path):
    """Parse an Autocall FQQ quote (.xlsm) and return
    {designation_name: [{"key":str, "qty":int}, ...]} for its conventional
    NAC devices. Raises ValueError if the file doesn't look like an FQQ."""
    import openpyxl
    from openpyxl.utils import column_index_from_string
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
    if "NAC" not in wb.sheetnames:
        raise ValueError("No 'NAC' sheet found — this doesn't look like an Autocall FQQ file.")
    ws = wb["NAC"]
    col_start, col_end = column_index_from_string("J"), column_index_from_string("DE")
    col_part = column_index_from_string("DI")
    col_cd = column_index_from_string("DM")
    col_ma = column_index_from_string("DN")

    names = {}
    for c in range(col_start, col_end + 1):
        name = (ws.cell(row=24, column=c).value or ws.cell(row=11, column=c).value
                or ws.cell(row=9, column=c).value)
        if name:
            names[c] = str(name)

    result = {}
    for r in range(1, ws.max_row + 1):
        part = ws.cell(row=r, column=col_part).value
        ma = ws.cell(row=r, column=col_ma).value
        if not part or not isinstance(ma, (int, float)) or ma <= 0:
            continue
        candela = ws.cell(row=r, column=col_cd).value
        key = _fqq_register_device(part, candela, ma)
        for c in range(col_start, col_end + 1):
            if c not in names:
                continue
            qty = ws.cell(row=r, column=c).value
            if isinstance(qty, (int, float)) and qty > 0:
                result.setdefault(names[c], []).append({"key": key, "qty": int(qty)})
    return result


def bin_pack_nac_devices(devices, capacity):
    """Greedy first-fit-decreasing pack of a [{'key','qty'}] tally into NAC
    circuits not exceeding capacity mA. Returns a list of tally lists, one
    per circuit."""
    items = []
    for d in devices:
        ma = FA_DEVICE_TYPES.get(d["key"], {}).get("ma", 0.0)
        items.extend([(d["key"], ma)] * d["qty"])
    items.sort(key=lambda x: -x[1])
    bins = []  # [(remaining_capacity, {key: qty})]
    for key, ma in items:
        for i, (remaining, tally) in enumerate(bins):
            if ma <= remaining:
                tally[key] = tally.get(key, 0) + 1
                bins[i] = (remaining - ma, tally)
                break
        else:
            bins.append((capacity - ma, {key: 1}))
    return [[{"key": k, "qty": q} for k, q in tally.items()] for _, tally in bins]


# ═══════════════════════════════════════════════════════════════════════════════
#  Project info
# ═══════════════════════════════════════════════════════════════════════════════

class ProjectInfoDialog(QDialog):
    def __init__(self, meta=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Project Info"); self.setMinimumWidth(340)
        meta = meta or {}
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.customer = QLineEdit(meta.get("customer",""))
        self.location = QLineEdit(meta.get("location",""))
        self.job_number = QLineEdit(meta.get("job_number",""))
        self.designer = QLineEdit(meta.get("designer",""))
        l.addRow("Customer / Project:", self.customer)
        l.addRow("Location:", self.location)
        l.addRow("Job #:", self.job_number)
        l.addRow("Designer:", self.designer)
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def values(self):
        return {"customer": self.customer.text().strip(), "location": self.location.text().strip(),
                "job_number": self.job_number.text().strip(), "designer": self.designer.text().strip()}


class WallSettingsDialog(QDialog):
    def __init__(self, current_in, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wall Thickness"); self.setMinimumWidth(340)
        l = QFormLayout(self); l.setSpacing(10); l.setContentsMargins(16,16,16,16)
        self.combo = QComboBox()
        for label, val in WALL_THICKNESS_OPTIONS:
            self.combo.addItem(label, val)
        idx = self.combo.findData(current_in)
        if idx >= 0: self.combo.setCurrentIndex(idx)
        self.custom = QDoubleSpinBox(); self.custom.setRange(0.5, 24.0); self.custom.setSuffix("  in")
        self.custom.setValue(current_in)
        l.addRow("Preset:", self.combo)
        l.addRow("Or custom:", self.custom)
        self.combo.currentIndexChanged.connect(lambda: self.custom.setValue(self.combo.currentData()))
        br = QHBoxLayout()
        ok = QPushButton("OK"); ok.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        ok.clicked.connect(self.accept)
        ca = QPushButton("Cancel"); ca.clicked.connect(self.reject)
        br.addStretch(); br.addWidget(ca); br.addWidget(ok)
        l.addRow(br)

    def value_in(self):
        return self.custom.value()


# ═══════════════════════════════════════════════════════════════════════════════
#  One-Line Diagram PDF export — not to scale, auto-picks the smallest
#  standard sheet the tree fits on (shrinking only if it exceeds ARCH E).
# ═══════════════════════════════════════════════════════════════════════════════

def _oneline_node_color(n):
    if isinstance(n, PanelNode):
        over = n.nac_budget_ma > 0 and n.direct_nac_load_ma() > n.nac_budget_ma
        return (0.75,0.22,0.17) if over else (0.17,0.24,0.31)
    if isinstance(n, BoosterNode):
        over = n.capacity_ma > 0 and n.total_load() > n.capacity_ma
        return (0.75,0.22,0.17) if over else (0.49,0.24,0.60)
    if isinstance(n, CircuitNode):
        pct = n.utilization_pct()
        if pct > 100: return (0.75,0.22,0.17)
        if pct > 80: return (0.90,0.49,0.13)
        return (0.15,0.68,0.38)
    return (0,0,0)


def export_oneline_pdf(scene, path, project_meta=None, sheet_title="ONE-LINE DIAGRAM"):
    project_meta = project_meta or {}
    nodes = scene.all_nodes()
    xs = [n.pos().x() for n in nodes] + [n.pos().x()+n._w for n in nodes]
    ys = [n.pos().y() for n in nodes] + [n.pos().y()+n._h for n in nodes]
    min_x, max_x, min_y, max_y = min(xs), max(xs), min(ys), max(ys)
    content_w, content_h = max(1.0, max_x-min_x), max(1.0, max_y-min_y)

    margin_pt = 36
    title_h_pt = 80
    candidates = ["Letter (8.5x11)", "Tabloid (11x17)", "ANSI C (17x22)",
                  "ANSI D (22x34)", "ARCH D (24x36)", "ARCH E (36x48)"]
    chosen = None
    for key in candidates:
        pw0, ph0 = PAPER_SIZES_IN[key]
        for pw_in, ph_in in ((max(pw0,ph0), min(pw0,ph0)), (min(pw0,ph0), max(pw0,ph0))):
            avail_w = pw_in*72 - 2*margin_pt
            avail_h = ph_in*72 - 2*margin_pt - title_h_pt
            if content_w <= avail_w and content_h <= avail_h:
                chosen = (pw_in, ph_in, 1.0); break
        if chosen: break
    if not chosen:
        pw_in, ph_in = 48.0, 36.0   # ARCH E landscape
        avail_w = pw_in*72 - 2*margin_pt
        avail_h = ph_in*72 - 2*margin_pt - title_h_pt
        scale = min(avail_w/content_w, avail_h/content_h, 1.0)
        chosen = (pw_in, ph_in, scale)
    pw_in, ph_in, scale = chosen

    doc = fitz.open()
    page = doc.new_page(width=pw_in*72, height=ph_in*72)
    origin_x = margin_pt + max(0.0, (pw_in*72-2*margin_pt - content_w*scale)/2)
    origin_y = margin_pt + max(0.0, (ph_in*72-2*margin_pt-title_h_pt - content_h*scale)/2)

    def tx(x, y):
        return fitz.Point(origin_x+(x-min_x)*scale, origin_y+(y-min_y)*scale)

    # Pass 1: connector lines + node boxes, batched through one Shape and
    # committed BEFORE any text is inserted — insert_text() writes straight to
    # the page's content stream, so text added before commit() would end up
    # UNDER the shape's opaque box fills once they're painted.
    shape = page.new_shape()

    def walk_lines(node):
        for c in node.children:
            if not getattr(node, "hidden", False) and not getattr(c, "hidden", False):
                p1 = tx(node.pos().x()+node._w/2, node.pos().y()+node._h)
                p2 = tx(c.pos().x()+c._w/2, c.pos().y())
                shape.draw_line(p1, p2); shape.finish(color=(0.55,0.55,0.55), width=1.2)
            walk_lines(c)   # keep recursing even when node/c is hidden
    walk_lines(scene.panel)
    for s in scene.standalone:
        walk_lines(s)

    for n in scene.all_nodes():
        if getattr(n, "hidden", False):
            continue
        p1 = tx(n.pos().x(), n.pos().y())
        p2 = tx(n.pos().x()+n._w, n.pos().y()+n._h)
        col = _oneline_node_color(n)
        shape.draw_rect(fitz.Rect(p1.x, p1.y, p2.x, p2.y))
        shape.finish(color=col, fill=(0.97,0.97,0.97), width=1.4)
        if isinstance(n, CircuitNode):
            # Individual device tick boxes
            ticks, cols, rows = circuit_grid_layout(n.devices, n._w, OL_TICK_W, OL_TICK_H)
            for i, abbr in enumerate(ticks):
                col_i, row_i = i % cols, i // cols
                tx0 = n.pos().x() + 8 + col_i*OL_TICK_W
                ty0 = n.pos().y() + 36 + row_i*OL_TICK_H
                tp1 = tx(tx0, ty0); tp2 = tx(tx0+OL_TICK_W-3, ty0+OL_TICK_H-3)
                shape.draw_rect(fitz.Rect(tp1.x, tp1.y, tp2.x, tp2.y))
                shape.finish(color=(0.4,0.4,0.4), fill=(1,1,1), width=0.6)
            # Terminus: Class B EOL resistor zigzag, Class A return arc
            term_y = n.pos().y() + n._h - 16
            if n.circuit_class == "B":
                zx = n.pos().x() + n._w/2 - 16
                pts = [tx(zx+i*4, term_y+6+((-1)**i)*4) for i in range(9)]
                shape.draw_polyline(pts); shape.finish(color=(0.3,0.3,0.3), width=1.0)
            else:
                c = tx(n.pos().x()+n._w/2, term_y+6)
                r = 8*scale
                shape.draw_circle(c, r); shape.finish(color=(0.3,0.3,0.3), width=1.0)
    shape.commit()

    # Pass 2: text, now safely on top of the committed boxes/lines.
    for n in scene.all_nodes():
        if getattr(n, "hidden", False):
            continue
        p1 = tx(n.pos().x(), n.pos().y())
        col = _oneline_node_color(n)
        if isinstance(n, PanelNode):
            page.insert_text(fitz.Point(p1.x+6, p1.y+16*scale+4), n.name,
                              fontsize=9, fontname="helv", color=(0,0,0))
            page.insert_text(fitz.Point(p1.x+6, p1.y+32*scale+4),
                              f"Direct NAC: {n.direct_nac_load_ma():.0f}/{n.nac_budget_ma:.0f} mA",
                              fontsize=7, color=col)
        elif isinstance(n, BoosterNode):
            page.insert_text(fitz.Point(p1.x+6, p1.y+18*scale+4), f"Booster: {n.name}",
                              fontsize=9, fontname="helv", color=(0,0,0))
            page.insert_text(fitz.Point(p1.x+6, p1.y+34*scale+4),
                              f"{n.total_load():.0f}/{n.capacity_ma:.0f} mA", fontsize=7, color=col)
        elif isinstance(n, CircuitNode):
            info = CIRCUIT_TYPE_INFO[n.circuit_type]
            title = n.name
            if not n.has_visible_parent() and n.source_label:
                title = f"{n.name}  —  {n.source_label}"
            page.insert_text(fitz.Point(p1.x+5, p1.y+13*scale+4),
                              f"{title} [{info['name']} · Class {n.circuit_class}]",
                              fontsize=7.5, fontname="helv", color=(0,0,0))
            page.insert_text(fitz.Point(p1.x+5, p1.y+27*scale+4),
                              f"{n.total_load():.0f}/{n.capacity:.0f} {info['unit']} ({n.utilization_pct():.0f}%)",
                              fontsize=6.5, color=col)
            ticks, cols, rows = circuit_grid_layout(n.devices, n._w, OL_TICK_W, OL_TICK_H)
            for i, abbr in enumerate(ticks):
                col_i, row_i = i % cols, i // cols
                tx0 = n.pos().x() + 8 + col_i*OL_TICK_W
                ty0 = n.pos().y() + 36 + row_i*OL_TICK_H
                tp = tx(tx0+2, ty0+(OL_TICK_H-3)*0.7)
                page.insert_text(tp, abbr, fontsize=5, color=(0.2,0.2,0.2))
            if not ticks:
                page.insert_text(fitz.Point(p1.x+5, p1.y+40*scale+4), "(no devices)",
                                  fontsize=6, color=(0.6,0.6,0.6))
            term_y_scene = n.pos().y() + n._h - 16
            if n.circuit_class == "B":
                lbl_x = n.pos().x()+n._w/2 - 16 + 36
                page.insert_text(tx(lbl_x, term_y_scene+2), "EOL", fontsize=6, fontname="helv", color=(0.3,0.3,0.3))
            else:
                lbl_x = n.pos().x()+n._w/2 + 12
                page.insert_text(tx(lbl_x, term_y_scene+2), "Class A return",
                                  fontsize=6, fontname="helv", color=(0.3,0.3,0.3))

    tb_y = ph_in*72 - title_h_pt
    page.draw_line(fitz.Point(margin_pt, tb_y), fitz.Point(pw_in*72-margin_pt, tb_y), color=(0,0,0), width=1.0)
    page.insert_text(fitz.Point(margin_pt, tb_y+18), sheet_title, fontsize=13, fontname="helv", color=(0,0,0))
    page.insert_text(fitz.Point(margin_pt, tb_y+36),
                      f"Project: {project_meta.get('customer','')}   {project_meta.get('job_number','')}",
                      fontsize=8, color=(0.2,0.2,0.2))
    page.insert_text(fitz.Point(margin_pt, tb_y+50),
                      f"Date: {datetime.date.today().strftime('%b %d, %Y')}", fontsize=8, color=(0.2,0.2,0.2))

    # Device schedule page(s) — spells out what each on-diagram abbreviation
    # tick actually is (full name, manufacturer part #, mA, quantity) per
    # circuit, so the diagram is usable by a contractor without this app.
    circuits = [n for n in scene.all_nodes() if isinstance(n, CircuitNode)]
    if circuits:
        sched_pw, sched_ph = 8.5*72, 11*72
        row_h = 14
        state = {"page": None, "y": 0.0}

        def _new_sched_page():
            state["page"] = doc.new_page(width=sched_pw, height=sched_ph)
            state["page"].insert_text(fitz.Point(margin_pt, margin_pt+14), "DEVICE SCHEDULE",
                                       fontsize=13, fontname="helv", color=(0,0,0))
            state["page"].insert_text(
                fitz.Point(margin_pt, margin_pt+30),
                f"Project: {project_meta.get('customer','')}   {project_meta.get('job_number','')}",
                fontsize=8, color=(0.2,0.2,0.2))
            state["y"] = margin_pt + 54

        _new_sched_page()
        for c in circuits:
            info = CIRCUIT_TYPE_INFO[c.circuit_type]
            if state["y"] > sched_ph - margin_pt - 60:
                _new_sched_page()
            state["page"].insert_text(
                fitz.Point(margin_pt, state["y"]),
                f"{c.name}  —  {info['name']} · Class {c.circuit_class}  —  "
                f"{c.total_load():.0f}/{c.capacity:.0f} {info['unit']} ({c.utilization_pct():.0f}%)",
                fontsize=9.5, fontname="helv", color=(0,0,0))
            state["y"] += row_h + 4
            if not c.devices:
                state["page"].insert_text(fitz.Point(margin_pt+14, state["y"]), "(no devices)",
                                           fontsize=8, color=(0.6,0.6,0.6))
                state["y"] += row_h
            for d in c.devices:
                if state["y"] > sched_ph - margin_pt - 30:
                    _new_sched_page()
                dev = FA_DEVICE_TYPES.get(d["key"], {})
                ma = d.get("ma")
                if ma is None:
                    ma = dev.get("ma", 0.0)
                label = f"{dev.get('abbr','?')}   {dev.get('name', d['key'])}"
                part = dev.get("part")
                if part:
                    label += f"   [{part}]"
                line = f"{label}   x{d['qty']}   @ {ma:.0f} mA ea   = {d['qty']*ma:.0f} mA"
                state["page"].insert_text(fitz.Point(margin_pt+14, state["y"]), line,
                                           fontsize=8, color=(0.1,0.1,0.1))
                state["y"] += row_h
            state["y"] += 10

    doc.save(path)
    doc.close()


class OneLinePreviewDialog(QDialog):
    """Export preview — a throwaway clone of the diagram the user can drag
    and resize freely (via the same drag/resize behavior as the live canvas)
    before generating the PDF. Nothing here touches the working diagram."""
    def __init__(self, source_scene, project_meta, sheet_title, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Export Preview — arrange boxes, then Export")
        self.resize(1150, 780)
        self.project_meta = project_meta
        self.sheet_title = sheet_title
        self.export_path = None

        self.scene = OneLineScene()
        self.scene.load_dict(source_scene.to_dict())

        l = QVBoxLayout(self); l.setContentsMargins(8,8,8,8); l.setSpacing(6)
        hint = QLabel("Drag a box to move it · drag its bottom-right corner to resize it · "
                       "this only affects the exported PDF, not your working diagram.")
        hint.setStyleSheet("color:#666;font-size:11px;")
        l.addWidget(hint)
        self.canvas = OneLineCanvas(self.scene)
        l.addWidget(self.canvas, 1)

        btns = QHBoxLayout()
        auto_btn = QPushButton("Auto-Arrange"); auto_btn.clicked.connect(self.scene.auto_arrange)
        btns.addWidget(auto_btn)
        fit_btn = QPushButton("Fit View"); fit_btn.clicked.connect(self.canvas.fit_all)
        btns.addWidget(fit_btn)
        btns.addStretch()
        cancel_btn = QPushButton("Cancel"); cancel_btn.clicked.connect(self.reject)
        btns.addWidget(cancel_btn)
        export_btn = QPushButton("Export PDF…")
        export_btn.setStyleSheet("background:#ff7002;color:white;padding:6px 18px;font-weight:bold;")
        export_btn.clicked.connect(self._do_export)
        btns.addWidget(export_btn)
        l.addLayout(btns)

        QTimer.singleShot(0, self.canvas.fit_all)

    def _do_export(self):
        cust = self.project_meta.get("customer","OneLine") or "OneLine"
        bad = re.compile(r'[<>:"/\\|?*\s]+')
        default_name = bad.sub("_", cust) + "_OneLine.pdf"
        path = os.path.join(_submittals_dir(), default_name)
        path, _ = QFileDialog.getSaveFileName(self, "Export One-Line Diagram PDF", path, "PDF Files (*.pdf)")
        if not path:
            return
        try:
            export_oneline_pdf(self.scene, path, self.project_meta, self.sheet_title)
        except Exception as e:
            _log_error("OneLinePreviewDialog._do_export", e)
            QMessageBox.critical(self, "Export Failed", str(e))
            return
        self.export_path = path
        self.accept()


# ═══════════════════════════════════════════════════════════════════════════════
#  Main window
# ═══════════════════════════════════════════════════════════════════════════════

class DrawingDesigner(QDialog):
    def __init__(self, parent=None, project_name=""):
        super().__init__(parent)
        self.setMinimumSize(1150, 700)
        self.setWindowState(Qt.WindowMaximized)
        self.setWindowTitle("DFP TakeoffPro — Drawing Designer")
        self.project_name = project_name
        self._project_meta = {"customer": project_name}
        self._current_file = None
        self._dirty = False

        self.fp_scene = FloorPlanScene()
        self.fp_scene.layout_changed.connect(self._on_changed)
        self.fp_scene.status_changed.connect(self._on_status)
        self.fp_canvas = FloorPlanCanvas(self.fp_scene)

        self.wire_scene = WiringScene()
        self.wire_scene.layout_changed.connect(self._on_changed)
        self.wire_scene.status_changed.connect(self._on_status)
        self.wire_canvas = WiringCanvas(self.wire_scene)

        self.oneline_scene = OneLineScene()
        self.oneline_scene.layout_changed.connect(self._on_changed)
        self.oneline_scene.status_changed.connect(self._on_status)
        self.oneline_canvas = OneLineCanvas(self.oneline_scene)

        self._build_ui()
        self._update_title()

    # ── UI construction ─────────────────────────────────────────────────────

    def _build_ui(self):
        main = QVBoxLayout(self); main.setContentsMargins(0,0,0,0); main.setSpacing(0)

        tbl = FlowBar(hspacing=5, vspacing=5); tbl.setStyleSheet("background:#232728;")
        tbl.setContentsMargins(8,5,8,5)

        def _btn(txt, slot, color="#333738", checkable=False):
            b = QPushButton(txt); b.setCheckable(checkable)
            b.setStyleSheet(f"QPushButton{{background:{color};color:#efe6e1;border-radius:3px;"
                            f"padding:5px 10px;border:none;font-weight:bold;font-size:11px;}}"
                            f"QPushButton:hover{{background:#ff7002;color:white;}}"
                            f"QPushButton:checked{{background:#ff7002;color:white;}}")
            b.clicked.connect(slot); return b

        def _sep():
            s = QFrame(); s.setFrameShape(QFrame.VLine); s.setStyleSheet("color:#555;"); return s

        # Floor Plan drawing tools — grouped behind one dropdown since every
        # one of these only applies on the Floor Plan tab (see
        # _on_tab_changed, which enables/disables the whole group at once
        # instead of 6 separate buttons individually).
        # Menus kept as self.* attributes (not locals) — Qt's C++ parent-
        # child ownership is supposed to keep them alive on its own once
        # attached via setMenu(), but a persisted Python reference is a
        # cheap extra guard against the classic "menu/action + its
        # connected lambda got garbage-collected out from under Qt" crash.
        self._fp_tools_btn = HoverMenuButton("Floor Plan Tools ▾", color="#1a5276")
        self._fp_menu = QMenu(self._fp_tools_btn); self._fp_menu.setStyleSheet(TOOLBAR_MENU_STYLE)
        self._wall_btn = self._fp_menu.addAction("📐 Draw Wall")
        self._wall_btn.setCheckable(True); self._wall_btn.triggered.connect(self._toggle_wall_tool)
        self._thickness_btn = self._fp_menu.addAction("Wall: 4.5\"")
        self._thickness_btn.triggered.connect(self._pick_wall_thickness)
        self._fp_menu.addAction("▭ Add Rectangle Room").triggered.connect(self._add_rect_room)
        self._fp_menu.addSeparator()
        self._snap_btn = self._fp_menu.addAction("Snap ON")
        self._snap_btn.setCheckable(True); self._snap_btn.setChecked(True)
        self._snap_btn.triggered.connect(self._toggle_snap)
        self._grid_btn = self._fp_menu.addAction("Grid ON")
        self._grid_btn.setCheckable(True); self._grid_btn.setChecked(True)
        self._grid_btn.triggered.connect(self._toggle_grid)
        self._fp_menu.addSeparator()
        self._zone_btn = self._fp_menu.addAction("🟥 Draw Fire Zone")
        self._zone_btn.setCheckable(True); self._zone_btn.triggered.connect(self._toggle_zone_tool)
        self._fp_tools_btn.setMenu(self._fp_menu)
        tbl.addWidget(self._fp_tools_btn)

        # Background trace tools — also Floor-Plan-tab only.
        self._bg_tools_btn = HoverMenuButton("Background Trace ▾", color="#7d3c98")
        self._bg_menu = QMenu(self._bg_tools_btn); self._bg_menu.setStyleSheet(TOOLBAR_MENU_STYLE)
        self._bg_menu.addAction("Import Background…").triggered.connect(self._import_background)
        self._bg_calibrate_btn = self._bg_menu.addAction("Calibrate Scale")
        self._bg_calibrate_btn.setCheckable(True); self._bg_calibrate_btn.triggered.connect(self._toggle_bg_calibrate)
        self._bg_menu.addAction("Set Scale…").triggered.connect(self._set_bg_scale_manual)
        self._bg_visible_btn = self._bg_menu.addAction("Background ON")
        self._bg_visible_btn.setCheckable(True); self._bg_visible_btn.setChecked(True)
        self._bg_visible_btn.triggered.connect(self._toggle_bg_visible)
        self._bg_menu.addAction("Remove Background").triggered.connect(self._remove_background)
        self._bg_tools_btn.setMenu(self._bg_menu)
        tbl.addWidget(self._bg_tools_btn)
        tbl.addWidget(_sep())

        self._connect_btn = _btn("🔌 Connect Wires", self._toggle_connect_tool, color="#7d3c98", checkable=True)
        tbl.addWidget(self._connect_btn)
        tbl.addWidget(_sep())

        # One-Line Diagram tools — grouped the same way; Connect Wires above
        # stays its own button since it's the only Wiring-tab-specific tool
        # (not worth a 1-item dropdown).
        self._ol_tools_btn = HoverMenuButton("One-Line Tools ▾", color="#1a5276")
        self._ol_menu = QMenu(self._ol_tools_btn); self._ol_menu.setStyleSheet(TOOLBAR_MENU_STYLE)
        self._add_loop_btn = self._ol_menu.addAction("+ SLC Loop")
        self._add_loop_btn.triggered.connect(lambda: self._ol_add_circuit("slc"))
        self._add_nac_btn = self._ol_menu.addAction("+ NAC Circuit")
        self._add_nac_btn.triggered.connect(lambda: self._ol_add_circuit("nac"))
        self._add_booster_btn = self._ol_menu.addAction("+ Booster")
        self._add_booster_btn.triggered.connect(self._ol_add_booster)
        self._add_standalone_btn = self._ol_menu.addAction("+ Standalone NAC Circuit")
        self._add_standalone_btn.setToolTip(
            "A NAC circuit with no panel/booster box on this diagram — right-click it to set a\n"
            "free-text source label (e.g. \"Panel — NAC 3\") instead of drawing a connector line.")
        self._add_standalone_btn.triggered.connect(lambda: self._ol_add_standalone_circuit("nac"))
        self._ol_menu.addSeparator()
        self._loading_btn = self._ol_menu.addAction("⚡ Check Loading")
        self._loading_btn.triggered.connect(self._ol_check_loading)
        self._ol_menu.addAction("Auto-Arrange").triggered.connect(self._ol_auto_arrange)
        self._ol_menu.addAction("Import FQQ (.xlsm)").triggered.connect(self._import_fqq)
        self._ol_tools_btn.setMenu(self._ol_menu)
        tbl.addWidget(self._ol_tools_btn)
        tbl.addWidget(_sep())

        tbl.addWidget(_btn("Fit View", self._fit))
        tbl.addWidget(_btn("Clear Tab", self._clear_current))
        tbl.addWidget(_sep())
        tbl.addWidget(_btn("Project Info", self._edit_project_info))
        tbl.addWidget(_btn("New", self._new_project))
        tbl.addWidget(_btn("Open", self._open_project, color="#1a5276"))
        tbl.addWidget(_btn("Save", self._save_project, color="#1a5276"))
        tbl.addWidget(_btn("Save As", self._save_project_as))
        tbl.addWidget(_sep())
        tbl.addWidget(_btn("Export PDF", self._export_pdf, color="#c0392b"))
        tbl.addWidget(_btn("🚨 Export Fire Safety Plan", self._export_fire_safety_plan, color="#922b21"))
        tbl.addWidget(_sep())
        tbl.addWidget(_btn("Help", self._show_help, color="#2c3e50"))
        tbl.addStretch()
        self._mode_lbl = QLabel("  Select  ·  Scroll=zoom  ·  Middle-drag or Ctrl+drag=pan  ·  Del=delete  ·  Right-click=options  ·  Esc=cancel")
        self._mode_lbl.setStyleSheet("color:#888;font-size:10px;")
        tbl.addWidget(self._mode_lbl)
        main.addWidget(tbl)

        body = QSplitter(Qt.Horizontal)
        body.setStyleSheet("QSplitter::handle{background:#d0c8c0;width:4px;}")

        self.layer_panel = LayerPanel()
        self.layer_panel.visibility_changed.connect(self._on_layer_toggle)
        body.addWidget(self.layer_panel)

        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(
            "QTabWidget::pane{border:none;margin:0;}"
            "QTabBar::tab{background:#3a3d3e;color:#efe6e1;padding:6px 14px;min-width:100px;"
            "border:1px solid #555;border-bottom:none;"
            "border-top-left-radius:3px;border-top-right-radius:3px;"
            "font-size:11px;font-weight:bold;}"
            "QTabBar::tab:selected{background:#ff7002;color:white;}"
            "QTabBar::tab:hover:!selected{background:#555;}")
        # Tabs default to expanding to fill the pane, which was shrinking
        # "Wiring Diagram" / "One-Line Diagram" below their text width and
        # clipping them raggedly (no ellipsis) when the palette/layer panels
        # left little room. Fixed-size tabs + scroll arrows if truly tight
        # keep the labels always fully readable.
        self.tabs.tabBar().setExpanding(False)
        self.tabs.tabBar().setUsesScrollButtons(True)
        self.tabs.setElideMode(Qt.ElideNone)
        self.tabs.addTab(self.fp_canvas, "Floor Plan")
        self.tabs.addTab(self.wire_canvas, "Wiring Diagram")
        self.tabs.addTab(self.oneline_canvas, "One-Line Diagram")
        self.tabs.currentChanged.connect(self._on_tab_changed)
        body.addWidget(self.tabs)

        self.palette = SymbolPalette()
        self.palette.symbol_clicked.connect(self._on_symbol_clicked)
        body.addWidget(self.palette)

        body.setStretchFactor(0, 0); body.setStretchFactor(1, 1); body.setStretchFactor(2, 0)
        main.addWidget(body, 1)

        self._status = QLabel("  Draw a wall, drop a room, or place symbols from the right-hand palette.")
        self._status.setStyleSheet("background:#efe6e1;color:#333;padding:4px 8px;font-size:10px;")
        main.addWidget(self._status)

        # QTabWidget's own currentChanged(0) fires while adding the FIRST tab
        # (before the connect() above even runs), so without this every
        # tab-specific button/group starts out enabled regardless of which
        # tab is actually active until the user manually switches tabs once.
        self._on_tab_changed(self.tabs.currentIndex())

    # ── Status / mode plumbing ──────────────────────────────────────────────

    def _active_scene(self):
        if self.tabs.currentIndex() == 0: return self.fp_scene
        if self.tabs.currentIndex() == 1: return self.wire_scene
        return self.oneline_scene

    def _on_tab_changed(self, idx):
        self.layer_panel.setVisible(idx == 0)
        self.palette.setVisible(idx in (0, 1))
        self._fp_tools_btn.setEnabled(idx == 0)
        self._bg_tools_btn.setEnabled(idx == 0)
        self._ol_tools_btn.setEnabled(idx == 2)
        self._wall_btn.setEnabled(idx == 0); self._wall_btn.setChecked(False)
        self._thickness_btn.setEnabled(idx == 0)
        self._zone_btn.setEnabled(idx == 0); self._zone_btn.setChecked(False)
        self._bg_calibrate_btn.setEnabled(idx == 0); self._bg_calibrate_btn.setChecked(False)
        self._connect_btn.setEnabled(idx == 1); self._connect_btn.setChecked(False)
        for b in (self._add_loop_btn, self._add_nac_btn, self._add_booster_btn, self._loading_btn):
            b.setEnabled(idx == 2)
        self.palette.clear_all()
        self._active_scene().set_mode_select()
        if idx == 2:
            self._status.setText("  Right-click the panel/booster to add loops, NAC circuits, or boosters — "
                                  "or use the toolbar buttons. Right-click a circuit to edit its devices.")

    def _on_status(self, text):
        self._status.setText(text)

    def _on_changed(self):
        self._dirty = True
        self._update_title()

    def _on_layer_toggle(self, name, visible):
        self.fp_scene.set_layer_visible(name, visible)
        self._dirty = True; self._update_title()

    def _on_symbol_clicked(self, kind):
        sc = self._active_scene()
        if not kind:
            sc.set_mode_select(); return
        if sc is self.wire_scene:
            self._connect_btn.setChecked(False)
        else:
            self._wall_btn.setChecked(False)
            self._zone_btn.setChecked(False)
            self._bg_calibrate_btn.setChecked(False)
        sc.set_mode_symbol(kind)

    def _toggle_wall_tool(self, checked):
        self.palette.clear_all()
        if checked:
            self._connect_btn.setChecked(False)
            self._zone_btn.setChecked(False)
            self._bg_calibrate_btn.setChecked(False)
            self.fp_scene.set_mode_wall(self.fp_scene.default_wall_thickness_in)
        else:
            self.fp_scene.set_mode_select()
            self._status.setText("  Select  ·  Scroll=zoom  ·  Middle-drag or Ctrl+drag=pan  ·  Del=delete  ·  Right-click=options  ·  Esc=cancel")

    def _toggle_zone_tool(self, checked):
        self.palette.clear_all()
        if checked:
            self._connect_btn.setChecked(False)
            self._wall_btn.setChecked(False)
            self._bg_calibrate_btn.setChecked(False)
            self.fp_scene.set_mode_zone()
        else:
            self.fp_scene.set_mode_select()
            self._status.setText("  Select  ·  Scroll=zoom  ·  Middle-drag or Ctrl+drag=pan  ·  Del=delete  ·  Right-click=options  ·  Esc=cancel")

    def _import_background(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Background to Trace", "",
            "Drawings (*.pdf *.png *.jpg *.jpeg *.tif *.tiff *.bmp);;All Files (*)")
        if not path:
            return
        try:
            if path.lower().endswith(".pdf"):
                doc = fitz.open(path)
                page_num = 0
                if doc.page_count > 1:
                    num, ok = QInputDialog.getInt(
                        self, "Choose Page",
                        f"This PDF has {doc.page_count} pages — which page do you want to trace?",
                        1, 1, doc.page_count)
                    if not ok:
                        return
                    page_num = num - 1
                dpi = 150
                pix = doc[page_num].get_pixmap(dpi=dpi)
                fmt = QImage.Format_RGBA8888 if pix.alpha else QImage.Format_RGB888
                qimg = QImage(pix.samples, pix.width, pix.height, pix.stride, fmt).copy()
                qpix = QPixmap.fromImage(qimg)
                doc.close()
            else:
                qpix = QPixmap(path)
                if qpix.isNull():
                    raise ValueError("Couldn't read that as an image.")
                dpi = 150
            if qpix.isNull():
                raise ValueError("Nothing rendered from that file.")
            self.fp_scene.bg_pixmap = qpix
            self.fp_scene.bg_dpi = dpi
            self.fp_scene.bg_pos_ft = (0.0, 0.0)
            self.fp_scene.bg_px_per_ft = 10.0
            self.fp_scene.bg_visible = True
            self._bg_visible_btn.setChecked(True)
            self._bg_visible_btn.setText("Background ON")
            self.fp_scene.update()
            self._status.setText(
                "  Background imported (not saved with the project — it's a tracing aid only). "
                "Use 'Calibrate Scale' or 'Set Scale…' before tracing walls over it precisely.")
        except Exception as e:
            _log_error("_import_background", e)
            QMessageBox.critical(self, "Import Failed", str(e))

    def _toggle_bg_calibrate(self, checked):
        self.palette.clear_all()
        if checked:
            self._wall_btn.setChecked(False); self._zone_btn.setChecked(False)
            self._connect_btn.setChecked(False)
            self.fp_scene.set_mode_bg_calibrate()
        else:
            self.fp_scene.set_mode_select()
            self._status.setText("  Select  ·  Scroll=zoom  ·  Middle-drag or Ctrl+drag=pan  ·  Del=delete  ·  Right-click=options  ·  Esc=cancel")

    def _set_bg_scale_manual(self):
        if self.fp_scene.bg_pixmap is None:
            QMessageBox.information(self, "Set Scale", "Import a background image/PDF first.")
            return
        labels = [lbl for lbl, _fpi in SCALE_OPTIONS]
        label, ok = QInputDialog.getItem(
            self, "Set Background Scale",
            "Architectural scale the imported drawing was printed at:", labels, 0, False)
        if not ok:
            return
        fpi = dict(SCALE_OPTIONS)[label]
        self.fp_scene.bg_px_per_ft = self.fp_scene.bg_dpi / fpi
        self.fp_scene.update()
        self._status.setText(f"  Background scale set to {label}. Fine-tune with 'Calibrate Scale' if it's still off.")

    def _toggle_bg_visible(self, checked):
        self.fp_scene.bg_visible = checked
        self._bg_visible_btn.setText("Background ON" if checked else "Background OFF")
        self.fp_scene.update()

    def _remove_background(self):
        self.fp_scene.bg_pixmap = None
        self.fp_scene.update()
        self._status.setText("  Background removed.")

    def _add_rect_room(self):
        center = self.fp_canvas.mapToScene(self.fp_canvas.viewport().rect().center())
        dlg = RectRoomDialog(parent=self)
        if dlg.exec_() == QDialog.Accepted:
            v = dlg.values()
            x0_ft = center.x()/PX_PER_FT - v["width_ft"]/2
            y0_ft = center.y()/PX_PER_FT - v["length_ft"]/2
            self.fp_scene.add_rect_room(x0_ft, y0_ft, v["width_ft"], v["length_ft"], v["name"],
                                         v["ceiling_in"], v["ceiling_type"], v["floor_type"])
            self._dirty = True; self._update_title()

    def _toggle_connect_tool(self, checked):
        self.palette.clear_all()
        if checked:
            self._wall_btn.setChecked(False)
            self._zone_btn.setChecked(False)
            self._bg_calibrate_btn.setChecked(False)
            self.wire_scene.set_mode_connect()
        else:
            self.wire_scene.set_mode_select()
            self._status.setText("  Select  ·  Scroll=zoom  ·  Middle-drag or Ctrl+drag=pan  ·  Del=delete  ·  Right-click=options  ·  Esc=cancel")

    def _pick_wall_thickness(self):
        dlg = WallSettingsDialog(self.fp_scene.default_wall_thickness_in, self)
        if dlg.exec_() == QDialog.Accepted:
            t = dlg.value_in()
            self.fp_scene.default_wall_thickness_in = t
            self._thickness_btn.setText(f"Wall: {t:g}\"")
            if self._wall_btn.isChecked():
                self.fp_scene.set_mode_wall(t)

    def _toggle_snap(self, checked):
        self.fp_scene.snap_to_grid = checked
        self._snap_btn.setText("Snap ON" if checked else "Snap OFF")

    def _toggle_grid(self, checked):
        self.fp_scene.show_grid = checked
        self._grid_btn.setText("Grid ON" if checked else "Grid OFF")
        self.fp_scene.update()

    def _fit(self):
        canvases = [self.fp_canvas, self.wire_canvas, self.oneline_canvas]
        canvases[self.tabs.currentIndex()].fit_all()

    def _clear_current(self):
        sc = self._active_scene()
        ans = QMessageBox.question(self, "Clear Tab", "Clear everything on this tab?",
                                    QMessageBox.Yes | QMessageBox.No)
        if ans == QMessageBox.Yes:
            sc.clear_all()

    # ── One-Line Diagram ────────────────────────────────────────────────────

    def _ol_target_node(self):
        """The node new circuits/boosters should attach to: the selected
        Panel/Booster, or the panel itself if nothing suitable is selected."""
        sel = self.oneline_scene.selectedItems()
        for it in sel:
            if isinstance(it, (PanelNode, BoosterNode)):
                return it
        return self.oneline_scene.panel

    def _ol_add_circuit(self, circuit_type):
        self.oneline_scene.add_circuit(self._ol_target_node(), circuit_type)

    def _ol_add_standalone_circuit(self, circuit_type):
        self.oneline_scene.add_standalone_circuit(circuit_type)

    def _ol_add_booster(self):
        self.oneline_scene.add_booster(self._ol_target_node())

    def _ol_check_loading(self):
        self.oneline_scene.show_loading_report()

    def _ol_auto_arrange(self):
        self.oneline_scene.auto_arrange()

    def _import_fqq(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Autocall FQQ Quote", "",
                                               "Autocall FQQ (*.xlsm);;All Files (*)")
        if not path:
            return
        try:
            by_designation = import_fqq_xlsm(path)
        except Exception as e:
            QMessageBox.critical(self, "Import Failed",
                                  f"Couldn't read that file as an FQQ quote:\n{e}")
            return
        if not by_designation:
            QMessageBox.information(self, "Import FQQ",
                                     "No NAC device quantities were found in that file's NAC sheet.")
            return
        capacity = CIRCUIT_TYPE_INFO["nac"]["default_capacity"]
        sc = self.oneline_scene
        n_circuits = 0
        for dname, devices in by_designation.items():
            for i, tally in enumerate(bin_pack_nac_devices(devices, capacity), start=1):
                c = sc.add_circuit(sc.panel, "nac")
                c.name = f"{dname} NAC-{i}"; c.capacity = capacity; c.devices = tally
                c.circuit_class = "B"
                c.prepareGeometryChange(); c._recompute_height(); c.update()
                n_circuits += 1
        sc.auto_arrange()
        sc.layout_changed.emit()
        self._dirty = True
        self._update_title()
        QMessageBox.information(
            self, "Import FQQ",
            f"Imported {n_circuits} NAC circuit(s) across {len(by_designation)} designation(s) "
            f"from {os.path.basename(path)}.\n\n"
            "Devices were packed into circuits at the standard NAC capacity — review circuit "
            "capacity, Class A/B wiring, spare-capacity margin, and rearrange under booster(s) "
            "as needed for the actual job.")

    # ── Project menu ─────────────────────────────────────────────────────────

    def _edit_project_info(self):
        dlg = ProjectInfoDialog(self._project_meta, self)
        if dlg.exec_() == QDialog.Accepted:
            self._project_meta.update(dlg.values())
            self._dirty = True; self._update_title()

    def _check_unsaved(self):
        if not self._dirty:
            return True
        ans = QMessageBox.question(self, "Unsaved Changes", "You have unsaved changes. Save before continuing?",
                                    QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
        if ans == QMessageBox.Save:
            return self._save_project()
        return ans == QMessageBox.Discard

    def _new_project(self):
        if not self._check_unsaved():
            return
        dlg = ProjectInfoDialog(parent=self)
        if dlg.exec_() != QDialog.Accepted:
            return
        self._project_meta = dlg.values()
        self._current_file = None
        self.fp_scene.clear_all(); self.wire_scene.clear_all()
        self._dirty = False; self._update_title()

    def _project_to_dict(self):
        return {
            "version": 1,
            "meta": self._project_meta,
            "saved": datetime.datetime.now().isoformat(),
            "floor_plan": self.fp_scene.to_dict(),
            "wiring": self.wire_scene.to_dict(),
            "oneline": self.oneline_scene.to_dict(),
        }

    def _load_project_dict(self, d):
        self._project_meta = d.get("meta", {})
        self.fp_scene.load_dict(d.get("floor_plan", {}))
        self.wire_scene.load_dict(d.get("wiring", {}))
        self.oneline_scene.load_dict(d.get("oneline", {}))
        for name in LAYER_ORDER:
            self.layer_panel.set_checked(name, self.fp_scene.layers_visible.get(name, True))
        self._dirty = False; self._update_title()
        self.fp_canvas.fit_all()

    def _save_project(self):
        if not self._current_file:
            return self._save_project_as()
        try:
            with open(self._current_file, "w", encoding="utf-8") as f:
                json.dump(self._project_to_dict(), f, indent=2)
            self._dirty = False; self._update_title()
            return True
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", str(e)); return False

    def _save_project_as(self):
        meta = self._project_meta
        bad = re.compile(r'[<>:"/\\|?*\s]+')
        cust = bad.sub("_", meta.get("customer","Floor Plan")).strip("_") or "Floor_Plan"
        default = f"{cust}.dfpplan"
        name, ok = QInputDialog.getText(self, "Save Project As", "Filename:", text=default)
        if not ok or not name.strip():
            return False
        name = name.strip()
        if not name.endswith(".dfpplan"):
            name += ".dfpplan"
        self._current_file = os.path.join(_projects_dir(), name)
        return self._save_project()

    def _open_project(self):
        if not self._check_unsaved():
            return
        path, _ = QFileDialog.getOpenFileName(self, "Open Floor Plan", _projects_dir(),
                                               "Floor Plan Project (*.dfpplan)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
            self._load_project_dict(d)
            self._current_file = path
        except Exception as e:
            QMessageBox.critical(self, "Open Error", str(e))

    def _update_title(self):
        fn = os.path.basename(self._current_file) if self._current_file else "Untitled"
        dirty = " •" if self._dirty else ""
        cust = self._project_meta.get("customer","")
        parts = [p for p in (cust,) if p]
        self.setWindowTitle(f"DFP TakeoffPro v{APP_VERSION}  —  Drawing Designer   |   "
                             f"{' · '.join(parts)}{'  ' if parts else ''}[{fn}]{dirty}")

    # ── Export ───────────────────────────────────────────────────────────────

    def _export_pdf(self):
        idx = self.tabs.currentIndex()
        if idx == 1:
            QMessageBox.information(self, "Export PDF", "PDF export isn't available for the Wiring Diagram tab yet.\n"
                                     "Switch to Floor Plan or One-Line Diagram to export.")
            return
        if idx == 2:
            self._export_oneline_pdf()
            return
        if not self.fp_scene._wall_items and not self.fp_scene._symbol_items:
            QMessageBox.information(self, "Nothing to Export", "Draw something on the floor plan first.")
            return
        dlg = PrintExportDialog(self.fp_scene, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        vals = dlg.values()
        cust = self._project_meta.get("customer","Floor_Plan") or "Floor_Plan"
        bad = re.compile(r'[<>:"/\\|?*\s]+')
        default_name = bad.sub("_", cust) + "_FloorPlan.pdf"
        path = os.path.join(_submittals_dir(), default_name)
        path, _ = QFileDialog.getSaveFileName(self, "Export Floor Plan PDF", path, "PDF Files (*.pdf)")
        if not path:
            return
        try:
            export_floor_plan_pdf(self.fp_scene, path, vals["paper_key"], vals["orientation"],
                                   vals["feet_per_inch"], vals["region_bbox"], self._project_meta,
                                   sheet_title=self._project_meta.get("customer","FLOOR PLAN") or "FLOOR PLAN")
            QMessageBox.information(self, "Exported", f"PDF saved:\n{path}")
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as e:
            _log_error("_export_pdf", e)
            QMessageBox.critical(self, "Export Failed", str(e))

    def _export_fire_safety_plan(self):
        """Dedicated export for the code-required Fire Safety Plan / 'You Are
        Here' evacuation diagram — reuses the same Floor Plan geometry as a
        regular export (so walls/rooms stay single-sourced), but forces the
        Fire Safety Plan layer on, requires a 'You Are Here' marker be placed
        first, and titles/names the sheet distinctly from a generic export."""
        if self.tabs.currentIndex() != 0:
            self.tabs.setCurrentIndex(0)
        scene = self.fp_scene
        if not scene._wall_items and not scene._symbol_items:
            QMessageBox.information(self, "Nothing to Export",
                                     "Draw the building's walls/rooms and place Fire Safety Plan "
                                     "symbols first.")
            return
        has_you_are_here = any(s.kind == "fsp_you_are_here" for s in scene._symbol_items)
        if not has_you_are_here:
            ans = QMessageBox.question(
                self, "Missing 'You Are Here' Marker",
                "No 'You Are Here' marker has been placed on this plan. A Fire Safety Plan "
                "normally requires one at every posted location.\n\n"
                "Export anyway?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans != QMessageBox.Yes:
                return
        if not scene._zone_items:
            ans = QMessageBox.question(
                self, "No Fire Zones Drawn",
                "No fire zone / separation boundaries have been drawn. If this building isn't "
                "zoned that's fine — otherwise draw them with the Fire Zone tool first.\n\n"
                "Export anyway?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if ans != QMessageBox.Yes:
                return
        dlg = FireSafetyPlanExportDialog(scene, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        vals = dlg.values()
        cust = self._project_meta.get("customer","Fire_Safety_Plan") or "Fire_Safety_Plan"
        bad = re.compile(r'[<>:"/\\|?*\s]+')
        default_name = bad.sub("_", cust) + "_FireSafetyPlan.pdf"
        path = os.path.join(_submittals_dir(), default_name)
        path, _ = QFileDialog.getSaveFileName(self, "Export Fire Safety Plan PDF", path, "PDF Files (*.pdf)")
        if not path:
            return
        was_fire_visible = scene.layers_visible.get(LAYER_FIRE, True)
        scene.layers_visible[LAYER_FIRE] = True
        try:
            export_fire_safety_plan_pdf(scene, path, vals["paper_key"], vals["orientation"],
                                         self._project_meta, sheet_title="FIRE SAFETY PLAN",
                                         symbol_scale=vals["symbol_scale"], text_pt=vals["text_pt"])
            QMessageBox.information(self, "Exported", f"PDF saved:\n{path}")
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as e:
            _log_error("_export_fire_safety_plan", e)
            QMessageBox.critical(self, "Export Failed", str(e))
        finally:
            scene.layers_visible[LAYER_FIRE] = was_fire_visible

    def _export_oneline_pdf(self):
        if len(self.oneline_scene.all_nodes()) <= 1 and not self.oneline_scene.standalone:
            QMessageBox.information(self, "Nothing to Export", "Add a loop, NAC circuit, or booster first.")
            return
        sheet_title = self._project_meta.get("customer","ONE-LINE DIAGRAM") or "ONE-LINE DIAGRAM"
        dlg = OneLinePreviewDialog(self.oneline_scene, self._project_meta, sheet_title, self)
        if dlg.exec_() == QDialog.Accepted and dlg.export_path:
            QMessageBox.information(self, "Exported", f"PDF saved:\n{dlg.export_path}")
            try:
                os.startfile(dlg.export_path)
            except Exception:
                pass

    def _show_help(self):
        from help_system import HelpDialog, DRAWING_DESIGNER_MANUAL
        dlg = HelpDialog(DRAWING_DESIGNER_MANUAL, "Drawing Designer", self)
        dlg.exec_()

    def closeEvent(self, event):
        if self._check_unsaved():
            event.accept()
        else:
            event.ignore()
