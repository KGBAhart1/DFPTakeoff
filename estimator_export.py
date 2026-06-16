"""
Excel export for DFP TakeoffPro estimator data.
Produces one workbook per project with:
  - PMA Quote sheet (with Uptick cost/sell input columns)
  - One Install Estimate sheet per estimate
  - Programming & V.I. sheet
Blue cells = values to copy into Uptick.
"""

import db

DISCIPLINES = [
    "Fire Alarm", "Sprinkler", "Emergency Lighting",
    "Extinguisher / Hose", "Kitchen Suppression",
]

DISC_RATE_KEY = {
    "Fire Alarm": "fa", "Emergency Lighting": "fa",
    "Sprinkler": "sp",
    "Extinguisher / Hose": "sg", "Kitchen Suppression": "sg",
}
DISC_MARGIN_KEY = {
    "Fire Alarm": "margin_fa", "Emergency Lighting": "margin_fa",
    "Sprinkler": "margin_sp",
    "Extinguisher / Hose": "margin_sg", "Kitchen Suppression": "margin_sg",
}


def export_estimates(project_id, output_path, discipline=None):
    """Build the export workbook. Returns (True, path) or (False, error)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ORANGE = "FF7002"; DARK = "232728"; GREY = "F2F2F2"
        DBLUE  = "2F5496"; GREEN = "E2EFDA"; DGREY = "D0D0D0"
        WHITE  = "FFFFFF"; LBLUE = "D9E1F2"; LGREEN = "C6EFCE"

        thin = Side(style="thin",  color="BDBDBD")
        med  = Side(style="medium", color=DARK)

        def bdr():  return Border(left=thin, right=thin, top=thin, bottom=thin)
        def bdr_b(): return Border(left=thin, right=thin, top=thin, bottom=med)
        def fill(c): return PatternFill("solid", fgColor=c)
        def ft(bold=False, size=10, color=DARK):
            return Font(name="Arial", bold=bold, size=size, color=color)
        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def set_cell(ws, row, col, val, bold=False, bg=WHITE, color=DARK,
                     fmt=None, align="left", wrap=False, border=True):
            c = ws.cell(row=row, column=col, value=val)
            c.font = ft(bold=bold, color=color)
            c.fill = fill(bg)
            c.alignment = al(align, wrap=wrap)
            if border:
                c.border = bdr()
            if fmt:
                c.number_format = fmt
            return c

        def title_row(ws, row, text, ncols=12):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = ft(bold=True, size=13, color=WHITE)
            c.fill = fill(DARK)
            c.alignment = al("center")
            ws.row_dimensions[row].height = 28

        def section_row(ws, row, text, ncols=12):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = ft(bold=True, size=11, color=WHITE)
            c.fill = fill(DBLUE)
            c.alignment = al("left")
            ws.row_dimensions[row].height = 20

        def hdr(ws, row, col, text, width=None):
            c = ws.cell(row=row, column=col, value=text)
            c.font = ft(bold=True, size=9, color=WHITE)
            c.fill = fill(DARK)
            c.alignment = al("center")
            c.border = bdr()
            ws.row_dimensions[row].height = 20
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width

        def inp(ws, row, col, val, fmt=None):
            """Blue input cell — for Uptick data entry."""
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", color="0070C0")
            c.fill = fill(LBLUE)
            c.alignment = al("center")
            c.border = bdr()
            if fmt:
                c.number_format = fmt
            return c

        def tot(ws, row, col, val, fmt=None, merge_to=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font = ft(bold=True)
            c.fill = fill(DGREY)
            c.alignment = al("center")
            c.border = bdr_b()
            if fmt:
                c.number_format = fmt
            if merge_to:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=merge_to)
            return c

        # ── Project name ──────────────────────────────────────────────────
        proj_name = ""
        with db.get_conn() as conn:
            prow = conn.execute(
                "SELECT name FROM projects WHERE id=?", (project_id,)
            ).fetchone()
            if prow:
                proj_name = prow["name"]

        wb = Workbook()
        wb.remove(wb.active)

        # ═══════════════════════════════════════════════════════════════════
        # SHEET: PMA QUOTE
        # ═══════════════════════════════════════════════════════════════════
        with db.get_conn() as conn:
            qrow = conn.execute(
                "SELECT * FROM pma_quotes WHERE project_id=? ORDER BY id DESC LIMIT 1",
                (project_id,)
            ).fetchone()

        if qrow:
            q = dict(qrow)
            ws = wb.create_sheet("PMA Quote")
            ws.sheet_view.showGridLines = False
            ws.column_dimensions["A"].width = 3

            title_row(ws, 1, f"DEFENSE FIRE PROTECTION  |  PMA INSPECTION QUOTE  |  {proj_name}")
            r = 2

            # Quote header info
            info_pairs = [
                ("Customer:", q.get("customer", "")),
                ("Address:",  q.get("address",  "")),
                ("Term (years):", q.get("term_years", 5)),
                ("Escalation/yr:", f"{q.get('escalation', 0.03)*100:.1f}%"),
                ("Site Multiplier:", q.get("site_multiplier", 1.0)),
            ]
            for lbl_text, val in info_pairs:
                ws.row_dimensions[r].height = 17
                set_cell(ws, r, 2, lbl_text, bold=True, bg=GREY, align="right")
                c = ws.cell(row=r, column=3, value=val)
                c.font = Font(name="Arial", color="0070C0")
                c.fill = fill(LBLUE)
                c.alignment = al("left")
                c.border = bdr()
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
                r += 1

            r += 1
            # Rate + margin info (right side)
            rates_info = [
                ("FA Lead Rate (burdened):", f"${q.get('rate_fa_lead', 90):.0f}/hr"),
                ("FA Helper Rate (burdened):", f"${q.get('rate_fa_help', 65):.0f}/hr"),
                ("SP Rate (burdened):", f"${q.get('rate_sp', 110):.0f}/hr"),
                ("SG/EXT Rate (burdened):", f"${q.get('rate_sg', 65):.0f}/hr"),
                ("FA Margin:", f"{q.get('margin_fa', 0.35)*100:.0f}%"),
                ("SP Margin:", f"{q.get('margin_sp', 0.32)*100:.0f}%"),
                ("SG Margin:", f"{q.get('margin_sg', 0.32)*100:.0f}%"),
            ]
            for lbl_text, val in rates_info:
                set_cell(ws, r, 8, lbl_text, bold=True, bg=GREY, align="right")
                ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
                set_cell(ws, r, 10, val, bg=LBLUE, align="center", color="0070C0")
                ws.row_dimensions[r].height = 17
                r += 1

            site_mult = q.get("site_multiplier", 1.0)

            disc_rate_map = {
                "Fire Alarm":           (q["rate_fa_lead"] + q["rate_fa_help"]) / 2,
                "Emergency Lighting":   (q["rate_fa_lead"] + q["rate_fa_help"]) / 2,
                "Sprinkler":             q["rate_sp"],
                "Extinguisher / Hose":  q["rate_sg"],
                "Kitchen Suppression":  q["rate_sg"],
            }
            disc_margin_map = {
                "Fire Alarm":           q["margin_fa"],
                "Emergency Lighting":   q["margin_fa"],
                "Sprinkler":            q["margin_sp"],
                "Extinguisher / Hose":  q["margin_sg"],
                "Kitchen Suppression":  q["margin_sg"],
            }

            import json as _json
            LABOUR_RATE_KH = q.get("rate_sg", 65.0); LINK_COST_KH = 9.0
            CART_COSTS_KH  = {"cart_rg": 14.0, "cart_a12": 12.0,
                               "cart_pyro": 28.0, "cart_buck": 30.0}
            CART_NAMES_KH  = {"cart_rg": "RG Test Cartridge",
                               "cart_a12": "A+ 12g CO2 Cartridge",
                               "cart_pyro": "Pyrochem 16g CO2 Cartridge",
                               "cart_buck": "Buckeye AC-s Cartridge"}
            KH_LOCATIONS   = ["Calgary", "Cochrane", "Airdrie", "Okotoks", "Other"]

            summary_sells = {}
            discs_to_export = [disc for disc in DISCIPLINES
                               if discipline is None or disc == discipline]

            for disc in discs_to_export:
                # ── Kitchen Suppression: KH calculator layout ──────────────
                if disc == "Kitchen Suppression":
                    kh_raw = q.get("kh_data")
                    try:
                        kh = _json.loads(kh_raw) if kh_raw else {}
                    except Exception:
                        kh = {}
                    num_sys    = kh.get("num_systems", 1)
                    loc_idx    = kh.get("location_idx", 0)
                    travel_hrs = kh.get("travel_hrs", 0.5)
                    loc_name   = KH_LOCATIONS[loc_idx] if 0 <= loc_idx < len(KH_LOCATIONS) else "Other"
                    systems    = kh.get("systems", [])
                    total_sell_ks = 0.0

                    r += 1
                    section_row(ws, r, "  KITCHEN SUPPRESSION — KH INSPECTIONS PRICING")
                    r += 1

                    # Info row: location + travel
                    ws.row_dimensions[r].height = 16
                    set_cell(ws, r, 1, "Location:", bold=True, bg=GREY)
                    set_cell(ws, r, 2, f"{loc_name}  (travel: {travel_hrs:.3f} hr one-way)",
                             bg=GREY, wrap=True)
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
                    set_cell(ws, r, 6, "Systems on site:", bold=True, bg=GREY)
                    set_cell(ws, r, 7, num_sys, align="center", bg=GREY)
                    r += 1

                    for si in range(num_sys):
                        sd = systems[si] if si < len(systems) else {}
                        on_site  = sd.get("on_site_hrs", 1.5) * site_mult
                        links    = sd.get("fusible_links", 6)
                        carts    = sd.get("carts", {})
                        margin_s = sd.get("margin", 35.0) / 100.0

                        labour   = on_site * LABOUR_RATE_KH
                        if si == 0:
                            labour += travel_hrs * LABOUR_RATE_KH
                        link_cost  = links * LINK_COST_KH
                        cart_cost  = sum(CART_COSTS_KH[k] for k, v in carts.items() if v)
                        active_carts = [CART_NAMES_KH[k] for k, v in carts.items() if v]
                        cogs     = labour + link_cost + cart_cost
                        sell_s   = cogs / (1 - margin_s) if margin_s < 1 else cogs
                        total_sell_ks += sell_s

                        sys_title = f"  System {si+1}" + ("  (includes call-out / travel fee)" if si == 0 else "")
                        ws.row_dimensions[r].height = 16
                        c = ws.cell(row=r, column=1, value=sys_title)
                        c.font = ft(bold=True, color=WHITE, size=9)
                        c.fill = fill("5D6D7E"); c.alignment = al("left"); c.border = bdr()
                        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
                        r += 1

                        detail_rows = [
                            ("On-site hours:",    f"{on_site:.2f} hr"),
                            ("Travel / call-out:", f"{travel_hrs:.3f} hr" if si == 0 else "n/a"),
                            ("Fusible links:",    f"{links}  ×  $9.00 = ${links*LINK_COST_KH:.2f}"),
                            ("Cartridges:",       ", ".join(active_carts) if active_carts else "None"),
                            ("Labour cost:",      f"${labour:.2f}"),
                            ("Cartridge cost:",   f"${cart_cost:.2f}"),
                            ("COGS:",             f"${cogs:.2f}"),
                            ("Gross margin:",     f"{margin_s*100:.1f}%"),
                        ]
                        for lbl, val in detail_rows:
                            ws.row_dimensions[r].height = 15
                            set_cell(ws, r, 2, lbl, bold=False, bg=GREY, align="right")
                            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                            set_cell(ws, r, 4, val, bg=WHITE, wrap=True)
                            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
                            r += 1

                        ws.row_dimensions[r].height = 18
                        tot(ws, r, 2, f"System {si+1} Sell Price", merge_to=7)
                        tot(ws, r, 8, "")
                        tot(ws, r, 9, round(cogs, 2), "$#,##0.00")
                        tot(ws, r, 10, round(sell_s, 2), "$#,##0.00")
                        r += 1

                    summary_sells[disc] = total_sell_ks
                    ws.row_dimensions[r].height = 20
                    tot(ws, r, 1, "")
                    tot(ws, r, 2, "TOTAL – KITCHEN SUPPRESSION", merge_to=9)
                    tot(ws, r, 10, round(total_sell_ks, 2), "$#,##0.00")
                    r += 1
                    continue

                # ── Standard labour-hours disciplines ─────────────────────
                with db.get_conn() as conn:
                    items = [dict(x) for x in conn.execute(
                        "SELECT * FROM pma_line_items "
                        "WHERE quote_id=? AND discipline=? ORDER BY sort_order",
                        (q["id"], disc)
                    ).fetchall()]
                if not items:
                    continue

                r += 1
                section_row(ws, r, f"  {disc.upper()}")
                r += 1

                for col, txt, w in [
                    (1, "#", 4), (2, "Item Description", 34), (3, "Techs Req'd", 9),
                    (4, "Quantity", 9), (5, "Unit Hours", 10), (6, "Adj. Hours", 10),
                    (7, "Frequency", 10), (8, "Ext. Hours", 10),
                    (9, "Labour Cost", 13), (10, "Sell Price", 13),
                ]:
                    hdr(ws, r, col, txt, w)
                r += 1

                total_man = 0.0
                for idx, item in enumerate(items):
                    techs = item.get("techs", 1) or 1
                    qty   = item.get("qty", 0) or 0
                    uh    = item.get("unit_hrs", 0) or 0
                    freq  = item.get("frequency", 1) or 1
                    adj_h = site_mult * uh
                    ext_h = adj_h * freq * qty * techs
                    total_man += ext_h

                    bg = GREY if idx % 2 == 0 else WHITE
                    ws.row_dimensions[r].height = 16
                    set_cell(ws, r, 1, idx+1, align="center", bg=bg)
                    set_cell(ws, r, 2, item["item_name"], bg=bg, wrap=True)
                    set_cell(ws, r, 3, techs, align="center", bg=bg)
                    inp(ws, r, 4, qty if qty else "")
                    set_cell(ws, r, 5, uh, bg=bg, fmt="0.0000", align="center")
                    set_cell(ws, r, 6, round(adj_h, 4), bg=GREY, fmt="0.0000", align="center")
                    set_cell(ws, r, 7, freq, bg=bg, fmt="0.####", align="center")
                    set_cell(ws, r, 8, round(ext_h, 2) if ext_h else "", bg=GREY,
                             fmt="0.00", align="center")
                    r += 1

                rate   = disc_rate_map[disc]
                margin = disc_margin_map[disc]
                cost   = total_man * rate
                sell   = cost / (1 - margin) if margin < 1 else cost
                summary_sells[disc] = sell

                ws.row_dimensions[r].height = 18
                tot(ws, r, 1, "")
                tot(ws, r, 2, f"TOTAL – {disc}", merge_to=7)
                tot(ws, r, 8, round(total_man, 2), "0.00")
                tot(ws, r, 9, round(cost, 2), "$#,##0.00")
                tot(ws, r, 10, round(sell, 2), "$#,##0.00")
                r += 1

            # ── Pricing Summary (full export only) ───────────────────────
            if discipline is None:
                r += 1
                section_row(ws, r, "  ANNUAL PMA PRICING SUMMARY — UPTICK ENTRY")
                r += 1

                for col, txt, w in [
                    (2, "Discipline", 28), (9, "Calc. Sell Price", 15),
                    (10, "Uptick — Cost (enter)", 18),
                    (11, "Uptick — Sell (enter)", 18),
                ]:
                    hdr(ws, r, col, txt, w)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
                r += 1

                total_annual = 0.0
                for disc, sell in summary_sells.items():
                    total_annual += sell
                    ws.row_dimensions[r].height = 18
                    set_cell(ws, r, 2, disc, bold=True)
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
                    set_cell(ws, r, 9, round(sell, 2), fmt="$#,##0.00", bg=LGREEN,
                             bold=True, align="center")
                    inp(ws, r, 10, "")
                    inp(ws, r, 11, round(sell, 2), "$#,##0.00")
                    r += 1

                term  = q.get("term_years", 5)
                escal = q.get("escalation", 0.0)
                if escal > 0:
                    contract_val = total_annual * ((1 + escal) ** term - 1) / escal
                else:
                    contract_val = total_annual * term
                escal_note = f"  ({escal*100:.1f}%/yr escalation)" if escal > 0 else ""
                ws.row_dimensions[r].height = 20
                tot(ws, r, 2, "TOTAL ANNUAL PMA", merge_to=8)
                tot(ws, r, 9, round(total_annual, 2), "$#,##0.00")
                tot(ws, r, 10, "")
                tot(ws, r, 11, "")
                r += 1
                ws.row_dimensions[r].height = 20
                tot(ws, r, 2, f"CONTRACT VALUE ({term} years{escal_note})", merge_to=8)
                tot(ws, r, 9, round(contract_val, 2), "$#,##0.00")
                tot(ws, r, 10, "")
                tot(ws, r, 11, "")

            ws.freeze_panes = "B3"

        # ═══════════════════════════════════════════════════════════════════
        # SHEET(S): INSTALL ESTIMATES  (full export only)
        # ═══════════════════════════════════════════════════════════════════
        if discipline is not None:
            wb.save(output_path)
            return True, output_path

        with db.get_conn() as conn:
            ests = [dict(x) for x in conn.execute(
                "SELECT * FROM install_estimates WHERE project_id=? ORDER BY id",
                (project_id,)
            ).fetchall()]

        diff_lu = {"Regular": "lu_reg", "Difficult": "lu_diff", "Hard": "lu_hard"}

        for est in ests:
            disc     = est.get("discipline", "Install")
            diff     = est.get("difficulty", "Regular")
            rate     = est.get("labour_rate", 90.0)
            mat_mu   = est.get("material_markup", 0.25)
            lab_mg   = est.get("labour_margin", 0.35)
            factor   = est.get("labour_factor", 1.0)
            prog_s   = est.get("prog_sell", 0.0)
            lu_key   = diff_lu.get(diff, "lu_reg")

            safe = disc[:28].replace("/", "-").strip()
            ws = wb.create_sheet(f"{safe} Install")
            ws.sheet_view.showGridLines = False
            ws.column_dimensions["A"].width = 3

            title_row(ws, 1,
                f"DEFENSE FIRE PROTECTION  |  {disc.upper()} INSTALLATION  |  {proj_name}")
            r = 2

            settings = [
                ("Discipline:", disc), ("Difficulty:", diff),
                ("Burdened Labour Rate:", f"${rate:.0f}/hr"),
                ("Material Markup:", f"{mat_mu*100:.0f}%"),
                ("Labour Margin:", f"{lab_mg*100:.0f}%"),
                ("Labour Factor:", f"{factor:.2f}"),
            ]
            for lbl_text, val in settings:
                ws.row_dimensions[r].height = 17
                set_cell(ws, r, 2, lbl_text, bold=True, bg=GREY, align="right")
                c = ws.cell(row=r, column=3, value=val)
                c.font = Font(name="Arial", color="0070C0")
                c.fill = fill(LBLUE); c.alignment = al("left"); c.border = bdr()
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
                r += 1

            r += 1
            for col, txt, w in [
                (1, "#", 4), (2, "Category", 16), (3, "Description", 34),
                (4, "Qty", 8), (5, "Unit", 6), (6, "Unit Cost", 11),
                (7, "Ext. Material", 13), (8, f"LU ({diff})", 9),
                (9, "Ext. Labour Hrs", 13), (10, "Labour Cost", 13),
                (11, "Uptick — Cost", 14), (12, "Uptick — Sell", 14),
            ]:
                hdr(ws, r, col, txt, w)
            r += 1

            with db.get_conn() as conn:
                line_items = [dict(x) for x in conn.execute(
                    "SELECT * FROM install_line_items "
                    "WHERE estimate_id=? ORDER BY sort_order",
                    (est["id"],)
                ).fetchall()]

            total_mat = 0.0
            total_hrs = 0.0
            prev_cat  = None

            for idx, item in enumerate(line_items):
                qty   = item.get("qty", 0) or 0
                uc    = item.get("unit_cost", 0) or 0
                lu    = (item.get(lu_key, 0) or 0) * factor
                ext_m = qty * uc
                ext_h = qty * lu
                lab_c = ext_h * rate
                total_mat += ext_m
                total_hrs += ext_h

                cat = item.get("category", "")
                if cat and cat != prev_cat:
                    ws.row_dimensions[r].height = 15
                    c = ws.cell(row=r, column=2, value=f"  {cat}")
                    c.font = ft(bold=True, color=WHITE, size=9)
                    c.fill = fill("5D6D7E")
                    c.alignment = al("left"); c.border = bdr()
                    ws.merge_cells(start_row=r, start_column=2,
                                   end_row=r, end_column=12)
                    r += 1
                    prev_cat = cat

                row_bg = LGREEN if qty and qty > 0 else WHITE
                ws.row_dimensions[r].height = 16
                set_cell(ws, r, 1, idx+1, align="center")
                set_cell(ws, r, 2, cat)
                set_cell(ws, r, 3, item["description"])
                inp(ws, r, 4, qty if qty else "")
                set_cell(ws, r, 5, item.get("unit", "E"), align="center")
                inp(ws, r, 6, uc if uc else "", "$#,##0.00")
                set_cell(ws, r, 7, round(ext_m, 2) if ext_m else "",
                         bg=GREY, fmt="$#,##0.00", align="center")
                set_cell(ws, r, 8, round(lu, 3), bg=GREY, fmt="0.000", align="center")
                set_cell(ws, r, 9, round(ext_h, 2) if ext_h else "",
                         bg=row_bg, fmt="0.00", align="center")
                set_cell(ws, r, 10, round(lab_c, 2) if lab_c else "",
                         bg=GREY, fmt="$#,##0.00", align="center")
                inp(ws, r, 11, "")
                # Pre-fill sell as a starting point
                inp(ws, r, 12,
                    round(lab_c / (1 - lab_mg), 2) if lab_c and lab_mg < 1 else "",
                    "$#,##0.00")
                r += 1

            # Cost summary
            r += 1
            mat_sell  = total_mat * (1 + mat_mu)
            lab_cost  = total_hrs * rate
            lab_sell  = lab_cost / (1 - lab_mg) if lab_mg < 1 else lab_cost
            total_sell = mat_sell + lab_sell + prog_s
            direct    = total_mat + lab_cost
            margin_pct = (total_sell - direct) / total_sell if total_sell else 0

            section_row(ws, r, "  COST SUMMARY")
            r += 1

            summary_rows = [
                ("Total Material Cost",     total_mat,  "$#,##0.00"),
                ("Material Markup",         mat_mu,     "0%"),
                ("Material Sell",           mat_sell,   "$#,##0.00"),
                ("", None, None),
                ("Total Labour Hours",      total_hrs,  "0.00"),
                ("Burdened Labour Rate",    rate,       "$#,##0.00"),
                ("Labour Cost",             lab_cost,   "$#,##0.00"),
                ("Labour Margin",           lab_mg,     "0%"),
                ("Labour Sell",             lab_sell,   "$#,##0.00"),
                ("", None, None),
                ("Programming / V.I. Sell", prog_s,     "$#,##0.00"),
                ("", None, None),
                ("Est. Gross Margin",       margin_pct, "0.0%"),
            ]
            ws.column_dimensions[get_column_letter(9)].width = 24
            ws.column_dimensions[get_column_letter(10)].width = 14
            for lbl_text, val, fmt in summary_rows:
                is_blank = not lbl_text
                ws.row_dimensions[r].height = 17 if not is_blank else 8
                if not is_blank:
                    set_cell(ws, r, 9, lbl_text, bold=False, bg=GREY, align="right")
                    if val is not None:
                        set_cell(ws, r, 10, val if not isinstance(val, float)
                                 else round(val, 4),
                                 bg=WHITE, align="center", fmt=fmt)
                r += 1

            ws.row_dimensions[r].height = 20
            tot(ws, r, 9, "TOTAL SELL PRICE")
            tot(ws, r, 10, round(total_sell, 2), "$#,##0.00")
            r += 1

            ws.row_dimensions[r].height = 20
            set_cell(ws, r, 9, "UPTICK — Enter Cost:", bold=True, bg=GREY, align="right")
            inp(ws, r, 10, "")
            r += 1
            ws.row_dimensions[r].height = 20
            set_cell(ws, r, 9, "UPTICK — Enter Sell:", bold=True, bg=GREY, align="right")
            inp(ws, r, 10, round(total_sell, 2), "$#,##0.00")

            ws.freeze_panes = "B9"

        # ═══════════════════════════════════════════════════════════════════
        # SHEET: PROGRAMMING & V.I.
        # ═══════════════════════════════════════════════════════════════════
        with db.get_conn() as conn:
            crow = conn.execute(
                "SELECT * FROM programming_calcs WHERE project_id=? LIMIT 1",
                (project_id,)
            ).fetchone()

        if crow:
            cd = dict(crow)
            ws = wb.create_sheet("Programming")
            ws.sheet_view.showGridLines = False
            ws.column_dimensions["A"].width = 3
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 10

            title_row(ws, 1,
                f"DEFENSE FIRE PROTECTION  |  PROGRAMMING & V.I.  |  {proj_name}",
                ncols=4)
            r = 2

            pt  = cd.get("panel_type", "4007es")
            hrs_map = {"4007es": cd["hrs_panel_4007"],
                       "4010es": cd["hrs_panel_4010"],
                       "4100 Bay": cd["hrs_panel_4100"]}
            min_map = {"4007es": cd["min_prog_vi_4007"],
                       "4010es": cd["min_prog_vi_4010"],
                       "4100 Bay": cd["min_prog_vi_4100"]}

            num_panels = cd.get("num_panels", 1)
            dev_pv     = cd.get("devices_prog_vi", 0)
            dev_p      = cd.get("devices_prog", 0)
            dev_v      = cd.get("devices_vi", 0)
            rate_prog  = cd.get("rate_prog", 140)
            rate_vi    = cd.get("rate_vi", 110)
            expenses   = cd.get("expenses", 0)
            margin     = cd.get("margin", 0.40)

            panel_hrs  = num_panels * hrs_map.get(pt, 3)
            device_hrs = (dev_pv * min_map.get(pt, 4) +
                          dev_p  * min_map.get(pt, 4) +
                          dev_v  * 6) / 60.0
            total_hrs  = panel_hrs + device_hrs
            prog_hrs   = total_hrs / 3.0
            vi_hrs     = total_hrs * 2.0 / 3.0
            days       = total_hrs / 8.0
            cost       = prog_hrs * rate_prog + vi_hrs * rate_vi + expenses
            sell       = cost / (1 - margin) if margin < 1 else cost

            data = [
                ("Panel Type:",             pt,                       ""),
                ("Number of Panels:",       num_panels,               ""),
                ("Job Type:",               cd.get("job_type", ""),   ""),
                ("Devices – Prog & V.I.:",  dev_pv,                   ""),
                ("Devices – Prog Only:",    dev_p,                    ""),
                ("Devices – V.I. Only:",    dev_v,                    ""),
                ("", None, ""),
                ("Panel Hours:",            round(panel_hrs, 1),      "0.0"),
                ("Device Hours:",           round(device_hrs, 1),     "0.0"),
                ("Total Hours:",            round(total_hrs, 1),      "0.0"),
                ("  → Programming (1/3):",  round(prog_hrs, 1),       "0.0"),
                ("  → V.I. (2/3):",         round(vi_hrs, 1),         "0.0"),
                ("Calendar Days (8hr/day):", round(days, 1),          "0.0"),
                ("", None, ""),
                ("Prog Rate (burdened):",   f"${rate_prog:.0f}/hr",   ""),
                ("V.I. Rate (burdened):",   f"${rate_vi:.0f}/hr",     ""),
                ("Expenses:",               expenses,                  "$#,##0.00"),
                ("Margin:",                 margin,                    "0%"),
                ("Labour Cost:",            round(cost, 2),            "$#,##0.00"),
                ("", None, ""),
                ("SELL PRICE:",             round(sell, 2),            "$#,##0.00"),
                ("", None, ""),
                ("UPTICK — Enter Cost:",    "",                        "$#,##0.00"),
                ("UPTICK — Enter Sell:",    round(sell, 2),            "$#,##0.00"),
            ]

            for lbl_text, val, fmt in data:
                is_blank  = not lbl_text
                is_sell   = lbl_text.startswith("SELL")
                is_uptick = lbl_text.startswith("UPTICK")
                ws.row_dimensions[r].height = 8 if is_blank else 18

                if not is_blank:
                    set_cell(ws, r, 2, lbl_text, bold=is_sell or is_uptick,
                             bg=DGREY if is_sell else GREY, align="right")
                    if val is not None:
                        if is_uptick:
                            inp(ws, r, 3, val, fmt if fmt else None)
                        elif is_sell:
                            set_cell(ws, r, 3, val, bold=True, bg=LGREEN,
                                     align="center", fmt=fmt if fmt else None)
                        else:
                            set_cell(ws, r, 3, val, bg=WHITE, align="center",
                                     fmt=fmt if fmt else None)
                r += 1

        wb.save(output_path)
        return True, output_path

    except Exception as e:
        import traceback
        return False, traceback.format_exc()
