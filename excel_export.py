from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

COLS = 7   # A-G  (#, Category, Product, Code, Qty, Unit Cost, Total)


def export_takeoff(project_name, items, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Takeoff"

    RED      = "FF7002"   # DFP brand orange
    DARK     = "232728"   # DFP brand charcoal
    LIGHT    = "ECF0F1"
    WHITE    = "FFFFFF"
    SUBTOTAL = "D5E8D4"
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def merge_row(row_num, value, fg, font_size=12, bold=True, indent=0):
        ws.merge_cells(f"A{row_num}:G{row_num}")
        c = ws[f"A{row_num}"]
        c.value = value
        c.font = Font(bold=bold, size=font_size, color=WHITE)
        c.fill = hfill(fg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)

    # Title rows  both span A:G (7 cols)
    merge_row(1, "DEFENSE FIRE PROTECTION  |  DFP TakeoffPro", RED, font_size=14)
    ws.row_dimensions[1].height = 28
    merge_row(2, f"Project:  {project_name}", DARK, font_size=11, indent=1)
    ws.row_dimensions[2].height = 20

    # Column headers
    headers    = ["#", "Category", "Product Name", "Code", "Qty", "Unit Cost", "Total"]
    col_widths = [5,    18,         36,              14,     8,     12,          14    ]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(bold=True, color=WHITE, size=10)
        c.fill      = hfill(DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 18

    # Data
    row            = 4
    current_cat    = None
    subtotal_start = 4
    grand_total    = 0.0
    item_num       = 1

    for item in items:
        cat = (item["category"] or "General")

        if cat != current_cat:
            if current_cat is not None:
                _subtotal(ws, row, subtotal_start, row - 1, border, SUBTOTAL)
                row += 1
            # Category header
            ws.merge_cells(f"A{row}:G{row}")
            ch = ws[f"A{row}"]
            ch.value     = f"  {cat.upper()}"
            ch.font      = Font(bold=True, size=10, color=WHITE)
            ch.fill      = hfill("5D6D7E")
            ch.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 16
            row         += 1
            subtotal_start = row
            current_cat = cat

        qty   = item["count"] or 0
        cost  = item["unit_cost"] or 0.0
        total = qty * cost
        grand_total += total

        fill = PatternFill("solid", fgColor=LIGHT if item_num % 2 == 0 else WHITE)
        values = [item_num, cat, item["name"], item["code"] or "", qty, cost, total]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill   = fill
            c.border = border
            c.alignment = Alignment(
                horizontal="center" if col in (1, 5) else
                "right"            if col in (6, 7) else "left",
                vertical="center",
            )
            if col in (6, 7):
                c.number_format = '"$"#,##0.00'
        ws.row_dimensions[row].height = 15
        row      += 1
        item_num += 1

    if current_cat is not None:
        _subtotal(ws, row, subtotal_start, row - 1, border, SUBTOTAL)
        row += 1

    # Grand total
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    gt = ws[f"A{row}"]
    gt.value     = "GRAND TOTAL"
    gt.font      = Font(bold=True, size=12, color=WHITE)
    gt.fill      = hfill(RED)
    gt.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    for col_ltr, val, fmt in [("F", sum(i["count"] or 0 for i in items), None),
                               ("G", grand_total, '"$"#,##0.00')]:
        c = ws[f"{col_ltr}{row}"]
        c.value = val
        c.font  = Font(bold=True, size=12 if col_ltr == "G" else 10, color=WHITE)
        c.fill  = hfill(RED)
        c.alignment = Alignment(horizontal="center" if col_ltr == "F" else "right")
        if fmt: c.number_format = fmt

    ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A4"
    wb.save(output_path)
    return output_path


def _subtotal(ws, row, start_row, end_row, border, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    ws.merge_cells(f"A{row}:E{row}")
    for col_ltr, val, fmt, align in [
        ("A", "Subtotal",                    None,            "right"),
        ("F", f"=SUM(F{start_row}:F{end_row})", None,         "center"),
        ("G", f"=SUM(G{start_row}:G{end_row})", '"$"#,##0.00', "right"),
    ]:
        c = ws[f"{col_ltr}{row}"]
        c.value     = val
        c.font      = Font(bold=True, size=10)
        c.fill      = fill
        c.alignment = Alignment(horizontal=align, indent=(1 if col_ltr == "A" else 0))
        c.border    = border
        if fmt: c.number_format = fmt
    ws.row_dimensions[row].height = 15
